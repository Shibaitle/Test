import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import threading

import requests
import json
import msal
import tempfile
import webbrowser
from urllib.parse import urlparse, quote

class ExcelComparisonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Compare and Replace Tool")
        self.root.geometry("900x700")  # Slightly larger default size
        
        # Set application style theme
        style = ttk.Style()
        style.theme_use('clam')  # Use a more modern theme
        
        # Configure custom styles
        style.configure("TButton", padding=6, relief="flat", background="#4a7ba7")
        style.configure("Accent.TButton", background="#3d6c96", foreground="white", font=('Segoe UI', 10, 'bold'))
        style.configure("TLabelframe", padding=10)
        style.configure("TLabelframe.Label", font=('Segoe UI', 10, 'bold'))
        style.configure("Header.TLabel", font=('Segoe UI', 12, 'bold'))
        style.configure("StepNumber.TLabel", font=('Segoe UI', 14, 'bold'), foreground="#3d6c96")
        
        # Create a header with application logo/title
        header_frame = ttk.Frame(root, padding="10")
        header_frame.pack(fill=tk.X)
        
        ttk.Label(header_frame, text="Excel Compare and Replace Tool", 
                font=('Segoe UI', 16, 'bold'), foreground="#3d6c96").pack(side=tk.LEFT, pady=10)
        
        # Create a wizard-like interface with step indicators
        self.step_frame = ttk.Frame(root, padding="5")
        self.step_frame.pack(fill=tk.X, padx=10)
        
        # Step indicators
        self.step1_label = ttk.Label(self.step_frame, text="1. Select Files", style="StepNumber.TLabel")
        self.step1_label.pack(side=tk.LEFT, padx=15)
        ttk.Separator(self.step_frame, orient=tk.HORIZONTAL).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.step2_label = ttk.Label(self.step_frame, text="2. Select Sheets", foreground="gray")
        self.step2_label.pack(side=tk.LEFT, padx=15)
        ttk.Separator(self.step_frame, orient=tk.HORIZONTAL).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.step3_label = ttk.Label(self.step_frame, text="3. Configure Criteria", foreground="gray")
        self.step3_label.pack(side=tk.LEFT, padx=15)
        ttk.Separator(self.step_frame, orient=tk.HORIZONTAL).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.step4_label = ttk.Label(self.step_frame, text="4. Update", foreground="gray")
        self.step4_label.pack(side=tk.LEFT, padx=15)

        # Variables to store file paths and sheet names
        self.old_file_path = tk.StringVar()
        self.new_file_path = tk.StringVar()
        self.selected_sheet = tk.StringVar()
        self.team_column = tk.StringVar()
        self.app_name_column = tk.StringVar()
        self.category_column = tk.StringVar()
        
        # Add header row configuration
        self.header_row = tk.IntVar(value=4)  # Set default to row 4
        
        # Variables for filter criteria
        self.team_filters = []
        self.app_name_filters = []
        self.category_filters = []

        # Initialize with default 1 filter entry for each type
        for _ in range(1):
            self.team_filters.append(tk.StringVar())
            self.app_name_filters.append(tk.StringVar())
            self.category_filters.append(tk.StringVar())
        
        # Add variables to track formula relationships
        self.formula_relationships = {}
        
        # Add checkbox to enable formula awareness
        self.formula_aware = tk.BooleanVar(value=True)
        
        # Create a main scrollable canvas
        self.main_canvas = tk.Canvas(root)
        self.main_scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.main_canvas.yview)
        
        # Configure the canvas
        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)
        self.main_canvas.bind('<Configure>', lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        
        # Create a frame inside the canvas to hold all content
        self.scrollable_frame = ttk.Frame(self.main_canvas)
        
        # Add mouse wheel scrolling to the canvas
        def _on_mousewheel(event):
            self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        self.main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Place the scrollable frame into the canvas
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Pack the scrollbar and canvas
        self.main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Create the main frame inside the scrollable area
        main_frame = ttk.Frame(self.scrollable_frame, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create and organize widgets
        self._create_file_selection_widgets(main_frame)
        self._create_sheet_selection_widgets(main_frame)
        self._create_criteria_widgets(main_frame)
        self._create_filter_widgets(main_frame)
        self._create_action_buttons(main_frame)
        
        # Progress bar and status label
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Ready")
        
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(status_frame, text="Status:").pack(side=tk.LEFT)
        ttk.Label(status_frame, textvariable=self.status_var).pack(side=tk.LEFT, padx=(5, 0))
        
        self.progress_bar = ttk.Progressbar(
            main_frame, 
            orient=tk.HORIZONTAL, 
            length=100, 
            mode='determinate',
            variable=self.progress_var
        )
        self.progress_bar.pack(fill=tk.X, pady=(10, 0))
        
        # Configure the binding for window resize to update scrollregion
        self.scrollable_frame.bind("<Configure>", self._configure_scrollregion)

        # Add this to the __init__ method after creating the canvas
        self.main_canvas.bind("<Enter>", self._bind_mousewheel)
        self.main_canvas.bind("<Leave>", self._unbind_mousewheel)

        # Add after other initializations
        # Initialize SharePoint integration
        self.sharepoint = SharePointIntegration(root)
        # Cleanup SharePoint files on exit
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        """Handle application close"""
        # Clean up temporary files
        self.sharepoint.cleanup()
        # Close the application
        self.root.destroy()

    def _configure_scrollregion(self, event):
        # Update the scrollregion to encompass the inner frame
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def _create_file_selection_widgets(self, parent):
        file_frame = ttk.LabelFrame(parent, text="Step 1: File Selection", padding="15")
        file_frame.pack(fill=tk.X, pady=(15, 10))
        
        # Old file selection with improved layout
        old_file_frame = ttk.Frame(file_frame)
        old_file_frame.pack(fill=tk.X, pady=(5, 10))
        
        ttk.Label(old_file_frame, text="Old File (to update):", width=20).pack(side=tk.LEFT, padx=5)
        entry = ttk.Entry(old_file_frame, textvariable=self.old_file_path, width=50)
        entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        button_frame = ttk.Frame(old_file_frame)
        button_frame.pack(side=tk.LEFT)
        
        browse_btn = ttk.Button(button_frame, text="Local File...", command=self._browse_old_file)
        browse_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(browse_btn, "Browse local files on your computer")
        
        sp_btn = ttk.Button(button_frame, text="SharePoint...", command=self._browse_old_file_sharepoint)
        sp_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(sp_btn, "Browse files from SharePoint")
        
        # Similar improvement for new file selection
        new_file_frame = ttk.Frame(file_frame)
        new_file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(new_file_frame, text="New File (reference):", width=20).pack(side=tk.LEFT, padx=5)
        ttk.Entry(new_file_frame, textvariable=self.new_file_path, width=50).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        button_frame = ttk.Frame(new_file_frame)
        button_frame.pack(side=tk.LEFT)
        
        browse_btn = ttk.Button(button_frame, text="Local File...", command=self._browse_new_file)
        browse_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(browse_btn, "Browse local files on your computer")
        
        sp_btn = ttk.Button(button_frame, text="SharePoint...", command=self._browse_new_file_sharepoint)
        sp_btn.pack(side=tk.LEFT, padx=2)
        self._add_tooltip(sp_btn, "Browse files from SharePoint")
        
        # Next step button with visual prominence
        next_btn = ttk.Button(file_frame, text="Next: Load Sheets ➜", command=self._load_sheets, style="Accent.TButton")
        next_btn.pack(side=tk.RIGHT, padx=5, pady=(10, 0))

    def _create_sheet_selection_widgets(self, parent):
        sheet_frame = ttk.LabelFrame(parent, text="Step 2: Sheet Selection", padding="15")
        sheet_frame.pack(fill=tk.X, pady=(5, 10))
        
        # Top controls with better layout
        control_frame = ttk.Frame(sheet_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        info_label = ttk.Label(control_frame, text="Select sheets to process:")
        info_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # Buttons with icons or clearer text
        select_all_btn = ttk.Button(control_frame, text="✓ Select All", command=self._select_all_sheets)
        select_all_btn.pack(side=tk.LEFT, padx=5)
        self._add_tooltip(select_all_btn, "Select all available sheets")
        
        deselect_all_btn = ttk.Button(control_frame, text="✗ Deselect All", command=self._deselect_all_sheets)
        deselect_all_btn.pack(side=tk.LEFT, padx=5)
        self._add_tooltip(deselect_all_btn, "Deselect all sheets")
        
        # Improved sheet selection area
        selection_frame = ttk.Frame(sheet_frame)
        selection_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Use a more visually appealing layout for checkboxes - grid layout with multiple columns
        self.sheet_checkbox_frame = ttk.Frame(selection_frame)
        self.sheet_checkbox_frame.pack(fill=tk.BOTH, expand=True, padx=5)
        
        # Next step button with visual prominence
        next_btn = ttk.Button(sheet_frame, text="Next: Configure Criteria ➜", command=self._load_columns,
                            style="Accent.TButton")
        next_btn.pack(side=tk.RIGHT, padx=5, pady=(10, 0))
        
        prev_btn = ttk.Button(sheet_frame, text="⬅ Back", command=self._go_to_file_selection)
        prev_btn.pack(side=tk.RIGHT, padx=5, pady=(10, 0))

    def _create_criteria_widgets(self, parent):
        criteria_frame = ttk.LabelFrame(parent, text="Step 3: Comparison Settings", padding="15")
        criteria_frame.pack(fill=tk.X, pady=(5, 10))
        
        # Create a tabbed interface for different comparison modes
        notebook = ttk.Notebook(criteria_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Tab for column-based comparison
        column_tab = ttk.Frame(notebook)
        notebook.add(column_tab, text="Column-Based Comparison")
        
        # Tab for row-based comparison
        row_tab = ttk.Frame(notebook)
        notebook.add(row_tab, text="Row-Based Comparison")
        
        # Configure the column tab
        self._configure_column_tab(column_tab)
        
        # Configure the row tab
        self._configure_row_tab(row_tab)
        
        # Bind tab change to update mode variable
        def on_tab_change(event):
            selected_tab = event.widget.index("current")
            self.use_row_mode.set(selected_tab == 1)  # 0 for column, 1 for row
        
        notebook.bind("<<NotebookTabChanged>>", on_tab_change)

    def _configure_column_tab(self, column_tab):
        # Add header row configuration
        ttk.Label(column_tab, text="Header Row:").grid(row=0, column=0, sticky=tk.W, pady=5)
        header_spin = ttk.Spinbox(column_tab, from_=1, to=20, textvariable=self.header_row, width=5)
        header_spin.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        # Add formula awareness checkbox
        ttk.Checkbutton(
            column_tab, 
            text="Formula-Aware Processing", 
            variable=self.formula_aware
        ).grid(row=0, column=15, sticky=tk.W, padx=5, pady=5)
        
        # Column mode widgets (default)
        self.column_frame = ttk.Frame(column_tab)
        self.column_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E)
        
        ttk.Label(self.column_frame, text="Team Column:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.team_combobox = ttk.Combobox(self.column_frame, textvariable=self.team_column, state="readonly", width=30)
        self.team_combobox.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(self.column_frame, text="App Name Column:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.app_name_combobox = ttk.Combobox(self.column_frame, textvariable=self.app_name_column, state="readonly", width=30)
        self.app_name_combobox.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(self.column_frame, text="Category Column:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.category_combobox = ttk.Combobox(self.column_frame, textvariable=self.category_column, state="readonly", width=30)
        self.category_combobox.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Button(self.column_frame, text="Load Columns", command=self._load_columns).grid(row=1, column=2, padx=5, pady=5)

    def _configure_row_tab(self, row_tab):
        # Add entries for row numbers instead of columns
        ttk.Label(row_tab, text="Team Row:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.team_row = tk.StringVar()
        ttk.Entry(row_tab, textvariable=self.team_row, width=10).grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(row_tab, text="App Name Row:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.app_name_row = tk.StringVar()
        ttk.Entry(row_tab, textvariable=self.app_name_row, width=10).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(row_tab, text="Category Row:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.category_row = tk.StringVar()
        ttk.Entry(row_tab, textvariable=self.category_row, width=10).grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(row_tab, text="Key Column:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.key_column = tk.StringVar()
        self.key_column_combobox = ttk.Combobox(row_tab, textvariable=self.key_column, state="readonly", width=30)
        self.key_column_combobox.grid(row=3, column=1, columnspan=2, sticky=tk.W, padx=5, pady=5)
    
    def _toggle_comparison_mode(self):
        if self.use_row_mode.get():
            # Switch to row-based comparison
            self.column_frame.grid_remove()
            self.row_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E)
            # Load columns for key selection
            self._load_key_columns()
        else:
            # Switch to column-based comparison
            self.row_frame.grid_remove()
            self.column_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E)
    
    def _load_key_columns(self):
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Read the Excel file to get column count
            df = pd.read_excel(old_file, sheet_name=selected_sheet, nrows=1)
            columns = list(df.columns)
            
            # Create letter-based column references (A, B, C, etc.)
            column_refs = [get_column_letter(i+1) for i in range(len(columns))]
            
            # Update key column combobox with both letter and name
            column_options = [f"{ref} - {name}" for ref, name in zip(column_refs, columns)]
            self.key_column_combobox['values'] = column_options
            
            if column_options:
                self.key_column_combobox.current(0)
            
            # Set default row values if empty
            if not self.team_row.get():
                self.team_row.set("1")  # Default to row 1
            if not self.app_name_row.get():
                self.app_name_row.set("2")  # Default to row 2
            if not self.category_row.get():
                self.category_row.set("3")  # Default to row 3
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load key columns: {str(e)}")
        
    def _create_filter_widgets(self, parent):
        filter_frame = ttk.LabelFrame(parent, text="Filter Criteria", padding="15")
        filter_frame.pack(fill=tk.X, pady=(5, 10))
        
        # Add a search bar at the top
        search_frame = ttk.Frame(filter_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Search filters:").pack(side=tk.LEFT, padx=5)
        
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        def filter_criteria_entries(*args):
            search_term = search_var.get().lower()
            # Apply the search filter to your criteria entries...
        
        search_var.trace_add("write", filter_criteria_entries)
        
        # Create scrollable frame for filters - INCREASE HEIGHT to 200 and add WIDTH of 750
        canvas = tk.Canvas(filter_frame, height=200, width=850)
        scrollbar = ttk.Scrollbar(filter_frame, orient="vertical", command=canvas.yview)
        self.scrollable_filter_frame = ttk.Frame(canvas)
        
        self.scrollable_filter_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_filter_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Make canvas expand to fill space
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Headers - make them stand out more
        ttk.Label(self.scrollable_filter_frame, text="Team Filters", font=("", 10, "bold")).grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(self.scrollable_filter_frame, text="App Name Filters", font=("", 10, "bold")).grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(self.scrollable_filter_frame, text="Category Filters", font=("", 10, "bold")).grid(row=0, column=2, padx=5, pady=5)
        
        # Add initial filter rows
        self._refresh_filter_widgets()
        
        # Add buttons row
        button_frame = ttk.Frame(filter_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(button_frame, text="+ Team Filter", command=self._add_team_filter).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="+ App Filter", command=self._add_app_filter).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="+ Category Filter", command=self._add_category_filter).pack(side=tk.LEFT, padx=5)
        
        # Add mouse wheel scrolling to this specific canvas as well
        def _on_filter_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_filter_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

    def _refresh_filter_widgets(self):
        # Clear existing widgets
        for widget in self.scrollable_filter_frame.winfo_children():
            if widget.winfo_class() != 'TLabel':  # Preserve headers
                widget.destroy()
        
        # Adjust layout to ensure equal spacing
        col_width = 30  # Increase width
        
        # Team filters
        for i, filter_var in enumerate(self.team_filters):
            row = i + 1
            entry = ttk.Entry(self.scrollable_filter_frame, textvariable=filter_var, width=col_width)
            entry.grid(row=row, column=0, padx=(5, 60), pady=5)
            ttk.Button(self.scrollable_filter_frame, text="Get Values", 
                    command=lambda idx=i: self._get_unique_values('team', idx)).grid(row=row, column=0, padx=(200, 5), pady=5)
        
        # App name filters
        for i, filter_var in enumerate(self.app_name_filters):
            row = i + 1
            entry = ttk.Entry(self.scrollable_filter_frame, textvariable=filter_var, width=col_width)
            entry.grid(row=row, column=1, padx=(5, 60), pady=5)
            ttk.Button(self.scrollable_filter_frame, text="Get Values", 
                    command=lambda idx=i: self._get_unique_values('app', idx)).grid(row=row, column=1, padx=(200, 5), pady=5)
        
        # Category filters
        for i, filter_var in enumerate(self.category_filters):
            row = i + 1
            entry = ttk.Entry(self.scrollable_filter_frame, textvariable=filter_var, width=col_width)
            entry.grid(row=row, column=2, padx=(5, 60), pady=5)
            ttk.Button(self.scrollable_filter_frame, text="Get Values", 
                    command=lambda idx=i: self._get_unique_values('category', idx)).grid(row=row, column=2, padx=(200, 5), pady=5)

    def _add_team_filter(self):
        self.team_filters.append(tk.StringVar())
        self._refresh_filter_widgets()

    def _add_app_filter(self):
        self.app_name_filters.append(tk.StringVar())
        self._refresh_filter_widgets()

    def _add_category_filter(self):
        self.category_filters.append(tk.StringVar())
        self._refresh_filter_widgets()

    def _select_all_sheets(self):
        for var in self.sheet_vars.values():
            var.set(True)
        
        # If any sheet is selected, load the columns
        if self.sheet_vars:
            self._on_sheet_selected()
    
    def _on_sheet_selected(self, *args):
        # Check if any sheet is selected
        selected = False
        for var in self.sheet_vars.values():
            if var.get():
                selected = True
                break
        
        if selected:
            self._load_columns()
    
    def _create_action_buttons(self, parent):
        action_frame = ttk.Frame(parent, padding="10")
        action_frame.pack(fill=tk.X, pady=15)
        
        # Save options with visual grouping and better layouts
        self.save_options_frame = ttk.LabelFrame(parent, text="Step 4: Save Options", padding="15")
        self.save_options_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Use clearer visual indicators for save options
        option_frame = ttk.Frame(self.save_options_frame)
        option_frame.pack(fill=tk.X, pady=5)
        
        # Option 1 with icon
        option1_frame = ttk.Frame(option_frame)
        option1_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(
            option1_frame, 
            text="Create a new updated file", 
            variable=self.save_mode, 
            value="new"
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(option1_frame, text="(Safe: Original file remains unchanged)", 
                 foreground="green").pack(side=tk.LEFT, padx=5)
        
        # Option 2 with warning
        option2_frame = ttk.Frame(option_frame)
        option2_frame.pack(fill=tk.X, pady=5)
        
        ttk.Radiobutton(
            option2_frame, 
            text="Replace original file", 
            variable=self.save_mode, 
            value="replace"
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(option2_frame, text="(Warning: Original file will be overwritten)", 
                 foreground="red").pack(side=tk.LEFT, padx=5)
        
        # Clearer post-update options
        ttk.Checkbutton(
            self.save_options_frame, 
            text="Reset form after update (clear all selections)",
            variable=self.clear_after_update
        ).pack(anchor=tk.W, padx=5, pady=5)
        
        # Action buttons with visual hierarchy
        button_frame = ttk.Frame(self.save_options_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="⬅ Back", 
                  command=self._go_to_criteria_selection).pack(side=tk.LEFT, padx=5)
        
        # Help button
        help_btn = ttk.Button(button_frame, text="?", width=3,
                             command=self._show_help)
        help_btn.pack(side=tk.RIGHT, padx=5)
        self._add_tooltip(help_btn, "Show help and usage instructions")
        
        # Exit button
        ttk.Button(button_frame, text="Exit",
                  command=self._confirm_exit).pack(side=tk.RIGHT, padx=5)
        
        # Primary action button with prominence
        update_btn = ttk.Button(
            button_frame, 
            text="Compare and Update ✓", 
            command=self._start_compare_update,
            style="Accent.TButton"
        )
        update_btn.pack(side=tk.RIGHT, padx=5)
        self._add_tooltip(update_btn, "Start the comparison and update process")
        
    def _browse_old_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Old Excel File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if file_path:
            # Close any potentially open workbook references
            import gc
            gc.collect()
            
            # Reset the formula relationships
            self.formula_relationships = {}
            
            # Clear sheet variables from previous selections
            self.sheet_vars.clear()
            
            # Set the new file path
            self.old_file_path.set(file_path)

    def _browse_old_file_sharepoint(self):
        """Browse and select old file from SharePoint"""
        sp_file = self.sharepoint.select_file(mode="open")
        if sp_file:
            self.old_file_path.set(sp_file)
            # Auto-load sheets when a file is selected
            self._load_sheets()
    
    def _browse_new_file(self):
        file_path = filedialog.askopenfilename(
            title="Select New Excel File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if file_path:
            self.new_file_path.set(file_path)

    def _browse_new_file_sharepoint(self):
        """Browse and select new file from SharePoint"""
        sp_file = self.sharepoint.select_file(mode="open")
        if sp_file:
            self.new_file_path.set(sp_file)
            
    def _load_sheets(self):
        old_file = self.old_file_path.get()
        new_file = self.new_file_path.get()
        
        if not old_file or not new_file:
            messagebox.showerror("Error", "Please select both old and new Excel files.")
            return
        
        try:
            # FIX: Add keep_vba and keep_links parameters
            old_wb = openpyxl.load_workbook(old_file, read_only=False, data_only=False, keep_vba=True, keep_links=True)
            new_wb = openpyxl.load_workbook(new_file, read_only=True, keep_vba=True, keep_links=True)
            
            old_sheets = set(old_wb.sheetnames)
            new_sheets = set(new_wb.sheetnames)
            
            # Find common sheets in both files
            common_sheets = list(old_sheets.intersection(new_sheets))
            
            if not common_sheets:
                messagebox.showerror("Error", "No common sheets found between the two Excel files.")
                return
            
            # Clear existing checkboxes
            for widget in self.sheet_checkbox_frame.winfo_children():
                widget.destroy()
            
            # Clear the variables dictionary
            self.sheet_vars.clear()
            
            # Create checkbox for each common sheet
            for i, sheet in enumerate(common_sheets):
                var = tk.BooleanVar(value=False)
                self.sheet_vars[sheet] = var
                checkbox = ttk.Checkbutton(
                    self.sheet_checkbox_frame, 
                    text=sheet, 
                    variable=var,
                    command=self._on_sheet_selected
                )
                checkbox.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
                
            # Detect formula relationships if enabled
            if self.formula_aware.get():
                header_row = self.header_row.get()
                
                # Clear existing relationships
                self.formula_relationships = {}
                
                # Load first sheet to detect relationships
                if old_wb.sheetnames:
                    first_sheet = old_wb[old_wb.sheetnames[0]]
                    self.formula_relationships = self._detect_formula_relationships(first_sheet, header_row)
                    
                    # Display detected relationships
                    if self.formula_relationships:
                        rel_msg = "Detected formula columns:\n"
                        for formula_col, source_col in self.formula_relationships.items():
                            rel_msg += f"- {formula_col} references {source_col}\n"
                        messagebox.showinfo("Formula Relationships", rel_msg)
            
            old_wb.close()
            new_wb.close()
            
            messagebox.showinfo("Success", f"Found {len(common_sheets)} common sheets.")
            
            # Add this to the _load_sheets method after detecting formulas
            if self.formula_relationships:
                formula_dialog = tk.Toplevel(self.root)
                formula_dialog.title("Formula Relationships Detected")
                formula_dialog.geometry("450x300")
                formula_dialog.transient(self.root)
                formula_dialog.grab_set()
                
                ttk.Label(formula_dialog, text="The following formula relationships were detected:", 
                        font=("", 10, "bold")).pack(padx=10, pady=10)
                
                # Create a frame with scrollbar for the formulas
                frame = ttk.Frame(formula_dialog)
                frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
                
                scrollbar = ttk.Scrollbar(frame)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # Use a Text widget instead of Listbox for better formatting
                formula_text = tk.Text(frame, wrap=tk.WORD, yscrollcommand=scrollbar.set)
                formula_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.config(command=formula_text.yview)
                
                # Add information about each formula relationship
                formula_text.insert(tk.END, "These columns will be preserved during updates:\n\n")
                
                for formula_col, source_col in self.formula_relationships.items():
                    formula_text.insert(tk.END, f"• Column '{formula_col}' references '{source_col}'\n")
                
                formula_text.insert(tk.END, "\n\nWhen updating, only source columns will be modified. Formula columns will be preserved.")
                formula_text.config(state=tk.DISABLED)  # Make read-only
                
                # Button to close the dialog
                ttk.Button(formula_dialog, text="OK", command=formula_dialog.destroy).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets: {str(e)}")

    def _get_unique_values(self, field_type, index=0):
        # Find the first selected sheet from checkboxes
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select a file and sheet first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Get formula relationships map if enabled
            formula_map = self.formula_relationships if self.formula_aware.get() else {}
            
            # Determine which column to get unique values from and its source
            if field_type == 'team':
                column = self.team_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select Team Column first.")
                    return
                source_column = formula_map.get(column, column)  # Get source column if it's a formula
                target_var = self.team_filters[index]
            elif field_type == 'app':
                column = self.app_name_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select App Name Column first.")
                    return
                source_column = formula_map.get(column, column)
                target_var = self.app_name_filters[index]
            elif field_type == 'category':
                column = self.category_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select Category Column first.")
                    return
                source_column = formula_map.get(column, column)
                target_var = self.category_filters[index]
                
            # Use openpyxl to read both raw values AND formula results
            wb = openpyxl.load_workbook(old_file, data_only=True)  # data_only=True gets formula results
            sheet = wb[selected_sheet]
            
            # Find the column index for the source column
            source_col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row+1, column=col).value
                if cell_value == source_column:
                    source_col_idx = col
                    break
                    
            if source_col_idx is None:
                messagebox.showerror("Error", f"Could not find column {source_column} in sheet.")
                return
                
            # Collect unique non-empty values from this column
            unique_values = set()
            for row in range(header_row+2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=source_col_idx).value
                if cell_value:  # Only add non-empty values
                    unique_values.add(str(cell_value))
                    
            # Sort the unique values
            unique_values = sorted(list(unique_values))
            
            # Create a selection dialog
            value_dialog = tk.Toplevel(self.root)
            value_dialog.title(f"Select {column} Value")
            value_dialog.geometry("400x300")
            value_dialog.transient(self.root)
            value_dialog.grab_set()
            
            # Create a frame with a scrollbar
            frame = ttk.Frame(value_dialog)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Create a scrollable listbox
            scrollbar = ttk.Scrollbar(frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=50, height=15)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar.config(command=listbox.yview)
            
            # Add search functionality
            search_frame = ttk.Frame(value_dialog)
            search_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
            
            ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
            search_var = tk.StringVar()
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
            search_entry.pack(side=tk.LEFT, padx=5)
            
            def filter_values(*args):
                search_term = search_var.get().lower()
                listbox.delete(0, tk.END)
                for value in unique_values:
                    if search_term in str(value).lower():
                        listbox.insert(tk.END, str(value))
                        
            search_var.trace_add("write", filter_values)
            
            # Populate the listbox initially
            for value in unique_values:
                listbox.insert(tk.END, str(value))
            
            # Function to set the selected value
            def set_selected_value():
                try:
                    selected_indices = listbox.curselection()
                    if selected_indices:
                        selected_value = listbox.get(selected_indices[0])
                        target_var.set(selected_value)
                    value_dialog.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to set value: {str(e)}")
                    value_dialog.destroy()
            
            # Add buttons
            button_frame = ttk.Frame(value_dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Select", command=set_selected_value).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=value_dialog.destroy).pack(side=tk.RIGHT, padx=5)
            
            # Focus on search box
            search_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get unique values: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _update_status(self, message, progress=None):
        self.status_var.set(message)
        if progress is not None:
            self.progress_var.set(progress)
        self.root.update_idletasks()
    
    def _start_compare_update(self):
        # Validate inputs
        if not self._validate_inputs():
            return
        
        # Start comparison and update in a separate thread to avoid freezing the UI
        threading.Thread(target=self._compare_and_update, daemon=True).start()
    
    def _validate_inputs(self):
        # Check if all required files are selected
        if not self.old_file_path.get():
            messagebox.showerror("Error", "Please select the old Excel file.")
            return False
        
        if not self.new_file_path.get():
            messagebox.showerror("Error", "Please select the new Excel file.")
            return False
        
        # Check if at least one sheet is selected
        selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
        if not selected_sheets:
            messagebox.showerror("Error", "Please select at least one sheet to compare.")
            return False
        
        if self.use_row_mode.get():
            # Validate row-based inputs
            try:
                int(self.team_row.get())
                int(self.app_name_row.get())
                int(self.category_row.get())
            except ValueError:
                messagebox.showerror("Error", "Row numbers must be integers.")
                return False
                
            if not self.key_column.get():
                messagebox.showerror("Error", "Please select a key column.")
                return False
        else:
            # Validate column-based inputs
            if not self.team_column.get() or not self.app_name_column.get() or not self.category_column.get():
                messagebox.showerror("Error", "Please select all three criteria columns.")
                return False
        
        return True
    
    def _compare_and_update(self):
        try:
            self._update_status("Starting comparison...", 0)
            
            # Get input values
            old_file = self.old_file_path.get()
            new_file = self.new_file_path.get()
            header_row = self.header_row.get()
        
            # Get selected sheets from checkbox variables
            selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
            if not selected_sheets:
                messagebox.showerror("Error", "No sheets selected for processing.")
                self._update_status("Ready", 0)
                return
            
            # Determine if we're using row or column mode
            is_row_mode = self.use_row_mode.get()
            
            # Get filter values - Extract the actual strings from StringVar objects
            team_filters = [f.get().strip() for f in self.team_filters if f.get().strip()]
            app_filters = [f.get().strip() for f in self.app_name_filters if f.get().strip()]
            category_filters = [f.get().strip() for f in self.category_filters if f.get().strip()]
            
            # Force close any previously open workbooks
            self._ensure_workbooks_closed()
            
            # Load workbooks with appropriate data_only settings
            self._update_status("Loading workbooks...", 10)
            old_wb_raw = openpyxl.load_workbook(old_file, data_only=False, keep_vba=True, keep_links=True)
            old_wb_eval = openpyxl.load_workbook(old_file, data_only=True, keep_vba=True, keep_links=True)
            new_wb = openpyxl.load_workbook(new_file, data_only=True, keep_vba=True, keep_links=True)
            
            # Track protected sheets and skipped cells
            protected_sheets = []
            skipped_protected_cells = 0
            
            # Get formula relationships map if enabled
            formula_map = self.formula_relationships if self.formula_aware.get() else {}
            
            total_updates = 0
            sheets_processed = 0
            
            for sheet_name in selected_sheets:
                self._update_status(f"Processing sheet: {sheet_name}...", 
                                20 + (sheets_processed / len(selected_sheets) * 60))
                
                # Get sheet objects - both raw and evaluated versions
                old_sheet_raw = old_wb_raw[sheet_name]  # Contains formulas
                old_sheet_eval = old_wb_eval[sheet_name]  # Contains formula results
                new_sheet = new_wb[sheet_name]
                
                # FIX 2: Check if sheet is protected and track it
                if old_sheet_raw.protection.sheet:
                    protected_sheets.append(sheet_name)
                    print(f"Warning: Sheet '{sheet_name}' is protected. Some cells may not be updated.")
                
                # COLUMN-BASED MODE
                if not self.use_row_mode.get():
                    # Column-based comparison
                    team_col = self.team_column.get()
                    app_name_col = self.app_name_column.get()
                    category_col = self.category_column.get()
                    
                    # Create mappings and process keys as before...
                    headers = {}
                    col_to_name = {}
                    
                    for col in range(1, old_sheet_raw.max_column + 1):
                        cell_value = old_sheet_raw.cell(row=header_row, column=col).value
                        if cell_value:
                            headers[cell_value] = col
                            col_to_name[col] = cell_value
                    
                    # Process column indexes, normalize values, etc. (existing code)
                    if team_col in headers:
                        team_idx = headers[team_col]
                    else:
                        messagebox.showerror("Error", f"Team column '{team_col}' not found in headers.")
                        continue
                        
                    if app_name_col in headers:
                        app_idx = headers[app_name_col]
                    else:
                        messagebox.showerror("Error", f"App name column '{app_name_col}' not found in headers.")
                        continue
                        
                    if category_col in headers:
                        cat_idx = headers[category_col]
                    else:
                        messagebox.showerror("Error", f"Category column '{category_col}' not found in headers.")
                        continue
                    
                    # Function to normalize cell value for consistent key generation
                    def normalize_value(value):
                        if value is None:
                            return ""
                        # Convert to string and strip
                        return str(value).strip()
                    
                    # Create keys and maps
                    old_keys = {}
                    new_keys = {}
                    
                    # Process rows and create keys
                    for row in range(header_row + 1, old_sheet_eval.max_row + 1):
                        team_value = normalize_value(old_sheet_eval.cell(row=row, column=team_idx).value)
                        app_value = normalize_value(old_sheet_eval.cell(row=row, column=app_idx).value)
                        cat_value = normalize_value(old_sheet_eval.cell(row=row, column=cat_idx).value)
                        key = f"{team_value}|{app_value}|{cat_value}"
                        if team_value or app_value or cat_value:
                            old_keys[key] = row
                    
                    for row in range(header_row + 1, new_sheet.max_row + 1):
                        team_value = normalize_value(new_sheet.cell(row=row, column=team_idx).value)
                        app_value = normalize_value(new_sheet.cell(row=row, column=app_idx).value)
                        cat_value = normalize_value(new_sheet.cell(row=row, column=cat_idx).value)
                        key = f"{team_value}|{app_value}|{cat_value}"
                        if team_value or app_value or cat_value:
                            new_keys[key] = row
                    
                    # Apply filters and find matching keys
                    all_common_keys = set(old_keys.keys()).intersection(set(new_keys.keys()))
                    
                    filtered_keys = set()
                    if team_filters or app_filters or category_filters:  # Use local variables, not self.team_filters
                        for key in all_common_keys:
                            parts = key.split('|')
                            if len(parts) == 3:
                                team, app, category = parts[0], parts[1], parts[2]
                                passes_team = True
                                passes_app = True
                                passes_category = True
                                
                                if team_filters:  # Use local variable instead of self.team_filters
                                    passes_team = any(team.lower() == filter_val.lower() for filter_val in team_filters)
                                
                                if app_filters and passes_team:  # Use local variable
                                    passes_app = any(app.lower() == filter_val.lower() for filter_val in app_filters)
                                
                                if category_filters and passes_team and passes_app:  # Use local variable
                                    passes_category = any(category.lower() == filter_val.lower() for filter_val in category_filters)
                                
                                if passes_team and passes_app and passes_category:
                                    filtered_keys.add(key)
                    else:
                        filtered_keys = all_common_keys
                    
                    # Identify formula cells
                    formula_cells = set()
                    for row in range(header_row + 1, old_sheet_raw.max_row + 1):
                        for col in range(1, old_sheet_raw.max_column + 1):
                            cell = old_sheet_raw.cell(row=row, column=col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                                formula_cells.add((row, col))
                    
                    formula_columns = set()
                    for formula_col, src_col in formula_map.items():
                        if formula_col in headers:
                            formula_columns.add(headers[formula_col])
                    
                    # Process updates
                    updates_made = 0
                    skipped_rows = 0
                    skipped_formula = 0
                    
                    # For each matching key, update cells
                    for key in filtered_keys:
                        old_row = old_keys[key]
                        new_row = new_keys[key]
                        cells_updated = False
                        
                        for col in range(1, min(old_sheet_raw.max_column, new_sheet.max_column) + 1):
                            # Skip header and formula cells
                            if old_row == header_row or col in formula_columns or (old_row, col) in formula_cells:
                                skipped_formula += 1
                                continue
                            
                            cell = old_sheet_raw.cell(row=old_row, column=col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                                formula_cells.add((old_row, col))
                                skipped_formula += 1
                                continue
                            
                            # Get the new value
                            new_value = new_sheet.cell(row=new_row, column=col).value
                            
                            # Handle merged cells
                            try:
                                # Check for merged cells
                                is_merged = False
                                merged_cell_addr = None
                                
                                for merged_range in old_sheet_raw.merged_cells.ranges:
                                    # Get the bounds of the merged range
                                    bounds = openpyxl.utils.cell.range_boundaries(merged_range.coord)
                                    min_col, min_row, max_col, max_row = bounds
                                    
                                    # Check if our cell is within those bounds
                                    if (min_row <= old_row <= max_row) and (min_col <= col <= max_col):
                                        is_merged = True
                                        merged_cell_addr = (min_row, min_col)
                                        break
                                
                                if is_merged and merged_cell_addr != (old_row, col):
                                    continue # Skip non-topleft merged cells
                                
                                # Only update if values are different
                                old_value = old_sheet_raw.cell(row=old_row, column=col).value
                                if old_value != new_value:
                                    # FIX 3: Handle protected sheets gracefully
                                    try:
                                        old_sheet_raw.cell(row=old_row, column=col).value = new_value
                                        cells_updated = True
                                    except AttributeError as e:
                                        if "protected" in str(e).lower():
                                            skipped_protected_cells += 1
                                            continue
                                        else:
                                            raise
                            
                            except Exception as e:
                                # Log error but continue processing
                                col_name = col_to_name.get(col, f"Column {col}")
                                print(f"Error updating row {old_row}, {col_name}: {str(e)}")
                                continue
                        
                        if cells_updated:
                            updates_made += 1
                        else:
                            skipped_rows += 1
                    
                    total_updates += updates_made

                # ROW-BASED MODE
                else:
                    team_row = int(self.team_row.get())
                    app_name_row = int(self.app_name_row.get())
                    category_row = int(self.category_row.get())
                    
                    # Get key column and key mapping
                    key_column_full = self.key_column.get()
                    key_column_letter = key_column_full.split(" - ")[0]
                    key_col_idx = openpyxl.utils.column_index_from_string(key_column_letter)
                    
                    # Create keys for all data columns in both sheets
                    def normalize_value(value):
                        if value is None:
                            return ""
                        return str(value).strip()
                    
                    old_keys = {}
                    new_keys = {}
                    
                    # Process the old sheet columns
                    for col in range(1, old_sheet_eval.max_column + 1):
                        try:
                            team_value = normalize_value(old_sheet_eval.cell(row=team_row, column=col).value)
                            app_value = normalize_value(old_sheet_eval.cell(row=app_name_row, column=col).value)
                            cat_value = normalize_value(old_sheet_eval.cell(row=category_row, column=col).value)
                            key = f"{team_value}|{app_value}|{cat_value}"
                            
                            if team_value or app_value or cat_value:
                                old_keys[key] = col
                        except Exception as e:
                            continue
                    
                    # Process the new sheet columns
                    for col in range(1, new_sheet.max_column + 1):
                        try:
                            team_value = normalize_value(new_sheet.cell(row=team_row, column=col).value)
                            app_value = normalize_value(new_sheet.cell(row=app_name_row, column=col).value)
                            cat_value = normalize_value(new_sheet.cell(row=category_row, column=col).value)
                            key = f"{team_value}|{app_value}|{cat_value}"
                            
                            if team_value or app_value or cat_value:
                                new_keys[key] = col
                        except Exception as e:
                            continue
                    
                    # Find common keys with filtering
                    common_key_values = set(old_keys.keys()).intersection(set(new_keys.keys()))
                    
                    # Apply filters if specified
                    if team_filters or app_filters or category_filters:  # FIX: Use local string variables
                        filtered_keys = set()
                        for key in common_key_values:
                            parts = key.split('|')
                            if len(parts) == 3:
                                team, app, category = parts[0], parts[1], parts[2]
                                
                                passes_team = True
                                passes_app = True
                                passes_category = True
                                
                                if team_filters:  # FIX: Use local string variables
                                    passes_team = any(team.lower() == filter_val.lower() for filter_val in team_filters)
                                
                                if app_filters and passes_team:  # FIX: Use local string variables
                                    passes_app = any(app.lower() == filter_val.lower() for filter_val in app_filters)
                                
                                if category_filters and passes_team and passes_app:  # FIX: Use local string variables
                                    passes_category = any(category.lower() == filter_val.lower() for filter_val in category_filters)
                                
                                if passes_team and passes_app and passes_category:
                                    filtered_keys.add(key)
                    else:
                        filtered_keys = common_key_values
                    
                    common_keys = filtered_keys
                    
                    # Find formula cells to avoid updating them
                    formula_cells = set()
                    for row in range(1, old_sheet_raw.max_row + 1):
                        for col in range(1, old_sheet_raw.max_column + 1):
                            cell = old_sheet_raw.cell(row=row, column=col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                                formula_cells.add((row, col))
                    
                    updates_made = 0
                    skipped_cols = 0
                    skipped_formula = 0
                    max_row = max(old_sheet_raw.max_row, new_sheet.max_row, 1000)
                    
                    # Process each key
                    for key in common_keys:
                        old_col = old_keys[key]
                        new_col = new_keys[key]
                        cells_updated = False
                        
                        # Update all rows for this matched column pair
                        for row in range(1, max_row + 1):
                            # Skip criteria rows
                            if row in (team_row, app_name_row, category_row):
                                continue
                            
                            # Skip formula cells
                            if (row, old_col) in formula_cells:
                                skipped_formula += 1
                                continue
                            
                            # Double check for formula
                            cell = old_sheet_raw.cell(row=row, column=old_col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                                formula_cells.add((row, old_col))
                                skipped_formula += 1
                                continue
                            
                            # Process if row exists in both sheets
                            if row <= new_sheet.max_row and row <= old_sheet_raw.max_row:
                                new_value = new_sheet.cell(row=row, column=new_col).value
                                
                                if new_value is not None:
                                    # Handle merged cells
                                    try:
                                        is_merged = False
                                        merged_cell_addr = None
                                        
                                        for merged_range in old_sheet_raw.merged_cells.ranges:
                                            bounds = openpyxl.utils.cell.range_boundaries(merged_range.coord)
                                            min_col, min_row, max_col, max_row_merged = bounds
                                            
                                            if (min_row <= row <= max_row_merged) and (min_col <= old_col <= max_col):
                                                is_merged = True
                                                merged_cell_addr = (min_row, min_col)
                                                break
                                        
                                        if is_merged and merged_cell_addr != (row, old_col):
                                            continue
                                        
                                        # Only update if values differ
                                        old_value = old_sheet_raw.cell(row=row, column=old_col).value
                                        if old_value != new_value:
                                            # FIX 3: Handle protected sheets gracefully
                                            try:
                                                old_sheet_raw.cell(row=row, column=old_col).value = new_value
                                                cells_updated = True
                                            except AttributeError as e:
                                                if "protected" in str(e).lower():
                                                    skipped_protected_cells += 1
                                                    continue
                                                else:
                                                    raise
                                    
                                    except Exception as e:
                                        print(f"Warning: Could not update cell at row {row}, col {old_col}: {str(e)}")
                                        continue
                        
                        if cells_updated:
                            updates_made += 1
                        else:
                            skipped_cols += 1
                    
                    total_updates += updates_made
                        
                sheets_processed += 1
            
            # Generate output filename based on selected save mode
            if self.save_mode.get() == "new":
                # Create new file with filter info
                base, ext = os.path.splitext(old_file)
                filter_info = ""
                if team_filters or app_filters or category_filters:  # FIX: Use local string variables
                    filter_parts = []
                    if team_filters:
                        filter_parts.append(f"Team-{'-'.join(team_filters)}")
                    if app_filters:
                        filter_parts.append(f"App-{'-'.join(app_filters)}")
                    if category_filters:
                        filter_parts.append(f"Cat-{'-'.join(category_filters)}")
                    filter_info = "_" + "_".join(filter_parts)
                    
                mode_info = "_row_based" if is_row_mode else ""
                output_file = f"{base}_updated{filter_info}{mode_info}{ext}"
            else:
                # Replace original file - but confirm first
                if not messagebox.askyesno("Confirm Replace", 
                    "Are you sure you want to overwrite the original file?\nThis cannot be undone.", 
                    icon="warning"):
                    self._update_status("Operation cancelled", 0)
                    return
                output_file = old_file
            
            # FIX 4: Preserve data connections when saving
            self._update_status("Saving updated workbook...", 90)
            try:
                # Preserve external links and data connections when saving
                old_wb_raw.save(output_file)
            except Exception as save_error:
                if "data connections" in str(save_error).lower():
                    # Special handling for data connection errors
                    messagebox.showwarning("Warning", 
                                         "External data connections may be affected in the saved file. "
                                         "Please verify data connections after saving.")
                    # Try again without keeping links
                    old_wb_raw._keep_links = False
                    old_wb_raw.save(output_file)
                else:
                    # Re-raise any other save errors
                    raise
            
            # Close all workbooks properly
            old_wb_raw.close()
            old_wb_eval.close()
            new_wb.close()
            
            # If replacing the original file, do cleanup
            if self.save_mode.get() == "replace":
                # Force Python's garbage collection to ensure file handles are released
                import gc
                gc.collect()
                
                # Reset the formula relationships to ensure they're redetected
                if self.formula_aware.get():
                    self.formula_relationships = {}
            
            # Check if we should upload to SharePoint
            sp_upload = False
            for temp_file in self.sharepoint.temp_files:
                if temp_file["path"] == old_file:
                    sp_upload = True
                    break
            
            if sp_upload:
                # SharePoint upload handling
                self._update_status("Uploading to SharePoint...", 95)
                if messagebox.askyesno("Upload to SharePoint", "Would you like to save the updated file to SharePoint?"):
                    output_name = os.path.basename(output_file)
                    if self.sharepoint.upload_file(output_file, output_name):
                        messagebox.showinfo("Success", f"File successfully uploaded to SharePoint as {output_name}")
                    else:
                        messagebox.warning("Upload Warning", "File was saved locally but SharePoint upload failed.")
            
            self._update_status("Complete!", 100)
            
            # Show success message with details
            success_message = f"Updated {total_updates} {'columns' if self.use_row_mode.get() else 'rows'} successfully!\nSaved to: {output_file}"
            
            # Add warning about protected sheets if any were encountered
            if protected_sheets:
                success_message += f"\n\nWARNING: {len(protected_sheets)} sheets were protected and may not have been fully updated:"
                success_message += f"\n- {', '.join(protected_sheets)}"
                success_message += f"\nSkipped {skipped_protected_cells} protected cells."
            
            # Add other message components
            sheet_message = f"\nProcessed sheets: {', '.join(selected_sheets)}"
            mode_message = "\nComparison mode: Row-based" if self.use_row_mode.get() else "\nComparison mode: Column-based"
            formula_message = "\nFormula columns preserved" if formula_map else ""
            
            filter_message = ""
            if team_filters or app_filters or category_filters:  # FIX: Use local string variables
                filter_message = "\n\nFilters applied:"
                if team_filters:
                    filter_message += f"\n- Teams: {', '.join(team_filters)}"
                if app_filters:
                    filter_message += f"\n- App Names: {', '.join(app_filters)}"
                if category_filters:
                    filter_message += f"\n- Categories: {', '.join(category_filters)}"
            
            messagebox.showinfo("Success", success_message + sheet_message + mode_message + formula_message + filter_message)
            
            # Clear file selections if option is enabled
            if self.clear_after_update.get():
                # Reset file paths
                self.old_file_path.set("")
                self.new_file_path.set("")
                
                # Clear sheet selections
                for var in self.sheet_vars.values():
                    var.set(False)
                
                # Clear sheet variables
                self.sheet_vars.clear()
                
                # Clear column selections
                self.team_column.set("")
                self.app_name_column.set("")
                self.category_column.set("")
                
                # Reset filters
                for filter_var in self.team_filters:
                    filter_var.set("")
                for filter_var in self.app_name_filters:
                    filter_var.set("")
                for filter_var in self.category_filters:
                    filter_var.set("")
            
            # Reset UI elements
            self._update_status("Ready", 0)
            
        except Exception as e:
            self._update_status("Error occurred", 0)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            import traceback
            traceback.print_exc()

    def _update_sheet(self, sheet, old_df, new_df, common_keys, team_col, app_name_col, category_col, formula_map=None):
        # Use the configured header row
        header_row = self.header_row.get()
        formula_map = formula_map or {}
        
        # Create a mapping of column names to column indices
        headers = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value:
                headers[col] = cell_value
        
        # Create a dictionary to store the column indices for faster lookup
        columns_indices = {col: headers[col] for col in new_df.columns if col in headers}
        
        # Create a set of columns that should NOT be updated (formula destination columns)
        formula_columns = set()
        for formula_col, source_col in formula_map.items():
            if formula_col in headers:
                formula_columns.add(formula_col)
        
        # Track which cells contain formulas (as backup detection)
        formula_cells = set()
        for row in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.data_type == 'f':  # 'f' indicates formula
                    formula_cells.add((row, col))
        
        updates_made = 0
        
        # For each row in the sheet (starting after the header)
        for row in range(header_row + 1, sheet.max_row + 1):
            # Get the key for this row using source columns (not formula columns)
            try:
                # Map formula columns to their source columns for key generation
                actual_team_col = formula_map.get(team_col, team_col)
                actual_app_col = formula_map.get(app_name_col, app_name_col)
                actual_cat_col = formula_map.get(category_col, category_col)
                
                if actual_team_col in headers and actual_app_col in headers and actual_cat_col in headers:
                    team_idx = headers[actual_team_col]
                    app_idx = headers[actual_app_col]
                    cat_idx = headers[actual_cat_col]
                    
                    team_value = str(sheet.cell(row=row, column=team_idx).value or "")
                    app_value = str(sheet.cell(row=row, column=app_idx).value or "")
                    cat_value = str(sheet.cell(row=row, column=cat_idx).value or "")
                    row_key = f"{team_value}|{app_value}|{cat_value}"
                    
                    # If this row has a match in the new data and passes our filters
                    if row_key in common_keys:
                        # Get the new data for this key
                        new_row_data = new_df[new_df['_key'] == row_key].iloc[0]
                        
                        # Update all applicable columns in the old sheet with values from the new data
                        for col_name, col_idx in columns_indices.items():
                            # Skip formula columns, artificial key column, and any cell with a formula
                            if (col_name != '_key' and 
                                col_name not in formula_columns and 
                                (row, col_idx) not in formula_cells):
                                
                                # Double-check this isn't a formula cell before updating
                                cell = sheet.cell(row=row, column=col_idx)
                                if cell.data_type != 'f':
                                    try:
                                        # Update with new value
                                        sheet.cell(row=row, column=col_idx).value = new_row_data[col_name]
                                    except Exception as cell_err:
                                        print(f"Error updating cell at row {row}, col {col_idx}: {str(cell_err)}")
                        
                        updates_made += 1
                else:
                    print(f"Warning: One or more selected columns not found in headers: {actual_team_col}, {actual_app_col}, {actual_cat_col}")
                    
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                continue
                    
        return updates_made

    def _load_columns(self):
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Read the Excel file skipping to the header row
            df = pd.read_excel(old_file, sheet_name=selected_sheet, header=header_row)
            columns = list(df.columns)
            
            if not columns:
                messagebox.showerror("Error", "No columns found in the selected sheet.")
                return
            
            # Update all three comboboxes with the column list
            self.team_combobox['values'] = columns
            self.app_name_combobox['values'] = columns
            self.category_combobox['values'] = columns
            
            # Set default selections if available
            for column in columns:
                if "team" in str(column).lower():
                    self.team_column.set(column)
                elif "app" in str(column).lower() or "name" in str(column).lower():
                    self.app_name_column.set(column)
                elif "category" in str(column).lower() or "test" in str(column).lower():
                    self.category_column.set(column)
            
            messagebox.showinfo("Success", f"Loaded {len(columns)} columns from sheet '{selected_sheet}'.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load columns: {str(e)}")

    def _bind_mousewheel(self, event):
        # Bind mousewheel scrolling when mouse enters the canvas
        self.main_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        # For Linux, bind Button-4 and Button-5
        self.main_canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.main_canvas.bind_all("<Button-5>", self._on_mousewheel)

    def _unbind_mousewheel(self, event):
        # Unbind mousewheel scrolling when mouse leaves the canvas
        self.main_canvas.unbind_all("<MouseWheel>")
        self.main_canvas.unbind_all("<Button-4>")
        self.main_canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        # Windows and macOS
        if event.num == 5 or event.delta < 0:
            self.main_canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.main_canvas.yview_scroll(-1, "units")

    def _detect_formula_relationships(self, sheet, header_row):
        """Detect formula relationships between columns in the sheet."""
        import re
        relationships = {}
        
        print("\n--- Starting Formula Relationship Detection ---")
        
        # Step 1: Create a mapping from column positions to header names
        headers = {}  # Column index to header name
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value:
                headers[col] = cell_value
        
        # Create reverse mapping from column letter to header name
        letter_to_header = {}
        for col_idx, header in headers.items():
            col_letter = get_column_letter(col_idx)
            letter_to_header[col_letter] = header
        
        print(f"Found {len(headers)} headers")
        print(f"Column letter mapping: {letter_to_header}")
        
        # Step 2: Analyze formula patterns in the first few rows
        formula_patterns = {}  # To track potential formula columns
        
        # Look at several rows to establish consistent patterns
        for row_num in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            print(f"Analyzing row {row_num}")
            
            for col_idx, header_name in headers.items():
                cell = sheet.cell(row=row_num, column=col_idx)
                cell_formula = None
                
                # Get formula using different methods
                if cell.data_type == 'f':
                    cell_formula = str(cell.value)
                elif hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_formula = cell.value
                
                if cell_formula:
                    # Look for simple cell references like =A5, =$A5, =A$5, or =$A$5
                    # The regex captures just the column letter part
                    match = re.search(r'=\$?([A-Za-z]+)\$?\d+', cell_formula)
                    if match:
                        ref_col_letter = match.group(1).upper()  # Ensure uppercase
                        print(f"  Column '{header_name}' contains formula: {cell_formula}")
                        print(f"  References column letter: {ref_col_letter}")
                        
                        # If we can map this reference to a header
                        if ref_col_letter in letter_to_header:
                            source_header = letter_to_header[ref_col_letter]
                            print(f"  Maps to header: '{source_header}'")
                            
                            # Add to our tracking dictionary
                            if header_name not in formula_patterns:
                                formula_patterns[header_name] = {}
                            
                            if source_header not in formula_patterns[header_name]:
                                formula_patterns[header_name][source_header] = 0
                            
                            formula_patterns[header_name][source_header] += 1
        
        # Step 3: Confirm relationships (if a column consistently references another column)
        for formula_col, references in formula_patterns.items():
            if references:
                # Find the most commonly referenced source column
                most_common = max(references.items(), key=lambda x: x[1])
                source_col, count = most_common
                
                # If we have at least 2 occurrences, establish a relationship
                if count >= 2:
                    relationships[formula_col] = source_col
                    print(f"Established relationship: '{formula_col}' references '{source_col}' ({count} occurrences)")
        
        print("\nFinal detected formula relationships:")
        for dest, source in relationships.items():
            print(f"  • '{dest}' references '{source}'")
        
        return relationships

    def _ensure_workbooks_closed(self):
        """Force closure of any open workbooks to avoid file locking issues"""
        try:
            # Force garbage collection to release file handles
            import gc
            gc.collect()
            
            # Reset openpyxl's _archive cache if accessible
            if hasattr(openpyxl, '_archive'):
                openpyxl._archive = {}
                
            # Reset zipfiles if possible
            import zipfile
            if hasattr(zipfile, '_zipfile_cache'):
                zipfile._zipfile_cache.clear()
        except:
            pass

    def _add_tooltip(self, widget, text):
        """Create a tooltip for a given widget"""
        tooltip = tk.Toplevel(widget, bg='#FFFFDD', relief='solid', borderwidth=1)
        tooltip.withdraw()
        tooltip.overrideredirect(True)
        label = ttk.Label(tooltip, text=text, background='#FFFFDD', padding=(5, 3))
        label.pack()
        
        def show_tooltip(event=None):
            x, y, _, _ = widget.bbox('insert')
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 25
            tooltip.geometry(f"+{x}+{y}")
            tooltip.deiconify()
        
        def hide_tooltip(event=None):
            tooltip.withdraw()
        
        widget.bind('<Enter>', show_tooltip)
        widget.bind('<Leave>', hide_tooltip)
    def _create_status_indicators(self, parent):
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill=tk.X, pady=(10, 5))
        
        # Status indicators with more visual feedback
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var, foreground="#3d6c96")
        self.status_label.pack(side=tk.LEFT, padx=10)
        
        # Activity indicator (spinning animation during processing)
        self.activity_frames = ["|", "/", "-", "\\"]
        self.activity_index = 0
        self.activity_var = tk.StringVar()
        self.activity_label = ttk.Label(status_frame, textvariable=self.activity_var)
        self.activity_label.pack(side=tk.LEFT)
        
        # Progress bar with better styling
        self.progress_bar = ttk.Progressbar(
            parent, 
            orient=tk.HORIZONTAL, 
            length=100, 
            mode='determinate',
            variable=self.progress_var,
            style="TProgressbar"
        )
        self.progress_bar.pack(fill=tk.X, padx=10, pady=5)
        
        # Define progress bar style
        style = ttk.Style()
        style.configure("TProgressbar", thickness=8, borderwidth=0, background="#4a7ba7")

    def _start_activity_animation(self):
        """Start the activity indicator animation"""
        self.animation_active = True
        self._update_activity_animation()
    
    def _stop_activity_animation(self):
        """Stop the activity indicator animation"""
        self.animation_active = False
        self.activity_var.set("")
    
    def _update_activity_animation(self):
        """Update the activity indicator animation frame"""
        if self.animation_active:
            self.activity_index = (self.activity_index + 1) % len(self.activity_frames)
            self.activity_var.set(self.activity_frames[self.activity_index])
            self.root.after(150, self._update_activity_animation)

    def _go_to_file_selection(self):
        """Navigate back to file selection step"""
        # Update step indicator highlighting
        self.step1_label.configure(style="StepNumber.TLabel")
        self.step2_label.configure(foreground="gray")
        self.step3_label.configure(foreground="gray")
        self.step4_label.configure(foreground="gray")
        
        # Show/hide relevant frames
        self.file_frame.pack(fill=tk.X, pady=(15, 10))
        self.sheet_frame.pack_forget()
        self.criteria_frame.pack_forget()
        self.filter_frame.pack_forget()
        self.save_options_frame.pack_forget()
        
    def _go_to_sheet_selection(self):
        """Navigate to sheet selection step"""
        self.step1_label.configure(foreground="gray")
        self.step2_label.configure(style="StepNumber.TLabel")
        self.step3_label.configure(foreground="gray")
        self.step4_label.configure(foreground="gray")
        
        # Show/hide relevant frames
        self.file_frame.pack_forget()
        self.sheet_frame.pack(fill=tk.X, pady=(15, 10))
        self.criteria_frame.pack_forget()
        self.filter_frame.pack_forget()
        self.save_options_frame.pack_forget()
    
    # Similar methods for other steps...

    def _show_help(self):
        """Display a help dialog with usage instructions"""
        help_dialog = tk.Toplevel(self.root)
        help_dialog.title("Help - Excel Compare and Replace Tool")
        help_dialog.geometry("700x500")
        help_dialog.transient(self.root)
        help_dialog.grab_set()
        
        notebook = ttk.Notebook(help_dialog)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Overview tab
        overview_frame = ttk.Frame(notebook, padding=10)
        notebook.add(overview_frame, text="Overview")
        
        ttk.Label(overview_frame, text="Using the Excel Compare and Replace Tool", 
                 font=('Segoe UI', 14, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        overview_text = """This tool helps you compare two Excel files and selectively update the old file with values from the new file.
    
    The process follows these steps:
    1. Select your source and target Excel files
    2. Choose which sheets to process
    3. Configure comparison criteria
    4. Set filtering options if needed
    5. Choose save options and run the update
    
    You can compare files using either column-based or row-based methods, depending on your data structure."""
        
        text_widget = tk.Text(overview_frame, wrap=tk.WORD, height=10, width=80)
        text_widget.pack(fill=tk.BOTH, expand=True, pady=5)
        text_widget.insert(tk.END, overview_text)
        text_widget.config(state=tk.DISABLED)
        
        # Add more tabs with specific help content
        # ...
        
        # Close button
        ttk.Button(help_dialog, text="Close", command=help_dialog.destroy).pack(pady=10)

    def _confirm_exit(self):
        """Show confirmation before exiting"""
        if messagebox.askyesno("Confirm Exit", "Are you sure you want to exit? Any unsaved progress will be lost."):
            self._on_close()
    
    def _start_compare_update(self):
        """Validate and start the comparison with better feedback"""
        # Validate inputs with specific error messages
        if not self.old_file_path.get():
            messagebox.showerror("Input Error", "Please select the old Excel file (file to update).")
            return
            
        if not self.new_file_path.get():
            messagebox.showerror("Input Error", "Please select the new Excel file (reference file).")
            return
        
        # Check if files exist and are accessible
        if not os.path.exists(self.old_file_path.get()):
            messagebox.showerror("File Error", "The selected old file cannot be found. Please check the path.")
            return
            
        if not os.path.exists(self.new_file_path.get()):
            messagebox.showerror("File Error", "The selected new file cannot be found. Please check the path.")
            return
        
        # Continue with other validations...
        
        # Show a confirmation dialog with the operation details
        message = f"Ready to compare and update with the following settings:\n\n"
        message += f"• Old file: {os.path.basename(self.old_file_path.get())}\n"
        message += f"• New file: {os.path.basename(self.new_file_path.get())}\n"
        message += f"• Mode: {'Row-based' if self.use_row_mode.get() else 'Column-based'} comparison\n"
        message += f"• Selected sheets: {len([sheet for sheet, var in self.sheet_vars.items() if var.get()])}\n"
        message += f"• Save mode: {'Create new file' if self.save_mode.get() == 'new' else 'Replace original'}\n"
        
        if messagebox.askyesno("Confirm Operation", message, icon="question"):
            # Start the operation with visual feedback
            self._start_activity_animation()
            threading.Thread(target=self._compare_and_update, daemon=True).start()

class SharePointIntegration:
    def __init__(self, parent):
        self.parent = parent
        self.access_token = None
        self.temp_files = []  # Track temporary files
        
        # SharePoint/Microsoft Graph API configuration
        self.client_id = "YOUR_CLIENT_ID"  # Register an app in Azure AD to get this
        self.tenant_id = "YOUR_TENANT_ID"  # Your Microsoft 365 tenant ID
        self.authority = f"https://login.microsoftonline.com/{self.tenant_id}"
        self.scopes = ["Sites.Read.All", "Files.Read.All", "Files.ReadWrite.All"]
        
        # For storing SharePoint site and drive information
        self.site_url = None
        self.drive_id = None
        self.selected_file_info = None
    
    def authenticate(self):
        """Authenticate the user to Microsoft Graph API"""
        try:
            # Create MSAL app for authentication
            app = msal.PublicClientApplication(
                self.client_id,
                authority=self.authority
            )
            
            # Attempt silent token acquisition first
            accounts = app.get_accounts()
            if accounts:
                # Use the first cached account
                result = app.acquire_token_silent(
                    self.scopes, 
                    account=accounts[0]
                )
            else:
                # No suitable token in cache, authenticate interactively
                # Use device code flow for better UX
                flow = app.initiate_device_flow(scopes=self.scopes)
                print(flow['message'])
                # Display message to user with the device code
                auth_dialog = tk.Toplevel(self.parent)
                auth_dialog.title("SharePoint Authentication")
                auth_dialog.geometry("500x300")
                auth_dialog.transient(self.parent)
                auth_dialog.grab_set()
                
                ttk.Label(
                    auth_dialog, 
                    text="Please authenticate to SharePoint:", 
                    font=("", 12, "bold")
                ).pack(pady=10)
                
                auth_text = tk.Text(auth_dialog, height=8, width=50)
                auth_text.pack(padx=20, pady=10)
                auth_text.insert("1.0", flow['message'])
                auth_text.config(state=tk.DISABLED)
                
                # Add button to open the verification URL
                url_button = ttk.Button(
                    auth_dialog, 
                    text="Open Authentication Page",
                    command=lambda: webbrowser.open(flow['verification_uri'])
                )
                url_button.pack(pady=10)
                
                # Continue with token acquisition
                result = app.acquire_token_by_device_flow(flow)
                auth_dialog.destroy()
            
            if "access_token" in result:
                self.access_token = result["access_token"]
                return True
            else:
                error_msg = result.get("error_description", "Unknown error during authentication")
                messagebox.showerror("Authentication Error", error_msg)
                return False
                
        except Exception as e:
            messagebox.showerror("Authentication Error", f"Failed to authenticate: {str(e)}")
            return False
    
    def select_file(self, mode="open"):
        """Open a dialog to browse and select SharePoint files"""
        if not self.access_token and not self.authenticate():
            return None
            
        # Create a SharePoint file browser dialog
        sp_dialog = tk.Toplevel(self.parent)
        sp_dialog.title("SharePoint File Browser")
        sp_dialog.geometry("800x600")
        sp_dialog.transient(self.parent)
        sp_dialog.grab_set()
        
        # Variables for site entry
        site_url_var = tk.StringVar()
        
        # Frame for site entry
        site_frame = ttk.Frame(sp_dialog)
        site_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(site_frame, text="SharePoint Site URL:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(site_frame, textvariable=site_url_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            site_frame, 
            text="Connect", 
            command=lambda: self._connect_to_site(site_url_var.get(), file_tree)
        ).pack(side=tk.LEFT, padx=5)
        
        # Frame for file browser with treeview
        browser_frame = ttk.Frame(sp_dialog)
        browser_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Path label
        path_var = tk.StringVar(value="Not connected to any site")
        ttk.Label(browser_frame, textvariable=path_var).pack(fill=tk.X, padx=5, pady=5)
        
        # Create treeview with scrollbars
        tree_frame = ttk.Frame(browser_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        tree_scroll_y = ttk.Scrollbar(tree_frame)
        tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        file_tree = ttk.Treeview(
            tree_frame,
            columns=("name", "type", "modified"),
            selectmode="browse",
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set
        )
        
        # Configure scrollbars
        tree_scroll_y.config(command=file_tree.yview)
        tree_scroll_x.config(command=file_tree.xview)
        
        # Configure columns
        file_tree.column("#0", width=50, stretch=tk.NO)  # Icon column
        file_tree.column("name", width=300, anchor=tk.W)
        file_tree.column("type", width=100, anchor=tk.W)
        file_tree.column("modified", width=150, anchor=tk.W)
        
        # Configure headers
        file_tree.heading("#0", text="")
        file_tree.heading("name", text="Name")
        file_tree.heading("type", text="Type")
        file_tree.heading("modified", text="Modified")
        
        file_tree.pack(fill=tk.BOTH, expand=True)
        
        # Handle double-click on treeview items
        file_tree.bind("<Double-1>", lambda e: self._on_tree_double_click(e, file_tree, path_var))
        
        # Buttons frame
        button_frame = ttk.Frame(sp_dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        result_var = tk.StringVar()
        
        def on_select():
            item_id = file_tree.focus()
            if item_id and file_tree.item(item_id, "values")[1] == "Excel File":
                # Get the selected file
                item_data = file_tree.item(item_id, "tags")
                if item_data and item_data[0] == "file":
                    file_info = item_data[1]
                    self.selected_file_info = file_info
                    result_var.set(file_info["name"])
                    sp_dialog.destroy()
                else:
                    messagebox.showerror("Error", "Please select an Excel file.")
            else:
                messagebox.showerror("Error", "Please select an Excel file.")
        
        ttk.Button(button_frame, text="Select", command=on_select).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=sp_dialog.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Wait for the dialog to close
        sp_dialog.wait_window()
        
        if self.selected_file_info:
            # Download the selected file to a temporary location
            return self._download_file(self.selected_file_info)
        
        return None
    
    def _connect_to_site(self, site_url, tree_widget):
        """Connect to a SharePoint site and load its document libraries"""
        try:
            # Parse and format the site URL
            parsed_url = urlparse(site_url)
            hostname = parsed_url.netloc
            
            # Extract site path
            path_parts = parsed_url.path.strip("/").split("/")
            site_name = path_parts[0] if path_parts else ""
            
            # Get site ID using Microsoft Graph API
            graph_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/sites/{site_name}"
            
            headers = {
                "Authorization": f"Bearer {self.access_token}",
                "Accept": "application/json"
            }
            
            response = requests.get(graph_url, headers=headers)
            
            if response.status_code == 200:
                site_data = response.json()
                self.site_url = site_url
                self.site_id = site_data["id"]
                
                # Clear the treeview
                for item in tree_widget.get_children():
                    tree_widget.delete(item)
                
                # Load document libraries
                self._load_document_libraries(tree_widget)
            else:
                messagebox.showerror("Error", f"Could not connect to site: {response.text}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to connect to SharePoint site: {str(e)}")
    
    def _load_document_libraries(self, tree_widget):
        """Load document libraries from the site"""
        if not self.site_id:
            return
            
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/json"
        }
        
        response = requests.get(graph_url, headers=headers)
        
        if response.status_code == 200:
            drives = response.json()["value"]
            
            for drive in drives:
                drive_name = drive["name"]
                drive_id = drive["id"]
                
                # Add the drive to the treeview
                drive_item = tree_widget.insert(
                    "", 
                    "end", 
                    text="", 
                    values=(drive_name, "Document Library", ""),
                    tags=("drive", drive_id)
                )
                
                # Add a dummy node to enable the expansion icon
                tree_widget.insert(drive_item, "end", text="", values=("Loading...", "", ""))
                
                # Bind expand event
                tree_widget.bind("<<TreeviewOpen>>", lambda e: self._on_tree_expand(e, tree_widget))
    
    def _on_tree_expand(self, event, tree_widget):
        """Handle expanding a tree node"""
        item_id = tree_widget.focus()
        if not item_id:
            return
            
        item_tags = tree_widget.item(item_id, "tags")
        if not item_tags:
            return
            
        item_type = item_tags[0]
        item_data = item_tags[1]
        
        # Only process if this node hasn't been loaded yet
        first_child = tree_widget.get_children(item_id)
        if len(first_child) == 1 and tree_widget.item(first_child[0], "values")[0] == "Loading...":
            # Remove the loading node
            tree_widget.delete(first_child[0])
            
            if item_type == "drive":
                # Load the root folder of this drive
                drive_id = item_data
                self._load_folder_items(tree_widget, item_id, drive_id, "")
            elif item_type == "folder":
                # Load the contents of this folder
                folder_info = item_data
                drive_id = folder_info["parentReference"]["driveId"]
                folder_path = folder_info["id"]
                self._load_folder_items(tree_widget, item_id, drive_id, folder_path)
    
    def _load_folder_items(self, tree_widget, parent_node, drive_id, folder_path):
        """Load items from a folder"""
        graph_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        if folder_path:
            graph_url += f"/items/{folder_path}/children"
        else:
            graph_url += "/root/children"
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/json"
        }
        
        response = requests.get(graph_url, headers=headers)
        
        if response.status_code == 200:
            items = response.json()["value"]
            
            # First add folders
            for item in items:
                if "folder" in item:
                    folder_name = item["name"]
                    modified = item.get("lastModifiedDateTime", "").split("T")[0]
                    
                    folder_node = tree_widget.insert(
                        parent_node, 
                        "end", 
                        text="📁", 
                        values=(folder_name, "Folder", modified),
                        tags=("folder", item)
                    )
                    
                    # Add dummy node
                    tree_widget.insert(folder_node, "end", text="", values=("Loading...", "", ""))
            
            # Then add Excel files
            for item in items:
                if "file" in item and item["name"].lower().endswith((".xlsx", ".xls", ".xlsm")):
                    file_name = item["name"]
                    modified = item.get("lastModifiedDateTime", "").split("T")[0]
                    
                    tree_widget.insert(
                        parent_node, 
                        "end", 
                        text="📊", 
                        values=(file_name, "Excel File", modified),
                        tags=("file", item)
                    )
    
    def _on_tree_double_click(self, event, tree_widget, path_var):
        """Handle double-click on tree item"""
        item_id = tree_widget.focus()
        if not item_id:
            return
            
        item_tags = tree_widget.item(item_id, "tags")
        if not item_tags:
            return
            
        item_type = item_tags[0]
        
        if item_type == "folder" or item_type == "drive":
            # Expand/collapse the node
            if tree_widget.item(item_id, "open"):
                tree_widget.item(item_id, open=False)
            else:
                tree_widget.item(item_id, open=True)
                
                # Update path label
                if item_type == "folder":
                    folder_info = item_tags[1]
                    path_var.set(f"Path: {folder_info.get('parentReference', {}).get('path', '')}/{folder_info.get('name', '')}")
                elif item_type == "drive":
                    drive_id = item_tags[1]
                    path_var.set(f"Document Library: {tree_widget.item(item_id, 'values')[0]}")
    
    def _download_file(self, file_info):
        """Download a file from SharePoint to a temporary location"""
        try:
            # Get the download URL
            file_id = file_info["id"]
            drive_id = file_info["parentReference"]["driveId"]
            
            graph_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}"
            
            headers = {
                "Authorization": f"Bearer {self.access_token}",
                "Accept": "application/json"
            }
            
            response = requests.get(graph_url, headers=headers)
            
            if response.status_code == 200:
                item_data = response.json()
                download_url = item_data.get("@microsoft.graph.downloadUrl")
                
                if download_url:
                    # Download file content
                    file_response = requests.get(download_url)
                    
                    if file_response.status_code == 200:
                        # Create a temporary file
                        temp_file = tempfile.NamedTemporaryFile(
                            delete=False, 
                            suffix=f"_{file_info['name']}"
                        )
                        temp_file.write(file_response.content)
                        temp_file.close()
                        
                        # Keep track of the temp file
                        self.temp_files.append({
                            "path": temp_file.name,
                            "sp_file_info": file_info
                        })
                        
                        return temp_file.name
            
            messagebox.showerror("Error", "Failed to download the file from SharePoint.")
            return None
                
        except Exception as e:
            messagebox.showerror("Download Error", f"Failed to download file: {str(e)}")
            return None
    
    def upload_file(self, local_file_path, target_name=None):
        """Upload a file to SharePoint"""
        try:
            # Check if this is a file we previously downloaded from SharePoint
            original_file_info = None
            for temp_file in self.temp_files:
                if temp_file["path"] == local_file_path:
                    original_file_info = temp_file["sp_file_info"]
                    break
                    
            if not original_file_info:
                # This is a new file, ask for upload location
                upload_location = self._select_upload_location()
                if not upload_location:
                    return None
                    
                drive_id = upload_location["drive_id"]
                parent_id = upload_location["folder_id"]
                folder_path = upload_location["folder_path"]
                
                # Use provided name or extract from local path
                if not target_name:
                    target_name = os.path.basename(local_file_path)
                    
                # Upload the file using Graph API
                with open(local_file_path, "rb") as f:
                    file_content = f.read()
                    
                # For files larger than 4MB, use upload session
                if len(file_content) > 4 * 1024 * 1024:
                    upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_id}:/{quote(target_name)}:/createUploadSession"
                    headers = {
                        "Authorization": f"Bearer {self.access_token}",
                        "Content-Type": "application/json"
                    }
                    
                    # Create upload session
                    session_response = requests.post(upload_url, headers=headers)
                    
                    if session_response.status_code == 200:
                        upload_session = session_response.json()
                        upload_url = upload_session["uploadUrl"]
                        
                        # Upload in chunks
                        chunk_size = 4 * 1024 * 1024  # 4MB
                        total_size = len(file_content)
                        
                        for i in range(0, total_size, chunk_size):
                            chunk = file_content[i:i+chunk_size]
                            
                            headers = {
                                "Content-Length": str(len(chunk)),
                                "Content-Range": f"bytes {i}-{i+len(chunk)-1}/{total_size}"
                            }
                            
                            chunk_response = requests.put(upload_url, headers=headers, data=chunk)
                            
                            if chunk_response.status_code not in [200, 201, 202]:
                                messagebox.showerror("Upload Error", f"Failed to upload chunk: {chunk_response.text}")
                                return None
                            
                        return True
                        
                else:
                    # Simple upload for smaller files
                    upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_id}:/{quote(target_name)}:/content"
                    
                    headers = {
                        "Authorization": f"Bearer {self.access_token}",
                        "Content-Type": "application/octet-stream"
                    }
                    
                    response = requests.put(upload_url, headers=headers, data=file_content)
                    
                    if response.status_code in [200, 201]:
                        return True
                    else:
                        messagebox.showerror("Upload Error", f"Failed to upload file: {response.text}")
                        return None
            else:
                # This is an update to an existing SharePoint file
                drive_id = original_file_info["parentReference"]["driveId"]
                file_id = original_file_info["id"]
                
                # Upload file content
                with open(local_file_path, "rb") as f:
                    file_content = f.read()
                
                upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
                
                headers = {
                    "Authorization": f"Bearer {self.access_token}",
                    "Content-Type": "application/octet-stream"
                }
                
                response = requests.put(upload_url, headers=headers, data=file_content)
                
                if response.status_code in [200, 201]:
                    return True
                else:
                    messagebox.showerror("Upload Error", f"Failed to update file: {response.text}")
                    return None
                    
        except Exception as e:
            messagebox.showerror("Upload Error", f"Failed to upload file: {str(e)}")
            return None
    
    def _select_upload_location(self):
        """Open a dialog to select a SharePoint location for upload"""
        # Similar to select_file but returns folder information
        # Implementation similar to select_file method
        # This would be implemented like select_file but for selecting folders
        pass
    
    def cleanup(self):
        """Remove temporary files"""
        for temp_file in self.temp_files:
            try:
                os.unlink(temp_file["path"])
            except:
                pass

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparisonApp(root)
    root.mainloop()