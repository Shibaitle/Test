import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
import os
import threading
import traceback
import gc
import difflib
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse, quote


class ExcelComparisonApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Compare and Replace Tool Phase 1")
        self.root.geometry("800x650")
        
        # Variables to store file paths and sheet names
        self.old_file_path = tk.StringVar()
        self.new_file_path = tk.StringVar()
        self.selected_sheet = tk.StringVar()
        self.team_column = tk.StringVar()
        self.app_name_column = tk.StringVar()
        self.category_column = tk.StringVar()

        # Add structure for additional comparison criteria columns
        self.additional_criteria = []  # Will hold tuples of (StringVar, label)
        
        # Add header row configuration
        self.header_row = tk.IntVar(value=4)  # Set default to row 4

        self._configure_button_styles()
        
        # Variables for filter criteria
        self.team_filters = []
        self.app_name_filters = []
        self.category_filters = []

        # Add these save option variables (before filter initialization)
        self.save_mode = tk.StringVar(value="new")  # Default to creating new file
        self.create_highlighted_file = tk.BooleanVar(value=False)
        self.show_update_popup = tk.BooleanVar(value=True)
        self.clear_after_update = tk.BooleanVar(value=True)

        # Storage for filters for additional criteria
        self.additional_filters = {}  # Dictionary mapping criteria label to list of filter StringVars
        self.additional_filter_frames = {}  # Dictionary mapping criteria label to its frame

        # Initialize with default 1 filter entry for each type
        for _ in range(1):
            self.team_filters.append(tk.StringVar())
            self.app_name_filters.append(tk.StringVar())
            self.category_filters.append(tk.StringVar())
        
        # Add variables to track formula relationships
        self.formula_relationships = {}
        
        # Add checkbox to enable formula awareness
        self.formula_aware = tk.BooleanVar(value=True)
        
        # Add variable to track which mode we're in (standard or custom)
        self.current_mode = tk.StringVar(value="standard")
        
        # Create frames for different modes
        self.standard_mode_frame = None
        self.custom_mode_frame = None
        
        # For custom mode
        self.key_columns = []  # Will store StringVars for key columns
        self.custom_filters = {}  # Will store custom filters
        
        # Create a main scrollable canvas
        self.main_canvas = tk.Canvas(root)
        self.main_scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.main_canvas.yview)
        
        # Configure the canvas
        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)
        self.main_canvas.bind('<Configure>', lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        
        # Create a frame inside the canvas to hold all content
        self.scrollable_frame = ttk.Frame(self.main_canvas)
        
        # Add mouse wheel scrolling to the canvas
        def _on_mousewheel(event):
            self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        self.main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Place the scrollable frame into the canvas
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Pack the scrollbar and canvas
        self.main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Create the main frame inside the scrollable area
        self.main_frame = ttk.Frame(self.scrollable_frame, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add mode selection at the top
        self._create_mode_selection(self.main_frame)
        
        # Create and organize widgets for standard mode
        self._create_standard_mode_ui()
        
        # Create custom mode UI (initially hidden)
        self._create_custom_mode_ui()
        
        # Show the default mode
        self._switch_mode(self.current_mode.get())
        
        # Configure the binding for window resize to update scrollregion
        self.scrollable_frame.bind("<Configure>", self._configure_scrollregion)

        # Add this to the __init__ method after creating the canvas
        self.main_canvas.bind("<Enter>", self._bind_mousewheel)
        self.main_canvas.bind("<Leave>", self._unbind_mousewheel)

        # Track the next column position for filter criteria
        self.next_filter_column = 3
        self.add_filter_btn = None  # Will be initialized in _create_filter_widgets
        self.first_filter_click = True  # Add this line: Flag for first filter criteria click

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
    
    def _create_mode_selection(self, parent):
        mode_frame = ttk.LabelFrame(parent, text="Tool Selection", padding="15")
        mode_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Add main title
        title_frame = ttk.Frame(mode_frame)
        title_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            title_frame, 
            text="Select Your Comparison Tool", 
            font=("Arial", 12, "bold"),
            foreground="#2C3E50"
        ).pack()
        
        # Excel Comparison Section
        excel_section = ttk.LabelFrame(mode_frame, text="Excel File Comparison", padding="15")
        excel_section.pack(fill=tk.X, pady=(0, 15))
        
        # Excel mode description
        excel_desc_frame = ttk.Frame(excel_section)
        excel_desc_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(
            excel_desc_frame,
            text="📊 Compare and update Excel files with advanced matching and filtering",
            font=("Arial", 9),
            foreground="#34495E"
        ).pack(anchor=tk.W)
        
        # Excel radio buttons in a more organized layout
        excel_options_frame = ttk.Frame(excel_section)
        excel_options_frame.pack(fill=tk.X, pady=5)
        
        # Standard mode
        standard_frame = ttk.Frame(excel_options_frame)
        standard_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        standard_radio = ttk.Radiobutton(
            standard_frame,
            text="📋 Standard Mode",
            variable=self.current_mode,
            value="standard",
            command=lambda: self._switch_mode("standard")
        )
        standard_radio.pack(anchor=tk.W)
        
        ttk.Label(
            standard_frame,
            text="Team/App Name/Category columns",
            font=("Arial", 8),
            foreground="#7F8C8D"
        ).pack(anchor=tk.W, padx=(20, 0))
        
        # Custom mode
        custom_frame = ttk.Frame(excel_options_frame)
        custom_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        
        custom_radio = ttk.Radiobutton(
            custom_frame,
            text="⚙️ Custom Mode", 
            variable=self.current_mode,
            value="custom",
            command=lambda: self._switch_mode("custom")
        )
        custom_radio.pack(anchor=tk.W)
        
        ttk.Label(
            custom_frame,
            text="Flexible column matching",
            font=("Arial", 8),
            foreground="#7F8C8D"
        ).pack(anchor=tk.W, padx=(20, 0))
        
        # Help button for Excel
        help_frame = ttk.Frame(excel_options_frame)
        help_frame.pack(side=tk.RIGHT)
        
        ttk.Button(
            help_frame,
            text="❓ Help",
            width=8,
            command=lambda: self._show_help_window("mode_selection")
        ).pack()
        
        # Text Comparison Section - More prominent
        text_section = ttk.LabelFrame(mode_frame, text="Text File Comparison", padding="15")
        text_section.pack(fill=tk.X, pady=(0, 10))
        
        # Text comparison content
        text_content_frame = ttk.Frame(text_section)
        text_content_frame.pack(fill=tk.X)
        
        # Left side - Description
        text_desc_frame = ttk.Frame(text_content_frame)
        text_desc_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        ttk.Label(
            text_desc_frame,
            text="📄 Text File Comparison Tool",
            font=("Arial", 11, "bold"),
            foreground="#2980B9"
        ).pack(anchor=tk.W)
        
        ttk.Label(
            text_desc_frame,
            text="Compare any text files (TXT, LOG, CSV, JSON, XML, Python, etc.)",
            font=("Arial", 9),
            foreground="#34495E"
        ).pack(anchor=tk.W, pady=(2, 0))
        
        # Features list
        features_frame = ttk.Frame(text_desc_frame)
        features_frame.pack(anchor=tk.W, pady=(5, 0))
        
        features = [
            "✓ Side-by-side comparison with syntax highlighting",
            "✓ Unified diff view and HTML reports", 
            "✓ Support for multiple file encodings",
            "✓ Advanced filtering and search options"
        ]
        
        for feature in features:
            ttk.Label(
                features_frame,
                text=feature,
                font=("Arial", 8),
                foreground="#27AE60"
            ).pack(anchor=tk.W)
        
        # Right side - Action buttons
        text_buttons_frame = ttk.Frame(text_content_frame)
        text_buttons_frame.pack(side=tk.RIGHT, padx=(20, 0))
        
        # Main text comparison button - more prominent
        text_compare_btn = ttk.Button(
            text_buttons_frame,
            text="🚀 Open Text Comparison Tool",
            command=self._open_text_comparison,
            width=25
        )
        text_compare_btn.pack(pady=(0, 5))
        
        # Style the button to make it more prominent
        text_compare_btn.configure(cursor="hand2")
        
        # Help button for text comparison
        ttk.Button(
            text_buttons_frame,
            text="❓ Help & Guide",
            command=lambda: self._show_help_window("text_comparison"),
            width=25
        ).pack()
    
    def _open_text_comparison(self):
        """Open the text comparison tool in a new window"""
        try:
            print("Attempting to open text comparison tool...")
            
            # Check if text_compare.py exists
            text_compare_path = os.path.join(os.path.dirname(__file__), 'text_compare.py')
            if not os.path.exists(text_compare_path):
                messagebox.showerror(
                    "Error", 
                    f"text_compare.py not found at: {text_compare_path}\n\n"
                    "Please make sure 'text_compare.py' is in the same directory as this file."
                )
                return
            
            print("text_compare.py found, importing...")
            
            # Import the text comparison module
            from text_compare import TextComparisonApp
            
            print("TextComparisonApp imported successfully")
            
            # Create a new window for text comparison
            text_window = tk.Toplevel(self.root)
            text_window.transient(self.root)
            text_window.grab_set()
            
            # Force window to be visible
            text_window.lift()
            text_window.focus_force()
            text_window.attributes('-topmost', True)
            text_window.after(100, lambda: text_window.attributes('-topmost', False))
            
            print("Creating TextComparisonApp instance...")
            
            # Create the text comparison app
            text_app = TextComparisonApp(text_window)
            text_app.parent_app = self.root
            
            print("Text comparison app created successfully")
            
            # Hide the main window while text comparison is open
            self.root.withdraw()
            
            # When text comparison window is closed, show main window again
            def on_text_window_close():
                print("Text comparison window closing...")
                self.root.deiconify()
                text_window.destroy()
            
            text_window.protocol("WM_DELETE_WINDOW", on_text_window_close)
            
            # Force update and make visible
            text_window.update()
            text_window.deiconify()
            
            print("Text comparison tool opened successfully!")
            
        except ImportError as e:
            print(f"Import error: {e}")
            messagebox.showerror(
                "Error", 
                f"Failed to load text comparison module: {str(e)}\n\n"
                "Make sure 'text_compare.py' is in the same directory as this file."
            )
        except Exception as e:
            print(f"General error: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to open text comparison tool: {str(e)}")
        
    def _switch_mode(self, mode):
        if mode == "standard":
            if self.custom_mode_frame:
                self.custom_mode_frame.pack_forget()
            if self.standard_mode_frame:
                self.standard_mode_frame.pack(fill=tk.BOTH, expand=True)
        else:  # custom mode
            if self.standard_mode_frame:
                self.standard_mode_frame.pack_forget()
            if self.custom_mode_frame:
                self.custom_mode_frame.pack(fill=tk.BOTH, expand=True)
    
    def _create_standard_mode_ui(self):
        # Create a container frame for standard mode
        self.standard_mode_frame = ttk.Frame(self.main_frame)
        
        # Create and organize widgets
        self._create_file_selection_widgets(self.standard_mode_frame)
        self._create_sheet_selection_widgets(self.standard_mode_frame)
        self._create_criteria_widgets(self.standard_mode_frame)
        self._create_filter_widgets(self.standard_mode_frame)
        self._create_action_buttons(self.standard_mode_frame)
        
        # Progress bar and status label
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Ready")

        # Use the new status section
        self._create_status_section(self.standard_mode_frame, custom_mode=False)
        
        status_frame = ttk.Frame(self.standard_mode_frame)
        status_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(status_frame, text="Status:").pack(side=tk.LEFT)
        ttk.Label(status_frame, textvariable=self.status_var).pack(side=tk.LEFT, padx=(5, 0))
        
        self.progress_bar = ttk.Progressbar(
            self.standard_mode_frame, 
            orient=tk.HORIZONTAL, 
            length=100, 
            mode='determinate',
            variable=self.progress_var
        )
        self.progress_bar.pack(fill=tk.X, pady=(10, 0))
    
    def _create_custom_mode_ui(self):
        # Create a container frame for custom mode
        self.custom_mode_frame = ttk.Frame(self.main_frame)
        
        # File selection section
        self._create_file_selection_widgets(self.custom_mode_frame, custom_mode=True)
        
        # Sheet selection section
        self._create_sheet_selection_widgets(self.custom_mode_frame, custom_mode=True)
        
        # Custom criteria selection section
        self._create_custom_criteria_widgets(self.custom_mode_frame)
        
        # Custom filters section
        self._create_custom_filter_widgets(self.custom_mode_frame)
        
        # Action buttons
        self._create_action_buttons(self.custom_mode_frame, custom_mode=True)
        
        # Progress bar and status
        self.custom_progress_var = tk.DoubleVar()
        self.custom_status_var = tk.StringVar(value="Ready")

        # Use the new status section
        self._create_status_section(self.custom_mode_frame, custom_mode=True)
        
        status_frame = ttk.Frame(self.custom_mode_frame)
        status_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(status_frame, text="Status:").pack(side=tk.LEFT)
        ttk.Label(status_frame, textvariable=self.custom_status_var).pack(side=tk.LEFT, padx=(5, 0))
        
        self.custom_progress_bar = ttk.Progressbar(
            self.custom_mode_frame, 
            orient=tk.HORIZONTAL, 
            length=100, 
            mode='determinate',
            variable=self.custom_progress_var
        )
        self.custom_progress_bar.pack(fill=tk.X, pady=(10, 0))
    
    def _on_close(self):
        """Handle application close"""
        # Close the application
        self.root.destroy()

    def _configure_scrollregion(self, event):
        # Update the scrollregion to encompass the inner frame
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def _create_file_selection_widgets(self, parent, custom_mode=False):
        # Main container with better spacing
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header with icon and title
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame, 
            text="📁 File Selection", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button with better styling
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("file_selection")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Select the Excel files you want to compare and update",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Content frame with better organization
        content_frame = ttk.Frame(container)
        content_frame.pack(fill=tk.X)
        
        # Old file section
        old_file_section = ttk.LabelFrame(
            content_frame, 
            text="📄 Source File (To Update)",
            padding=15,
            style="Card.TLabelframe"
        )
        old_file_section.pack(fill=tk.X, pady=(0, 10))
        
        old_desc = ttk.Label(
            old_file_section,
            text="This file will be updated with new data",
            font=("Segoe UI", 8),
            foreground="#605E5C"
        )
        old_desc.pack(anchor=tk.W, pady=(0, 8))
        
        old_file_frame = ttk.Frame(old_file_section)
        old_file_frame.pack(fill=tk.X)
        
        old_entry = ttk.Entry(
            old_file_frame, 
            textvariable=self.old_file_path, 
            font=("Segoe UI", 9),
            width=60
        )
        old_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        old_browse_btn = ttk.Button(
            old_file_frame,
            text="🗂️ Browse",
            style="Secondary.TButton",
            command=self._browse_old_file
        )
        old_browse_btn.pack(side=tk.RIGHT)
        
        # New file section
        new_file_section = ttk.LabelFrame(
            content_frame, 
            text="📊 Reference File (Source Data)",
            padding=15,
            style="Card.TLabelframe"
        )
        new_file_section.pack(fill=tk.X)
        
        new_desc = ttk.Label(
            new_file_section,
            text="Data will be copied from this file to the source file",
            font=("Segoe UI", 8),
            foreground="#605E5C"
        )
        new_desc.pack(anchor=tk.W, pady=(0, 8))
        
        new_file_frame = ttk.Frame(new_file_section)
        new_file_frame.pack(fill=tk.X)
        
        new_entry = ttk.Entry(
            new_file_frame, 
            textvariable=self.new_file_path, 
            font=("Segoe UI", 9),
            width=60
        )
        new_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        new_browse_btn = ttk.Button(
            new_file_frame,
            text="🗂️ Browse",
            style="Secondary.TButton",
            command=self._browse_new_file
        )
        new_browse_btn.pack(side=tk.RIGHT)
        

    def _create_sheet_selection_widgets(self, parent, custom_mode=False):
        # Main container
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header section
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame,
            text="📋 Sheet Selection", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("sheet_selection")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Load and select the worksheets you want to process",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Action buttons section
        actions_frame = ttk.Frame(container)
        actions_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Load sheets button
        load_sheets_btn = ttk.Button(
            actions_frame,
            text="🔄 Load Sheets",
            style="Accent.TButton",
            command=self._load_sheets
        )
        load_sheets_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Load columns button
        load_columns_btn = ttk.Button(
            actions_frame,
            text="📊 Load Columns",
            style="Secondary.TButton",
            command=self._on_sheet_selected
        )
        load_columns_btn.pack(side=tk.LEFT)
        
        # Status indicator
        self.sheet_status_label = ttk.Label(
            actions_frame,
            text="",
            font=("Segoe UI", 8),
            foreground="#107C10"
        )
        self.sheet_status_label.pack(side=tk.RIGHT, padx=(10, 0))
        
        # Sheets container with improved styling
        sheets_container = ttk.LabelFrame(
            container,
            text="Available Sheets",
            padding=15,
            style="Card.TLabelframe"
        )
        sheets_container.pack(fill=tk.X)
        
        # Canvas setup for scrollable sheet list
        canvas_frame = ttk.Frame(sheets_container)
        canvas_frame.pack(fill=tk.X)
        
        canvas = tk.Canvas(
            canvas_frame, 
            height=120, 
            borderwidth=0, 
            highlightthickness=0,
            background="#FAFAFA"
        )
        
        scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=canvas.yview)
        
        # Create frame for checkboxes
        if custom_mode:
            inner_frame = ttk.Frame(canvas)
            self.custom_sheet_checkbox_frame = inner_frame
        else:
            inner_frame = ttk.Frame(canvas)
            self.standard_sheet_checkbox_frame = inner_frame
        
        self.sheet_checkbox_frame = inner_frame
        
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        canvas_window = canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        
        # Configure scrolling
        def _configure_inner_frame(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        def _update_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        canvas.bind("<Configure>", _configure_inner_frame)
        inner_frame.bind("<Configure>", _update_scroll_region)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            if event.delta:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            elif event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel)
        canvas.bind_all("<Button-5>", _on_mousewheel)
        
        # Store references
        if custom_mode:
            self.custom_sheet_canvas = canvas
        else:
            self.standard_sheet_canvas = canvas
        
        self.sheet_vars = {}

    # Add this new method to handle sheet selection
    def _on_sheet_selected(self):
        """Handle when a sheet is selected via checkbox"""
        # Debug print to verify this method is running
        print("Sheet selection changed - loading columns...")
        
        # Check if any sheet is selected
        selected = False
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected = True
                selected_sheet = sheet
                break
        
        if not selected:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        # Explicitly load columns based on current mode
        current_mode = self.current_mode.get()
        print(f"Current mode: {current_mode}, Selected sheet: {selected_sheet}")
        
        try:
            if current_mode == "standard":
                print("Loading standard columns...")
                self._load_columns()
            else:
                # For custom mode
                print("Loading custom columns...")
                self._load_custom_columns()
                
            # Provide visual confirmation that columns were loaded
            messagebox.showinfo("Success", f"Columns loaded from sheet '{selected_sheet}'.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load columns: {str(e)}")
            import traceback
            traceback.print_exc()
        
    def _create_criteria_widgets(self, parent):
        # Main container
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header section
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame,
            text="🎯 Comparison Criteria", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("comparison_criteria")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Configure how rows are matched between files",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Configuration section
        config_frame = ttk.LabelFrame(
            container,
            text="⚙️ Configuration",
            padding=15,
            style="Card.TLabelframe"
        )
        config_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Header row and formula settings
        settings_frame = ttk.Frame(config_frame)
        settings_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Header row setting
        header_frame = ttk.Frame(settings_frame)
        header_frame.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(
            header_frame, 
            text="📍 Header Row:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        header_spin = ttk.Spinbox(
            header_frame, 
            from_=1, 
            to=20, 
            textvariable=self.header_row, 
            width=5,
            font=("Segoe UI", 9)
        )
        header_spin.pack(side=tk.LEFT)
        
        # Formula awareness
        formula_frame = ttk.Frame(settings_frame)
        formula_frame.pack(side=tk.LEFT)
        
        ttk.Checkbutton(
            formula_frame,
            text="🧮 Formula-Aware Processing",
            variable=self.formula_aware,
            style="Modern.TCheckbutton"
        ).pack()
        
        # Column matching section
        columns_frame = ttk.LabelFrame(
            container,
            text="🔗 Column Matching",
            padding=15,
            style="Card.TLabelframe"
        )
        columns_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Instructions
        instruction_label = ttk.Label(
            columns_frame,
            text="Select the columns used to match rows between files:",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        instruction_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Column selection grid
        self.column_frame = ttk.Frame(columns_frame)
        self.column_frame.pack(fill=tk.X)
        
        # Team column
        team_frame = ttk.Frame(self.column_frame)
        team_frame.grid(row=0, column=0, columnspan=3, sticky=tk.W+tk.E, pady=5)
        
        ttk.Label(
            team_frame, 
            text="👥 Team Column:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, anchor=tk.W, padx=(0, 10))
        
        self.team_combobox = ttk.Combobox(
            team_frame, 
            textvariable=self.team_column, 
            state="readonly", 
            width=35,
            font=("Segoe UI", 9)
        )
        self.team_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # App Name column
        app_frame = ttk.Frame(self.column_frame)
        app_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E, pady=5)
        
        ttk.Label(
            app_frame, 
            text="📱 App Name Column:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, anchor=tk.W, padx=(0, 10))
        
        self.app_name_combobox = ttk.Combobox(
            app_frame, 
            textvariable=self.app_name_column, 
            state="readonly", 
            width=35,
            font=("Segoe UI", 9)
        )
        self.app_name_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # Category column
        category_frame = ttk.Frame(self.column_frame)
        category_frame.grid(row=2, column=0, columnspan=3, sticky=tk.W+tk.E, pady=5)
        
        ttk.Label(
            category_frame, 
            text="📂 Category Column:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, anchor=tk.W, padx=(0, 10))
        
        self.category_combobox = ttk.Combobox(
            category_frame, 
            textvariable=self.category_column, 
            state="readonly", 
            width=35,
            font=("Segoe UI", 9)
        )
        self.category_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # Load columns button
        load_btn = ttk.Button(
            category_frame,
            text="🔄 Load Columns",
            style="Secondary.TButton",
            command=self._load_columns
        )
        load_btn.pack(side=tk.RIGHT)
        
        # Additional criteria section
        self.additional_criteria_frame = ttk.Frame(self.column_frame)
        self.additional_criteria_frame.grid(row=3, column=0, columnspan=3, sticky=tk.W+tk.E, pady=(15, 0))
        
        # Add criteria button
        add_criteria_frame = ttk.Frame(self.column_frame)
        add_criteria_frame.grid(row=4, column=0, columnspan=3, sticky=tk.W, pady=(15, 0))
        
        ttk.Button(
            add_criteria_frame,
            text="➕ Add Additional Criteria",
            style="Secondary.TButton",
            command=self._add_criteria_column
        ).pack(side=tk.LEFT)

    def _add_criteria_column(self):
        """Add a new comparison criteria column"""
        # Get current column values
        current_columns = self.team_combobox['values']
        if not current_columns:
            messagebox.showerror("Error", "Please load columns first")
            return
        
        # Create a new criteria row
        criteria_var = tk.StringVar()
        criteria_label = f"Additional Criteria {len(self.additional_criteria) + 1}"
        
        # Create a frame for this criteria with label, combobox, and remove button
        row_frame = ttk.Frame(self.additional_criteria_frame)
        row_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(row_frame, text=f"{criteria_label}:").pack(side=tk.LEFT, padx=5)
        combobox = ttk.Combobox(
            row_frame, 
            textvariable=criteria_var, 
            state="readonly", 
            width=30, 
            values=current_columns
        )
        combobox.pack(side=tk.LEFT, padx=5)
        
        # Add remove button
        ttk.Button(
            row_frame, 
            text="Remove", 
            command=lambda frame=row_frame, idx=len(self.additional_criteria): self._remove_criteria_column(frame, idx)
        ).pack(side=tk.LEFT, padx=5)
        
        # Store the criteria
        self.additional_criteria.append((criteria_var, criteria_label, row_frame))
        
        # Initialize filter list for this criteria
        self.additional_filters[criteria_label] = [tk.StringVar()]

    def _remove_criteria_column(self, frame, idx):
        """Remove a criteria column"""
        if idx < len(self.additional_criteria):
            # Get the criteria to remove
            _, criteria_label, _ = self.additional_criteria[idx]
            
            # Remove filters associated with this criteria
            if criteria_label in self.additional_filters:
                del self.additional_filters[criteria_label]
            
            # Remove from list
            self.additional_criteria.pop(idx)
            
            # Destroy the UI elements
            frame.destroy()
            
            # Renumber the remaining criteria
            for i, (var, _, frame) in enumerate(self.additional_criteria[idx:], idx+1):
                # Update label text
                for child in frame.winfo_children():
                    if isinstance(child, ttk.Label):
                        child.config(text=f"Additional Criteria {i}:")
                        break
    
    def _toggle_comparison_mode(self):
        if self.use_row_mode.get():
            # Switch to row-based comparison
            self.column_frame.grid_remove()
            self.row_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E)
            # Load columns for key selection
            self._load_key_columns()
        else:
            # Switch to column-based comparison
            self.row_frame.grid_remove()
            self.column_frame.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E)
    
    def _load_key_columns(self):
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Read the Excel file to get column count
            df = pd.read_excel(old_file, sheet_name=selected_sheet, nrows=1)
            columns = list(df.columns)
            
            # Create letter-based column references (A, B, C, etc.)
            column_refs = [get_column_letter(i+1) for i in range(len(columns))]
            
            # Update key column combobox with both letter and name
            column_options = [f"{ref} - {name}" for ref, name in zip(column_refs, columns)]
            self.key_column_combobox['values'] = column_options
            
            if column_options:
                self.key_column_combobox.current(0)
            
            # Set default row values if empty
            if not self.team_row.get():
                self.team_row.set("1")  # Default to row 1
            if not self.app_name_row.get():
                self.app_name_row.set("2")  # Default to row 2
            if not self.category_row.get():
                self.category_row.set("3")  # Default to row 3
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load key columns: {str(e)}")
        
    def _create_filter_widgets(self, parent):
        # Main container
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header section
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame,
            text="🔍 Filter Criteria", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("filter_criteria")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Filter which rows to process (leave empty to include all matching rows)",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Filter content
        filter_content = ttk.LabelFrame(
            container,
            text="🔧 Active Filters",
            padding=15,
            style="Card.TLabelframe"
        )
        filter_content.pack(fill=tk.X, pady=(0, 15))
        
        # Team filters
        self.team_filter_frame = ttk.LabelFrame(
            filter_content, 
            text="👥 Team Filters",
            padding=10,
            style="Nested.TLabelframe"
        )
        self.team_filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        # App Name filters
        self.app_filter_frame = ttk.LabelFrame(
            filter_content, 
            text="📱 App Name Filters",
            padding=10,
            style="Nested.TLabelframe"
        )
        self.app_filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Category filters
        self.category_filter_frame = ttk.LabelFrame(
            filter_content, 
            text="📂 Category Filters",
            padding=10,
            style="Nested.TLabelframe"
        )
        self.category_filter_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Additional filters container
        self.additional_filters_container = ttk.Frame(filter_content)
        self.additional_filters_container.pack(fill=tk.X, pady=5)
        
        # Add filter criteria button
        add_filter_frame = ttk.Frame(filter_content)
        add_filter_frame.pack(anchor=tk.W, pady=(10, 0))
        
        self.add_filter_btn = ttk.Button(
            add_filter_frame,
            text="➕ Add Custom Filter",
            style="Secondary.TButton",
            command=self._add_filter_criteria,
            width=20
        )
        self.add_filter_btn.pack(side=tk.LEFT)
        
        # Initialize filter widgets
        self._refresh_filter_widgets()

    def _refresh_filter_widgets(self):
        # Clear existing widgets from filter frames
        for widget in self.team_filter_frame.winfo_children():
            widget.destroy()
            
        for widget in self.app_filter_frame.winfo_children():
            widget.destroy()
            
        for widget in self.category_filter_frame.winfo_children():
            widget.destroy()
        
        # Team filters - horizontal arrangement
        for i, filter_var in enumerate(self.team_filters):
            # Create filter entry with compact buttons side by side
            entry = ttk.Entry(self.team_filter_frame, textvariable=filter_var, width=20)
            entry.grid(row=0, column=i*2, padx=2, pady=5)
            
            # Button frame for icons (vertical stack of buttons)
            btn_frame = ttk.Frame(self.team_filter_frame)
            btn_frame.grid(row=0, column=i*2+1, padx=2, pady=5)
            
            # Get Values button with icon
            get_btn = ttk.Button(
                btn_frame, 
                text="🔍", 
                width=2,
                command=lambda idx=i: self._get_unique_values('team', idx)
            )
            get_btn.pack(fill="x", pady=1)
            
            # Delete button with icon - only if we have more than one filter
            if len(self.team_filters) > 1:
                del_btn = ttk.Button(
                    btn_frame, 
                    text="❌", 
                    width=2,
                    command=lambda idx=i: self._delete_team_filter(idx)
                )
                del_btn.pack(fill="x", pady=1)
        
        # Add filter button at the end of the row
        add_btn = ttk.Button(
            self.team_filter_frame, 
            text="+", 
            width=2,
            command=self._add_team_filter
        )
        add_btn.grid(row=0, column=len(self.team_filters)*2, padx=2, pady=5)
        
        # App name filters - horizontal arrangement
        for i, filter_var in enumerate(self.app_name_filters):
            entry = ttk.Entry(self.app_filter_frame, textvariable=filter_var, width=20)
            entry.grid(row=0, column=i*2, padx=2, pady=5)
            
            # Button frame for icons
            btn_frame = ttk.Frame(self.app_filter_frame)
            btn_frame.grid(row=0, column=i*2+1, padx=2, pady=5)
            
            # Get Values button with icon
            get_btn = ttk.Button(
                btn_frame, 
                text="🔍", 
                width=2,
                command=lambda idx=i: self._get_unique_values('app', idx)
            )
            get_btn.pack(fill="x", pady=1)
            
            # Delete button with icon - only if we have more than one filter
            if len(self.app_name_filters) > 1:
                del_btn = ttk.Button(
                    btn_frame, 
                    text="❌", 
                    width=2,
                    command=lambda idx=i: self._delete_app_filter(idx)
                )
                del_btn.pack(fill="x", pady=1)
        
        # Add filter button at the end of the row
        add_btn = ttk.Button(
            self.app_filter_frame, 
            text="+", 
            width=2,
            command=self._add_app_filter
        )
        add_btn.grid(row=0, column=len(self.app_name_filters)*2, padx=2, pady=5)
        
        # Category filters - horizontal arrangement
        for i, filter_var in enumerate(self.category_filters):
            entry = ttk.Entry(self.category_filter_frame, textvariable=filter_var, width=20)
            entry.grid(row=0, column=i*2, padx=2, pady=5)
            
            # Button frame for icons
            btn_frame = ttk.Frame(self.category_filter_frame)
            btn_frame.grid(row=0, column=i*2+1, padx=2, pady=5)
            
            # Get Values button with icon
            get_btn = ttk.Button(
                btn_frame, 
                text="🔍", 
                width=2,
                command=lambda idx=i: self._get_unique_values('category', idx)
            )
            get_btn.pack(fill="x", pady=1)
            
            # Delete button with icon - only if we have more than one filter
            if len(self.category_filters) > 1:
                del_btn = ttk.Button(
                    btn_frame, 
                    text="❌", 
                    width=2,
                    command=lambda idx=i: self._delete_category_filter(idx)
                )
                del_btn.pack(fill="x", pady=1)
        
        # Add filter button at the end of the row
        add_btn = ttk.Button(
            self.category_filter_frame, 
            text="+", 
            width=2,
            command=self._add_category_filter
        )
        add_btn.grid(row=0, column=len(self.category_filters)*2, padx=2, pady=5)
        
        # Refresh additional filter criteria frames - horizontal arrangement
        for criteria_label, filter_frame in self.additional_filter_frames.items():
            # Clear existing widgets from this filter frame
            for widget in filter_frame.winfo_children():
                widget.destroy()
            
            # Add filters for this criteria horizontally
            filters = self.additional_filters.get(criteria_label, [])
            for i, filter_var in enumerate(filters):
                # Create entry horizontally
                entry = ttk.Entry(filter_frame, textvariable=filter_var, width=20)
                entry.grid(row=0, column=i*2, padx=2, pady=5)
                
                # Create button frame for icons
                btn_frame = ttk.Frame(filter_frame)
                btn_frame.grid(row=0, column=i*2+1, padx=2, pady=5)
                
                # Get Values button with icon
                get_btn = ttk.Button(
                    btn_frame, 
                    text="🔍", 
                    width=2,
                    command=lambda label=criteria_label, idx=i: self._get_additional_unique_values(label, idx)
                )
                get_btn.pack(fill="x", pady=1)
                
                # Delete button with icon
                if len(filters) > 1:
                    del_btn = ttk.Button(
                        btn_frame, 
                        text="❌", 
                        width=2,
                        command=lambda label=criteria_label, idx=i: self._delete_additional_filter(label, idx)
                    )
                    del_btn.pack(fill="x", pady=1)
            
            # Add button at the right end
            add_btn = ttk.Button(
                filter_frame, 
                text="+", 
                width=2,
                command=lambda label=criteria_label: self._add_additional_filter(label)
            )
            add_btn.grid(row=0, column=len(filters)*2, padx=2, pady=5)

    def _add_team_filter(self):
        self.team_filters.append(tk.StringVar())
        self._refresh_filter_widgets()

    def _add_app_filter(self):
        self.app_name_filters.append(tk.StringVar())
        self._refresh_filter_widgets()

    def _add_category_filter(self):
        self.category_filters.append(tk.StringVar())
        self._refresh_filter_widgets()
    
    def _create_action_buttons(self, parent, custom_mode=False):
        # Main container
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header section
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame,
            text="💾 Save Options", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("save_options")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Configure how the updated data will be saved",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Save options section
        save_options_frame = ttk.LabelFrame(
            container,
            text="💾 Output Configuration",
            padding=15,
            style="Card.TLabelframe"
        )
        save_options_frame.pack(fill=tk.X, pady=(0, 20))
        
        # File mode section
        file_mode_frame = ttk.Frame(save_options_frame)
        file_mode_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            file_mode_frame,
            text="📁 File Output Mode:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(anchor=tk.W, pady=(0, 8))
        
        mode_frame = ttk.Frame(file_mode_frame)
        mode_frame.pack(fill=tk.X, padx=15)
        
        ttk.Radiobutton(
            mode_frame,
            text="📄 Create new updated file (Recommended)",
            variable=self.save_mode,
            value="new",
            style="Modern.TRadiobutton"
        ).pack(anchor=tk.W, pady=2)
        
        ttk.Radiobutton(
            mode_frame,
            text="⚠️ Replace original file",
            variable=self.save_mode,
            value="replace",
            style="Modern.TRadiobutton"
        ).pack(anchor=tk.W, pady=2)
        
        # Additional options section
        additional_frame = ttk.Frame(save_options_frame)
        additional_frame.pack(fill=tk.X)
        
        ttk.Label(
            additional_frame,
            text="🎨 Additional Options:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(anchor=tk.W, pady=(0, 8))
        
        options_frame = ttk.Frame(additional_frame)
        options_frame.pack(fill=tk.X, padx=15)
        
        ttk.Checkbutton(
            options_frame,
            text="🌟 Create highlighted changes file",
            variable=self.create_highlighted_file,
            style="Modern.TCheckbutton"
        ).pack(anchor=tk.W, pady=2)
        
        ttk.Checkbutton(
            options_frame,
            text="📊 Show update summary popup",
            variable=self.show_update_popup,
            style="Modern.TCheckbutton"
        ).pack(anchor=tk.W, pady=2)
        
        ttk.Checkbutton(
            options_frame,
            text="🧹 Clear file selection after update",
            variable=self.clear_after_update,
            style="Modern.TCheckbutton"
        ).pack(anchor=tk.W, pady=2)
        
        # Action buttons section
        action_frame = ttk.Frame(container)
        action_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Left side - secondary actions
        left_actions = ttk.Frame(action_frame)
        left_actions.pack(side=tk.LEFT)
        
        # Right side - primary actions
        right_actions = ttk.Frame(action_frame)
        right_actions.pack(side=tk.RIGHT)
        
        # Exit button
        ttk.Button(
            left_actions,
            text="🚪 Exit",
            style="Secondary.TButton",
            command=self.root.quit,
            width=12
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        # Main action button
        ttk.Button(
            right_actions,
            text="🚀 Compare and Update",
            style="Success.TButton",
            command=self._start_compare_update,
            width=20
        ).pack(side=tk.RIGHT)
        
    def _browse_old_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Old Excel File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if file_path:
            # Close any potentially open workbook references
            import gc
            gc.collect()
            
            # Reset the formula relationships
            self.formula_relationships = {}
            
            # Clear sheet variables from previous selections
            self.sheet_vars.clear()
            
            # Set the new file path
            self.old_file_path.set(file_path)
    
    def _browse_new_file(self):
        file_path = filedialog.askopenfilename(
            title="Select New Excel File",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if file_path:
            self.new_file_path.set(file_path)

    def _load_sheets(self):
        old_file = self.old_file_path.get()
        new_file = self.new_file_path.get()
        
        if not old_file or not new_file:
            messagebox.showerror("Error", "Please select both old and new Excel files.")
            return
        
        try:
            # Load old workbook without read-only to allow editing
            old_wb = openpyxl.load_workbook(old_file, read_only=False, data_only=False)
            
            # Load new workbook in read-only mode initially for efficiency
            new_wb_readonly = openpyxl.load_workbook(new_file, read_only=True, data_only=True)
            
            old_sheets = set(old_wb.sheetnames)
            new_sheets = set(new_wb_readonly.sheetnames)
            
            # Find common sheets in both files
            common_sheets = list(old_sheets.intersection(new_sheets))
            
            if not common_sheets:
                messagebox.showerror("Error", "No common sheets found between the two Excel files.")
                return
            
            # Clear the variables dictionary
            self.sheet_vars.clear()
            
            # Determine which checkbox frame to use based on current mode
            current_mode = self.current_mode.get()
            
            # Clear existing checkboxes in BOTH frames
            if hasattr(self, 'standard_sheet_checkbox_frame'):
                for widget in self.standard_sheet_checkbox_frame.winfo_children():
                    widget.destroy()
                    
            if hasattr(self, 'custom_sheet_checkbox_frame'):
                for widget in self.custom_sheet_checkbox_frame.winfo_children():
                    widget.destroy()
            
            # Create checkbox for each common sheet
            for i, sheet in enumerate(common_sheets):
                var = tk.BooleanVar(value=False)
                self.sheet_vars[sheet] = var
                
                # Create checkbox in standard mode frame if it exists
                if hasattr(self, 'standard_sheet_checkbox_frame'):
                    checkbox = ttk.Checkbutton(
                        self.standard_sheet_checkbox_frame, 
                        text=sheet, 
                        variable=var
                    )
                    checkbox.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
                
                # Create checkbox in custom mode frame if it exists
                if hasattr(self, 'custom_sheet_checkbox_frame'):
                    checkbox = ttk.Checkbutton(
                        self.custom_sheet_checkbox_frame, 
                        text=sheet, 
                        variable=var
                    )
                    checkbox.grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            
            # Detect formula relationships if enabled
            if self.formula_aware.get():
                header_row = self.header_row.get()
                
                # Clear existing relationships
                self.formula_relationships = {}
                
                try:
                    # Load first sheet to detect relationships
                    if old_wb.sheetnames:
                        first_sheet = old_wb[old_wb.sheetnames[0]]
                        self.formula_relationships = self._detect_formula_relationships(first_sheet, header_row)
                except Exception as formula_error:
                    print(f"Warning: Could not detect formula relationships: {formula_error}")
                    # Don't fail the whole process if formula detection fails
                    self.formula_relationships = {}
                    
            # Always close the workbooks properly
            old_wb.close()
            new_wb_readonly.close()
            
            messagebox.showinfo("Success", f"Found {len(common_sheets)} common sheets.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets: {str(e)}")
            import traceback
            traceback.print_exc()

    def _get_unique_values(self, field_type, index=0):
        # Find the first selected sheet from checkboxes
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select a file and sheet first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Get formula relationships map if enabled
            formula_map = self.formula_relationships if self.formula_aware.get() else {}
            
            # Determine which column to get unique values from and its source
            if field_type == 'team':
                column = self.team_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select Team Column first.")
                    return
                source_column = formula_map.get(column, column)  # Get source column if it's a formula
                target_var = self.team_filters[index]
            elif field_type == 'app':
                column = self.app_name_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select App Name Column first.")
                    return
                source_column = formula_map.get(column, column)
                target_var = self.app_name_filters[index]
            elif field_type == 'category':
                column = self.category_column.get()
                if not column:
                    messagebox.showerror("Error", "Please select Category Column first.")
                    return
                source_column = formula_map.get(column, column)
                target_var = self.category_filters[index]
                
            # Use openpyxl to read both raw values AND formula results
            wb = openpyxl.load_workbook(old_file, data_only=True)  # data_only=True gets formula results
            sheet = wb[selected_sheet]
            
            # Find the column index for the source column
            source_col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row+1, column=col).value
                if cell_value == source_column:
                    source_col_idx = col
                    break
                    
            if source_col_idx is None:
                messagebox.showerror("Error", f"Could not find column {source_column} in sheet.")
                return
                
            # Collect unique non-empty values from this column
            unique_values = set()
            for row in range(header_row+2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=source_col_idx).value
                if cell_value:  # Only add non-empty values
                    unique_values.add(str(cell_value))
                    
            # Sort the unique values
            unique_values = sorted(list(unique_values))
            
            # Create a selection dialog
            value_dialog = tk.Toplevel(self.root)
            value_dialog.title(f"Select {column} Value")
            value_dialog.geometry("400x300")
            value_dialog.transient(self.root)
            value_dialog.grab_set()
            
            # Create a frame with a scrollbar
            frame = ttk.Frame(value_dialog)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Create a scrollable listbox
            scrollbar = ttk.Scrollbar(frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=50, height=15)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar.config(command=listbox.yview)
            
            # Add search functionality
            search_frame = ttk.Frame(value_dialog)
            search_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
            
            ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
            search_var = tk.StringVar()
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
            search_entry.pack(side=tk.LEFT, padx=5)
            
            def filter_values(*args):
                search_term = search_var.get().lower()
                listbox.delete(0, tk.END)
                for value in unique_values:
                    if search_term in str(value).lower():
                        listbox.insert(tk.END, str(value))
                        
            search_var.trace_add("write", filter_values)
            
            # Populate the listbox initially
            for value in unique_values:
                listbox.insert(tk.END, str(value))
            
            # Function to set the selected value
            def set_selected_value():
                try:
                    selected_indices = listbox.curselection()
                    if selected_indices:
                        selected_value = listbox.get(selected_indices[0])
                        target_var.set(selected_value)
                    value_dialog.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to set value: {str(e)}")
                    value_dialog.destroy()
            
            # Add buttons
            button_frame = ttk.Frame(value_dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Select", command=set_selected_value).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=value_dialog.destroy).pack(side=tk.RIGHT, padx=5)
            
            # Focus on search box
            search_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get unique values: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _update_status(self, message, progress=None):
        self.status_var.set(message)
        if progress is not None:
            self.progress_var.set(progress)
        self.root.update_idletasks()
    
    def _start_compare_update(self):
        """Start comparison and update process based on current mode"""
        # Validate inputs
        if not self._validate_inputs():
            return
        
        # Start in a thread based on mode
        if self.current_mode.get() == "standard":
            threading.Thread(target=self._compare_and_update, daemon=True).start()
        else:  # custom mode
            threading.Thread(target=self._custom_compare_and_update, daemon=True).start()
    
    def _validate_inputs(self):
        # Check if all required files are selected
        if not self.old_file_path.get():
            messagebox.showerror("Error", "Please select the old Excel file.")
            return False
        
        if not self.new_file_path.get():
            messagebox.showerror("Error", "Please select the new Excel file.")
            return False
        
        # Check if at least one sheet is selected
        selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
        if not selected_sheets:
            messagebox.showerror("Error", "Please select at least one sheet to compare.")
            return False
        
        return True
    
    def _compare_and_update(self):
        try:
            self._update_status("Starting comparison...", 0)
            
            # Get input values
            old_file = self.old_file_path.get()
            new_file = self.new_file_path.get()
            header_row = self.header_row.get()
            
            # Track updated cells and rows for highlighting and popup
            updated_cells = {}  # Dictionary mapping sheet name to list of updated cells (row, col)
            updated_rows = {}  # Dictionary mapping sheet name to set of updated rows
            
            # Get selected sheets from checkbox variables
            selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
            if not selected_sheets:
                messagebox.showerror("Error", "No sheets selected for processing.")
                self._update_status("Ready", 0)
                return
            
            # We only use column mode now (removed row mode toggle)
            is_row_mode = False
            
            # Get filter values
            team_filters = [f.get().strip() for f in self.team_filters if f.get().strip()]
            app_filters = [f.get().strip() for f in self.app_name_filters if f.get().strip()]
            category_filters = [f.get().strip() for f in self.category_filters if f.get().strip()]
            
            # Force close any previously open workbooks
            self._ensure_workbooks_closed()
            
            # Load workbooks with appropriate data_only settings
            self._update_status("Loading workbooks...", 10)
            old_wb_raw = openpyxl.load_workbook(old_file, data_only=False)  # For preserving formulas
            old_wb_eval = openpyxl.load_workbook(old_file, data_only=True)  # For evaluating formulas
            new_wb = openpyxl.load_workbook(new_file, data_only=True)  # Always use evaluated values
            
            # Get formula relationships map if enabled
            formula_map = self.formula_relationships if self.formula_aware.get() else {}
            
            total_updates = 0
            sheets_processed = 0
            
            for sheet_name in selected_sheets:
                # Initialize tracking for this sheet
                updated_cells[sheet_name] = []
                updated_rows[sheet_name] = set()
                
                self._update_status(f"Processing sheet: {sheet_name}...", 
                                20 + (sheets_processed / len(selected_sheets) * 60))
                
                # Get sheet objects - both raw and evaluated versions
                old_sheet_raw = old_wb_raw[sheet_name]  # Contains formulas
                old_sheet_eval = old_wb_eval[sheet_name]  # Contains formula results
                new_sheet = new_wb[sheet_name]
                
                # COLUMN-BASED MODE
                # Column-based comparison
                team_col = self.team_column.get()
                app_name_col = self.app_name_column.get()
                category_col = self.category_column.get()
                
                # Create a mapping of column names to column indices
                headers = {}
                col_to_name = {}  # Reverse mapping
                
                for col in range(1, old_sheet_raw.max_column + 1):
                    cell_value = old_sheet_raw.cell(row=header_row, column=col).value
                    if cell_value:
                        headers[cell_value] = col
                        col_to_name[col] = cell_value
                
                # Map formula columns to their source columns for key generation
                source_team_col = formula_map.get(team_col, team_col)
                source_app_col = formula_map.get(app_name_col, app_name_col)
                source_cat_col = formula_map.get(category_col, category_col)
                
                # Set up the team/app/cat column indexes
                if team_col in headers:
                    team_idx = headers[team_col]
                else:
                    messagebox.showerror("Error", f"Team column '{team_col}' not found in headers.")
                    continue
                    
                if app_name_col in headers:
                    app_idx = headers[app_name_col]
                else:
                    messagebox.showerror("Error", f"App name column '{app_name_col}' not found in headers.")
                    continue
                    
                if category_col in headers:
                    cat_idx = headers[category_col]
                else:
                    messagebox.showerror("Error", f"Category column '{category_col}' not found in headers.")
                    continue
                
                # Function to normalize cell value for consistent key generation
                def normalize_value(value):
                    if value is None:
                        return ""
                    # Convert to string and strip
                    return str(value).strip()
                
                # Create keys and maps based on the pattern in the Excel file
                old_keys = {}  # Maps key to row number
                new_keys = {}
                
                # Process additional criteria columns
                additional_col_indices = []
                for criteria_var, criteria_label, _ in self.additional_criteria:
                    col_name = criteria_var.get()
                    if col_name in headers:
                        additional_col_indices.append((criteria_label, headers[col_name]))
                    else:
                        if col_name:  # Only show error if a column was selected
                            messagebox.showerror("Error", f"Additional criteria column '{col_name}' not found in headers.")
                
                # For old file - use evaluated values (formula results)
                print(f"Processing {old_sheet_eval.max_row} rows in old sheet")
                for row in range(header_row + 1, old_sheet_eval.max_row + 1):
                    team_value = normalize_value(old_sheet_eval.cell(row=row, column=team_idx).value)
                    app_value = normalize_value(old_sheet_eval.cell(row=row, column=app_idx).value)
                    cat_value = normalize_value(old_sheet_eval.cell(row=row, column=cat_idx).value)
                    
                    # Add values for additional criteria
                    additional_values = []
                    for _, col_idx in additional_col_indices:
                        additional_values.append(normalize_value(old_sheet_eval.cell(row=row, column=col_idx).value))
                    
                    # Create key with all values even if empty
                    key_parts = [team_value, app_value, cat_value] + additional_values
                    key = "|".join(key_parts)
                    
                    # Only add non-empty keys (at least one component must be non-empty)
                    if any(key_parts):
                        old_keys[key] = row
                
                # For new file - use evaluated values
                print(f"Processing {new_sheet.max_row} rows in new sheet")
                for row in range(header_row + 1, new_sheet.max_row + 1):
                    team_value = normalize_value(new_sheet.cell(row=row, column=team_idx).value)
                    app_value = normalize_value(new_sheet.cell(row=row, column=app_idx).value)
                    cat_value = normalize_value(new_sheet.cell(row=row, column=cat_idx).value)
                    
                    # Add values for additional criteria
                    additional_values = []
                    for _, col_idx in additional_col_indices:
                        additional_values.append(normalize_value(new_sheet.cell(row=row, column=col_idx).value))
                    
                    # Create key with all values even if empty
                    key_parts = [team_value, app_value, cat_value] + additional_values
                    key = "|".join(key_parts)
                    
                    # Only add non-empty keys (at least one component must be non-empty)
                    if any(key_parts):
                        new_keys[key] = row
                
                # Find common keys
                all_common_keys = set(old_keys.keys()).intersection(set(new_keys.keys()))
                print(f"Found {len(old_keys)} keys in old file, {len(new_keys)} keys in new file")
                print(f"Common keys before filtering: {len(all_common_keys)}")
                
                # Apply filters if specified - use the EVALUATED values for filtering
                filtered_keys = set()

                # Check if we have any filters at all
                has_filters = (team_filters or app_filters or category_filters or 
                            any(filters for filters in self.additional_filters.values() if any(f.get().strip() for f in filters)))

                if has_filters:
                    for key in all_common_keys:
                        parts = key.split('|')
                        if len(parts) >= 3:  # We always have at least team, app, category
                            team, app, category = parts[0], parts[1], parts[2]
                            
                            # Additional parts correspond to additional criteria
                            additional_parts = parts[3:] if len(parts) > 3 else []
                            
                            # Check default filters
                            passes_team = True
                            passes_app = True
                            passes_category = True
                            
                            if team_filters:
                                passes_team = any(team.lower() == filter_val.lower() for filter_val in team_filters)
                            
                            if app_filters and passes_team:
                                passes_app = any(app.lower() == filter_val.lower() for filter_val in app_filters)
                            
                            if category_filters and passes_team and passes_app:
                                passes_category = any(category.lower() == filter_val.lower() for filter_val in category_filters)
                            
                            # Check additional filters if we have any
                            passes_additional = True
                            
                            # First check Team, App Name, Category additional filters
                            if "Team" in self.additional_filters and passes_team and passes_app and passes_category:
                                team_add_filters = [f.get().strip() for f in self.additional_filters["Team"] if f.get().strip()]
                                if team_add_filters:
                                    passes_additional = passes_additional and any(team.lower() == filter_val.lower() for filter_val in team_add_filters)
                            
                            if "App Name" in self.additional_filters and passes_team and passes_app and passes_category and passes_additional:
                                app_add_filters = [f.get().strip() for f in self.additional_filters["App Name"] if f.get().strip()]
                                if app_add_filters:
                                    passes_additional = passes_additional and any(app.lower() == filter_val.lower() for filter_val in app_add_filters)
                            
                            if "Category" in self.additional_filters and passes_team and passes_app and passes_category and passes_additional:
                                cat_add_filters = [f.get().strip() for f in self.additional_filters["Category"] if f.get().strip()]
                                if cat_add_filters:
                                    passes_additional = passes_additional and any(category.lower() == filter_val.lower() for filter_val in cat_add_filters)
                            
                            # Check other additional criteria filters
                            idx = 0
                            for criteria_var, criteria_label, _ in self.additional_criteria:
                                if (criteria_label in self.additional_filters and 
                                    passes_team and passes_app and passes_category and passes_additional and 
                                    idx < len(additional_parts)):
                                    
                                    add_filters = [f.get().strip() for f in self.additional_filters[criteria_label] if f.get().strip()]
                                    if add_filters:
                                        passes_additional = passes_additional and any(additional_parts[idx].lower() == filter_val.lower() for filter_val in add_filters)
                                
                                idx += 1
                            
                            # Add key if it passes all active filters
                            if passes_team and passes_app and passes_category and passes_additional:
                                filtered_keys.add(key)
                else:
                    filtered_keys = all_common_keys
                
                print(f"Keys after filtering: {len(filtered_keys)}")
                
                # Better detection of formula cells
                formula_cells = set()
                
                # First detect by data_type
                for row in range(header_row + 1, old_sheet_raw.max_row + 1):
                    for col in range(1, old_sheet_raw.max_column + 1):
                        cell = old_sheet_raw.cell(row=row, column=col)
                        if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                            formula_cells.add((row, col))
                
                # Create a set of formula columns to avoid
                formula_columns = set()
                for formula_col, src_col in formula_map.items():
                    if formula_col in headers:
                        formula_columns.add(headers[formula_col])
                        print(f"Excluding formula column: {formula_col} (column {headers[formula_col]})")
                
                # Process updates using the filtered keys
                updates_made = 0
                skipped_rows = 0
                skipped_formula = 0
                
                # For each matching key, update the cells in old sheet from new sheet
                for key in filtered_keys:
                    old_row = old_keys[key]
                    new_row = new_keys[key]
                    
                    cells_updated = False
                    
                    for col in range(1, min(old_sheet_raw.max_column, new_sheet.max_column) + 1):
                        # Skip header row
                        if old_row == header_row:
                            continue
                            
                        # Skip formula columns we identified
                        if col in formula_columns:
                            skipped_formula += 1
                            continue
                            
                        # Skip cells with formulas (additional check)
                        if (old_row, col) in formula_cells:
                            skipped_formula += 1
                            continue
                            
                        # Check if the source column has a name
                        if col in col_to_name:
                            # Double-check this isn't a formula cell
                            cell = old_sheet_raw.cell(row=old_row, column=col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                                skipped_formula += 1
                                continue
                                
                            # Get source and target cells
                            new_cell = new_sheet.cell(row=new_row, column=col)
                            old_cell = old_sheet_raw.cell(row=old_row, column=col)
                            
                            # Only update if values are different
                            if old_cell.value != new_cell.value:
                                # Store old value for reference
                                old_value = old_cell.value
                                
                                # Use helper method to update while preserving comments
                                self._update_cell_preserve_comments(new_cell, old_cell)
                                cells_updated = True
                                
                                # Track cell for highlighting and popup
                                updated_cells[sheet_name].append((old_row, col, old_value, new_cell.value))
                    
                    if cells_updated:
                        updates_made += 1
                        updated_rows[sheet_name].add(old_row)
                    else:
                        skipped_rows += 1
                
                total_updates += updates_made
                print(f"Updated {updates_made} rows, skipped {skipped_rows} rows, skipped {skipped_formula} formula cells")
                    
                sheets_processed += 1
            
            # Generate output filename based on selected save mode
            if self.save_mode.get() == "new":
                # Create new file with filter info
                base, ext = os.path.splitext(old_file)
                filter_info = ""
                if team_filters or app_filters or category_filters:
                    filter_parts = []
                    if team_filters:
                        filter_parts.append(f"Team-{'-'.join(team_filters)}")
                    if app_filters:
                        filter_parts.append(f"App-{'-'.join(app_filters)}")
                    if category_filters:
                        filter_parts.append(f"Cat-{'-'.join(category_filters)}")
                    filter_info = "_" + "_".join(filter_parts)
                    
                mode_info = "_row_based" if is_row_mode else ""
                output_file = f"{base}_updated{filter_info}{mode_info}{ext}"
            else:
                # Replace original file - but confirm first
                if not messagebox.askyesno("Confirm Replace", 
                    "Are you sure you want to overwrite the original file?\nThis cannot be undone.", 
                    icon="warning"):
                    self._update_status("Operation cancelled", 0)
                    return
                output_file = old_file
            
            # Save the updated workbook
            self._update_status("Saving updated workbook...", 90)
            old_wb_raw.save(output_file)
            
            # Create highlighted file if option is enabled
            if self.create_highlighted_file.get() and updated_cells:
                self._update_status("Creating highlighted changes file...", 95)
                highlighted_file = self._create_highlighted_file(old_file, output_file, updated_cells)
                if highlighted_file:
                    messagebox.showinfo("Highlighted File Created", f"A file with highlighted changes has been created at:\n{highlighted_file}")
            
            # Show popup with updated rows if option is enabled
            if self.show_update_popup.get() and any(rows for rows in updated_rows.values()):
                self._show_update_index_popup(updated_rows, header_row)
            
            # Make sure to close all workbooks properly
            old_wb_raw.close()
            old_wb_eval.close()
            new_wb.close()
            
            # If replacing the original file, ensure the file is properly released
            if self.save_mode.get() == "replace":
                # Force Python's garbage collection to ensure file handles are released
                import gc
                gc.collect()
                
                # Reset the formula relationships to ensure they're redetected
                if self.formula_aware.get():
                    self.formula_relationships = {}
            
            self._update_status("Complete!", 100)
            
            # Show success message with details
            filter_message = ""
            if team_filters or app_filters or category_filters:
                filter_message = "\n\nFilters applied:"
                if team_filters:
                    filter_message += f"\n- Teams: {', '.join(team_filters)}"
                if app_filters:
                    filter_message += f"\n- App Names: {', '.join(app_filters)}"
                if category_filters:
                    filter_message += f"\n- Categories: {', '.join(category_filters)}"
            
            sheet_message = f"\nProcessed sheets: {', '.join(selected_sheets)}"
            mode_message = "\nComparison mode: Row-based" if is_row_mode else "\nComparison mode: Column-based"
            formula_message = "\nFormula columns preserved" if formula_map else ""
            
            messagebox.showinfo("Success", f"Updated {total_updates} {'columns' if is_row_mode else 'rows'} successfully!\nSaved to: {output_file}{sheet_message}{mode_message}{formula_message}{filter_message}")
            
            # Clear file selections if option is enabled
            if self.clear_after_update.get():
                # Reset file paths
                self.old_file_path.set("")
                self.new_file_path.set("")
                
                # Clear sheet selections
                for var in self.sheet_vars.values():
                    var.set(False)
                
                # Clear sheet variables
                self.sheet_vars.clear()
                
                # Clear column selections
                self.team_column.set("")
                self.app_name_column.set("")
                self.category_column.set("")
                
                # Reset filters
                for filter_var in self.team_filters:
                    filter_var.set("")
                for filter_var in self.app_name_filters:
                    filter_var.set("")
                for filter_var in self.category_filters:
                    filter_var.set("")
            
            # Reset UI elements
            self._update_status("Ready", 0)
            
        except Exception as e:
            self._update_status("Error occurred", 0)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            import traceback
            traceback.print_exc()

    def _update_sheet(self, sheet, old_df, new_df, common_keys, team_col, app_name_col, category_col, formula_map=None):
        # Use the configured header row
        header_row = self.header_row.get()
        formula_map = formula_map or {}
        
        # Create a mapping of column names to column indices
        headers = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value:
                headers[col] = cell_value
        
        # Create a dictionary to store the column indices for faster lookup
        columns_indices = {col: headers[col] for col in new_df.columns if col in headers}
        
        # Create a set of columns that should NOT be updated (formula destination columns)
        formula_columns = set()
        for formula_col, source_col in formula_map.items():
            if formula_col in headers:
                formula_columns.add(formula_col)
        
        # Track which cells contain formulas (as backup detection)
        formula_cells = set()
        for row in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.data_type == 'f':  # 'f' indicates formula
                    formula_cells.add((row, col))
        
        updates_made = 0
        
        # For each row in the sheet (starting after the header)
        for row in range(header_row + 1, sheet.max_row + 1):
            # Get the key for this row using source columns (not formula columns)
            try:
                # Map formula columns to their source columns for key generation
                actual_team_col = formula_map.get(team_col, team_col)
                actual_app_col = formula_map.get(app_name_col, app_name_col)
                actual_cat_col = formula_map.get(category_col, category_col)
                
                if actual_team_col in headers and actual_app_col in headers and actual_cat_col in headers:
                    team_idx = headers[actual_team_col]
                    app_idx = headers[actual_app_col]
                    cat_idx = headers[actual_cat_col]
                    
                    team_value = str(sheet.cell(row=row, column=team_idx).value or "")
                    app_value = str(sheet.cell(row=row, column=app_idx).value or "")
                    cat_value = str(sheet.cell(row=row, column=cat_idx).value or "")
                    row_key = f"{team_value}|{app_value}|{cat_value}"
                    
                    # If this row has a match in the new data and passes our filters
                    if row_key in common_keys:
                        # Get the new data for this key
                        new_row_data = new_df[new_df['_key'] == row_key].iloc[0]
                        
                        # Update all applicable columns in the old sheet with values from the new data
                        for col_name, col_idx in columns_indices.items():
                            # Skip formula columns, artificial key column, and any cell with a formula
                            if (col_name != '_key' and 
                                col_name not in formula_columns and 
                                (row, col_idx) not in formula_cells):
                                
                                # Double-check this isn't a formula cell before updating
                                cell = sheet.cell(row=row, column=col_idx)
                                if cell.data_type != 'f':
                                    try:
                                        # Update with new value
                                        sheet.cell(row=row, column=col_idx).value = new_row_data[col_name]
                                    except Exception as cell_err:
                                        print(f"Error updating cell at row {row, col_idx}: {str(cell_err)}")
                        
                        updates_made += 1
                else:
                    print(f"Warning: One or more selected columns not found in headers: {actual_team_col, actual_app_col, actual_cat_col}")
                    
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                continue
                    
        return updates_made

    def _load_columns(self):
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Read the Excel file skipping to the header row
            df = pd.read_excel(old_file, sheet_name=selected_sheet, header=header_row)
            columns = list(df.columns)
            
            if not columns:
                messagebox.showerror("Error", "No columns found in the selected sheet.")
                return
            
            # Update all three comboboxes with the column list
            self.team_combobox['values'] = columns
            self.app_name_combobox['values'] = columns
            self.category_combobox['values'] = columns
            
            # Update any additional criteria comboboxes
            for criteria_var, _, _ in self.additional_criteria:
                for child in self.additional_criteria_frame.winfo_children():
                    for widget in child.winfo_children():
                        if isinstance(widget, ttk.Combobox) and widget['textvariable'] == str(criteria_var):
                            widget['values'] = columns
            
            # Set default selections if available
            for column in columns:
                if "team" in str(column).lower():
                    self.team_column.set(column)
                elif "app" in str(column).lower() or "name" in str(column).lower():
                    self.app_name_column.set(column)
                elif "category" in str(column).lower() or "test" in str(column).lower():
                    self.category_column.set(column)
            
            messagebox.showinfo("Success", f"Loaded {len(columns)} columns from sheet '{selected_sheet}'.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load columns: {str(e)}")

    def _bind_mousewheel(self, event):
        # Bind mousewheel scrolling when mouse enters the canvas
        self.main_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        # For Linux, bind Button-4 and Button-5
        self.main_canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.main_canvas.bind_all("<Button-5>", self._on_mousewheel)

    def _unbind_mousewheel(self, event):
        # Unbind mousewheel scrolling when mouse leaves the canvas
        self.main_canvas.unbind_all("<MouseWheel>")
        self.main_canvas.unbind_all("<Button-4>")
        self.main_canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        # Windows and macOS
        if event.num == 5 or event.delta < 0:
            self.main_canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.main_canvas.yview_scroll(-1, "units")

    def _detect_formula_relationships(self, sheet, header_row):
        """Detect formula relationships between columns in the sheet."""
        import re
        relationships = {}
        
        print("\n--- Starting Formula Relationship Detection ---")
        
        # Check if we're working with a read-only worksheet
        is_readonly = getattr(sheet, 'read_only', False)
        if is_readonly:
            print("Warning: Formula detection limited on read-only worksheets")
            return relationships
        
        try:
            # Step 1: Create a mapping from column positions to header names
            headers = {}  # Column index to header name
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value:
                    headers[col] = cell_value
            
            # Create reverse mapping from column letter to header name
            letter_to_header = {}
            for col_idx, header in headers.items():
                col_letter = get_column_letter(col_idx)
                letter_to_header[col_letter] = header
            
            # Step 2: Analyze formula patterns in the first few rows
            formula_patterns = {}  # To track potential formula columns
            
            # Look at several rows to establish consistent patterns
            for row_num in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
                for col_idx, header_name in headers.items():
                    try:
                        cell = sheet.cell(row=row_num, column=col_idx)
                        cell_formula = None
                        
                        # Get formula using different methods
                        if getattr(cell, 'data_type', None) == 'f':
                            cell_formula = str(cell.value)
                        elif hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                            cell_formula = cell.value
                        
                        if cell_formula:
                            # Look for simple cell references like =A5, =$A5, =A$5, or =$A$5
                            # The regex captures just the column letter part
                            match = re.search(r'=\$?([A-Za-z]+)\$?\d+', cell_formula)
                            if match:
                                ref_col_letter = match.group(1).upper()  # Ensure uppercase
                                
                                # If we can map this reference to a header
                                if ref_col_letter in letter_to_header:
                                    source_header = letter_to_header[ref_col_letter]
                                    
                                    # Add to our tracking dictionary
                                    if header_name not in formula_patterns:
                                        formula_patterns[header_name] = {}
                                    
                                    if source_header not in formula_patterns[header_name]:
                                        formula_patterns[header_name][source_header] = 0
                                    
                                    formula_patterns[header_name][source_header] += 1
                    except Exception as cell_error:
                        print(f"Error analyzing cell at row {row_num, col_idx}: {cell_error}")
                        continue
            
            # Step 3: Confirm relationships (if a column consistently references another column)
            for formula_col, references in formula_patterns.items():
                if references:
                    # Find the most commonly referenced source column
                    most_common = max(references.items(), key=lambda x: x[1])
                    source_col, count = most_common
                    
                    # If we have at least 2 occurrences, establish a relationship
                    if count >= 2:
                        relationships[formula_col] = source_col
                        print(f"Established relationship: '{formula_col}' references '{source_col}' ({count} occurrences)")
            
        except Exception as e:
            print(f"Error in formula detection: {e}")
            import traceback
            traceback.print_exc()
        
        return relationships

    def _ensure_workbooks_closed(self):
        """Force closure of any open workbooks to avoid file locking issues"""
        try:
            # Force garbage collection to release file handles
            import gc
            gc.collect()
            
            # Reset openpyxl's _archive cache if accessible
            if hasattr(openpyxl, '_archive'):
                openpyxl._archive = {}
        except:
            pass

    def _update_cell_preserve_comments(self, source_cell, target_cell):
        """Update a cell's value and handle comments appropriately"""
        # Update the cell value
        target_cell.value = source_cell.value
        
        # Handle comments
        if hasattr(source_cell, 'comment') and source_cell.comment is not None:
            # Direct copy of the comment object from source
            try:
                import copy
                # Try to create a deep copy of the comment
                target_cell.comment = copy.copy(source_cell.comment)
            except Exception as e:
                print(f"Error copying comment: {e}")

    def _add_filter_criteria(self):
        """Add a new filter criteria based on the comparison criteria"""
        # Show warning on first click
        if self.first_filter_click:
            # Show confirmation dialog with custom title, message, and warning icon
            proceed = messagebox.askokcancel(
                title="Add Filter Criteria",
                message="This will create additional filter criteria based on your selected columns.\n\n"
                       "If an additional filter has already been created, the first one created cannot be deleted.\n\n"
                       "To delete it, the program must be closed and reopened.\n\n"
                       "Would you like to continue?",
                icon="warning"
            )
            
            # Set the flag to False so this warning doesn't show again
            self.first_filter_click = False
            
            # If user clicked Cancel, exit the method
            if not proceed:
                return
        
        # Find an available criteria column
        available_criteria = []
        
        # Add the default criteria columns
        if self.team_column.get():
            available_criteria.append(("Team", self.team_column.get()))
        if self.app_name_column.get():
            available_criteria.append(("App Name", self.app_name_column.get()))
        if self.category_column.get():
            available_criteria.append(("Category", self.category_column.get()))
        
        # Add additional criteria columns
        for criteria_var, criteria_label, _ in self.additional_criteria:
            col_name = criteria_var.get()
            if col_name:
                available_criteria.append((criteria_label, col_name))
        
        # If no criteria columns are available, show error
        if not available_criteria:
            messagebox.showerror("Error", "No comparison criteria columns defined.\nPlease add comparison criteria first.")
            return
        
        # Create selection dialog
        criteria_dialog = tk.Toplevel(self.root)
        criteria_dialog.title("Select Filter Criteria")
        criteria_dialog.geometry("400x300")
        criteria_dialog.transient(self.root)
        criteria_dialog.grab_set()
        
        ttk.Label(
            criteria_dialog,
            text="Select column to create filters for:",
            font=("", 10, "bold")
        ).pack(padx=10, pady=10)
        
        # Create listbox for selection
        frame = ttk.Frame(criteria_dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=40, height=10)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox with available criteria
        for label, column in available_criteria:
            listbox.insert(tk.END, f"{label}: {column}")
        
        # Function to create the filter criteria
        def create_filter_criteria():
            try:
                selected_idx = listbox.curselection()
                if not selected_idx:
                    messagebox.showerror("Error", "Please select a criteria column.")
                    return
                
                selected = available_criteria[selected_idx[0]]
                label, column = selected
                
                # Check if this criteria already has filters
                if label in self.additional_filter_frames:
                    messagebox.showerror("Error", f"Filters for {label} already exist.")
                    criteria_dialog.destroy()
                    return
                
                # Create a new filter frame for this criteria
                filter_frame = ttk.LabelFrame(self.additional_filters_container, text=f"{label} Filters")
                filter_frame.pack(fill=tk.X, padx=5, pady=5)
                
                # Store the frame reference
                self.additional_filter_frames[label] = filter_frame
                
                # Initialize with empty filter
                self.additional_filters[label] = [tk.StringVar()]
                
                # Refresh the filters UI
                self._refresh_filter_widgets()
                
                criteria_dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create filter: {str(e)}")
                import traceback
                traceback.print_exc()  # Print detailed error for debugging
                criteria_dialog.destroy()
        
        # Add buttons
        button_frame = ttk.Frame(criteria_dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(
            button_frame, 
            text="Create Filter", 
            command=create_filter_criteria
        ).pack(side=tk.RIGHT, padx=5)
        
        ttk.Button(
            button_frame,
            text="Cancel",
            command=criteria_dialog.destroy
        ).pack(side=tk.RIGHT, padx=5)

    def _add_additional_filter(self, criteria_label):
        """Add a new filter for an additional criteria"""
        if criteria_label in self.additional_filters:
            self.additional_filters[criteria_label].append(tk.StringVar())
            self._refresh_filter_widgets()

    def _get_additional_unique_values(self, criteria_label, index):
        """Get unique values for additional filter criteria"""
        # Find the matching comparison criteria column name
        column_name = None
        
        # Check default criteria
        if criteria_label == "Team":
            column_name = self.team_column.get()
        elif criteria_label == "App Name":
            column_name = self.app_name_column.get()
        elif criteria_label == "Category":
            column_name = self.category_column.get()
        else:
            # Check additional criteria
            for criteria_var, label, _ in self.additional_criteria:
                if label == criteria_label and criteria_var.get():
                    column_name = criteria_var.get()
                    break
        
        if not column_name:
            messagebox.showerror("Error", f"Could not find column for {criteria_label}")
            return
        
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        # Get unique values from the column
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Use openpyxl to read values
            old_file = self.old_file_path.get()
            wb = openpyxl.load_workbook(old_file, data_only=True)
            sheet = wb[selected_sheet]
            
            # Find the column index
            col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row+1, column=col).value
                if cell_value == column_name:
                    col_idx = col
                    break
            
            if col_idx is None:
                messagebox.showerror("Error", f"Could not find column {column_name} in sheet.")
                return
                
            # Collect unique values
            unique_values = set()
            for row in range(header_row+2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:  # Only add non-empty values
                    unique_values.add(str(cell_value))
                    
            # Sort the unique values
            unique_values = sorted(list(unique_values))
            
            # Create selection dialog
            value_dialog = tk.Toplevel(self.root)
            value_dialog.title(f"Select {criteria_label} Value")
            value_dialog.geometry("400x300")
            value_dialog.transient(self.root)
            value_dialog.grab_set()
            
            # Create a frame with a scrollbar
            frame = ttk.Frame(value_dialog)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Create a scrollable listbox
            scrollbar = ttk.Scrollbar(frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=50, height=15)
            listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            scrollbar.config(command=listbox.yview)
            
            # Add search functionality
            search_frame = ttk.Frame(value_dialog)
            search_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
            
            ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
            search_var = tk.StringVar()
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
            search_entry.pack(side=tk.LEFT, padx=5)
            
            def filter_values(*args):
                search_term = search_var.get().lower()
                listbox.delete(0, tk.END)
                for value in unique_values:
                    if search_term in str(value).lower():
                        listbox.insert(tk.END, str(value))
                        
            search_var.trace_add("write", filter_values)
            
            # Populate the listbox initially
            for value in unique_values:
                listbox.insert(tk.END, str(value))
            
            # Function to set the selected value
            def set_selected_value():
                try:
                    selected_indices = listbox.curselection()
                    if selected_indices:
                        selected_value = listbox.get(selected_indices[0])
                        target_var = self.additional_filters[criteria_label][index]
                        target_var.set(selected_value)
                    value_dialog.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to set value: {str(e)}")
                    value_dialog.destroy()
            
            # Add buttons
            button_frame = ttk.Frame(value_dialog)
            button_frame.pack(fill=tk.X, padx=10, pady=10)
            
            ttk.Button(button_frame, text="Select", command=set_selected_value).pack(side=tk.RIGHT, padx=5)
            ttk.Button(button_frame, text="Cancel", command=value_dialog.destroy).pack(side=tk.RIGHT, padx=5)
            
            # Focus on search box
            search_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get unique values: {str(e)}")
            import traceback
            traceback.print_exc()

    def _create_highlighted_file(self, original_file, updated_file, updated_cells):
        """Create a copy of the updated file with highlighted changes"""
        try:
            from openpyxl.styles import PatternFill
            
            # Create a filename for the highlighted file
            base, ext = os.path.splitext(updated_file)
            highlighted_file = f"{base}_highlighted{ext}"
            
            # Load the updated workbook
            wb = openpyxl.load_workbook(updated_file)
            
            # Yellow fill for highlighting changes
            highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Apply highlighting to all updated cells
            for sheet_name, cells in updated_cells.items():
                if sheet_name in wb.sheetnames and cells:
                    ws = wb[sheet_name]
                    
                    for row, col, old_val, new_val in cells:
                        cell = ws.cell(row=row, column=col)
                        cell.fill = highlight_fill
                        
                        # Optionally add a comment with the old value
                        from openpyxl.comments import Comment
                        comment = Comment(f"Previous value: {old_val}", "Excel Compare Tool")
                        cell.comment = comment
            
            # Save the highlighted workbook
            wb.save(highlighted_file)
            return highlighted_file
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create highlighted file: {str(e)}")
            return None

    def _show_update_index_popup(self, updated_rows, header_row):
        """Show a popup with the indexes of updated data"""
        try:
            # Create a dialog to show the updated rows
            update_dialog = tk.Toplevel(self.root)
            update_dialog.title("Updated Data Indexes")
            update_dialog.geometry("600x400")
            update_dialog.transient(self.root)
            update_dialog.grab_set()
            
            # Add a label at the top
            ttk.Label(
                update_dialog, 
                text="The following rows were updated:", 
                font=("", 12, "bold")
            ).pack(padx=10, pady=10)
            
            # Create a frame with scrollbar for the updates list
            frame = ttk.Frame(update_dialog)
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            scrollbar = ttk.Scrollbar(frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Use a Text widget instead of Listbox for more formatting options
            updates_text = tk.Text(frame, yscrollcommand=scrollbar.set)
            updates_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=updates_text.yview)
            
            # Format and add the updates information
            updates_text.tag_configure("header", font=("", 11, "bold"), foreground="blue")
            updates_text.tag_configure("row", font=("", 10))
            
            # Add the updates for each sheet
            for sheet_name, rows in updated_rows.items():
                if rows:
                    updates_text.insert(tk.END, f"Sheet: {sheet_name}\n", "header")
                    
                    # Convert row numbers to real rows (accounting for header row)
                    sorted_rows = sorted(rows)
                    
                    # Group consecutive rows for better readability
                    row_groups = []
                    current_group = [sorted_rows[0]]
                    
                    for row in sorted_rows[1:]:
                        if row == current_group[-1] + 1:
                            current_group.append(row)
                        else:
                            row_groups.append(current_group)
                            current_group = [row]
                    
                    # Add the last group
                    row_groups.append(current_group)
                    
                    # Display the groups
                    for group in row_groups:
                        if len(group) == 1:
                            updates_text.insert(tk.END, f"  Row {group[0]} (Excel row {group[0]})\n", "row")
                        else:
                            updates_text.insert(tk.END, f"  Rows {group[0]}-{group[-1]} (Excel rows {group[0]}-{group[-1]})\n", "row")
                    
                    updates_text.insert(tk.END, "\n")
            
            # Set the text widget to read-only
            updates_text.config(state=tk.DISABLED)
            
            # Add a close button
            ttk.Button(
                update_dialog, 
                text="Close", 
                command=update_dialog.destroy
            ).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to show update popup: {str(e)}")

    def _delete_team_filter(self, index):
        """Delete a team filter at the specified index"""
        # Ensure we always keep at least one filter
        if len(self.team_filters) <= 1:
            messagebox.showwarning("Warning", "Cannot delete the last filter entry.")
            return
        
        # Remove the filter at the specified index
        self.team_filters.pop(index)
        
        # Refresh the filter widgets
        self._refresh_filter_widgets()

    def _delete_app_filter(self, index):
        """Delete an app filter at the specified index"""
        # Ensure we always keep at least one filter
        if len(self.app_name_filters) <= 1:
            messagebox.showwarning("Warning", "Cannot delete the last filter entry.")
            return
        
        # Remove the filter at the specified index
        self.app_name_filters.pop(index)
        
        # Refresh the filter widgets
        self._refresh_filter_widgets()

    def _delete_category_filter(self, index):
        """Delete a category filter at the specified index"""
        # Ensure we always keep at least one filter
        if len(self.category_filters) <= 1:
            messagebox.showwarning("Warning", "Cannot delete the last filter entry.")
            return
        
        # Remove the filter at the specified index
        self.category_filters.pop(index)
        
        # Refresh the filter widgets
        self._refresh_filter_widgets()

    def _delete_additional_filter(self, criteria_label, index):
        """Delete an additional filter at the specified index"""
        filters = self.additional_filters.get(criteria_label, [])
        
        # Ensure we always keep at least one filter
        if len(filters) <= 1:
            messagebox.showwarning("Warning", "Cannot delete the last filter entry.")
            return
        
        # Remove the filter at the specified index
        filters.pop(index)
        
        # Refresh the filter widgets
        self._refresh_filter_widgets()
    
    def _show_help_window(self, section):
        """Display help information for the specified section"""
        from help_utils import show_help_window
        show_help_window(self.root, section)

    def _create_custom_criteria_widgets(self, parent):
        # Main container
        container = ttk.Frame(parent)
        container.pack(fill=tk.X, pady=(0, 20))
        
        # Header section
        header_frame = ttk.Frame(container)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT)
        
        title_label = ttk.Label(
            title_frame,
            text="🎯 Custom Comparison Criteria", 
            font=("Segoe UI", 12, "bold"),
            foreground="#323130"
        )
        title_label.pack(side=tk.LEFT)
        
        # Help button
        help_btn = ttk.Button(
            header_frame,
            text="❓",
            style="Icon.TButton",
            width=3,
            command=lambda: self._show_help_window("custom_comparison_criteria")
        )
        help_btn.pack(side=tk.RIGHT)
        
        # Description
        desc_label = ttk.Label(
            container,
            text="Define custom key columns for flexible row matching",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Configuration section
        config_frame = ttk.LabelFrame(
            container,
            text="⚙️ Configuration",
            padding=15,
            style="Card.TLabelframe"
        )
        config_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Settings row
        settings_frame = ttk.Frame(config_frame)
        settings_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Header row
        header_row_frame = ttk.Frame(settings_frame)
        header_row_frame.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(
            header_row_frame,
            text="📍 Header Row:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        header_spin = ttk.Spinbox(
            header_row_frame,
            from_=1,
            to=20,
            textvariable=self.header_row,
            width=5,
            font=("Segoe UI", 9)
        )
        header_spin.pack(side=tk.LEFT)
        
        # Formula awareness
        formula_frame = ttk.Frame(settings_frame)
        formula_frame.pack(side=tk.LEFT)
        
        ttk.Checkbutton(
            formula_frame,
            text="🧮 Formula-Aware Processing",
            variable=self.formula_aware,
            style="Modern.TCheckbutton"
        ).pack()
        
        # Key columns section
        key_columns_frame = ttk.LabelFrame(
            container,
            text="🔑 Key Columns for Matching",
            padding=15,
            style="Card.TLabelframe"
        )
        key_columns_frame.pack(fill=tk.X)
        
        # Instructions
        instruction_label = ttk.Label(
            key_columns_frame,
            text="Define which columns uniquely identify rows for matching:",
            font=("Segoe UI", 9),
            foreground="#605E5C"
        )
        instruction_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Action buttons row
        actions_frame = ttk.Frame(key_columns_frame)
        actions_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Button(
            actions_frame,
            text="➕ Add Key Column",
            style="Accent.TButton",
            command=self._add_custom_key_column
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            actions_frame,
            text="🔄 Load Columns",
            style="Secondary.TButton",
            command=self._load_custom_columns
        ).pack(side=tk.LEFT)
        
        # Key columns container
        self.custom_criteria_container = ttk.Frame(key_columns_frame)
        self.custom_criteria_container.pack(fill=tk.X, pady=(10, 0))

    def _add_custom_key_column(self):
        """Add a new key column for custom matching"""
        # Create a new key column
        key_var = tk.StringVar()
        
        # Create row index (number of existing key columns + 2 for the header rows)
        row_idx = len(self.key_columns) + 2
        
        # Create a frame for this key column
        column_frame = ttk.Frame(self.custom_criteria_container)
        column_frame.grid(row=row_idx, column=0, columnspan=3, sticky=tk.W+tk.E, pady=2)
        
        # Add label
        ttk.Label(
            column_frame, 
            text=f"Key Column {len(self.key_columns) + 1}:"
        ).pack(side=tk.LEFT, padx=5)
        
        # Add combobox
        combobox = ttk.Combobox(
            column_frame,
            textvariable=key_var,
            state="readonly",
            width=30
        )
        combobox.pack(side=tk.LEFT, padx=5)
        
        # Add remove button
        ttk.Button(
            column_frame,
            text="Remove",
            command=lambda: self._remove_custom_key_column(column_frame, len(self.key_columns))
        ).pack(side=tk.LEFT, padx=5)
        
        # Store the key column
        self.key_columns.append((key_var, column_frame))

    def _remove_custom_key_column(self, frame, idx):
        """Remove a custom key column"""
        # We should always keep at least one key column
        if len(self.key_columns) <= 1:
            messagebox.showwarning("Warning", "You must have at least one key column for matching.")
            return
        
        # Remove the key column
        self.key_columns.pop(idx)
        
        # Remove the frame
        frame.destroy()
        
        # Reindex the remaining key columns
        for i, (_, frame) in enumerate(self.key_columns):
            # Find and update the label
            for child in frame.winfo_children():
                if isinstance(child, ttk.Label):
                    child.config(text=f"Key Column {i + 1}:")
                    break

    def _load_custom_columns(self):
        """Load columns for custom mode"""
        # Similar to _load_columns but for custom mode
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Read the Excel file skipping to the header row
            df = pd.read_excel(old_file, sheet_name=selected_sheet, header=header_row)
            columns = list(df.columns)
            
            if not columns:
                messagebox.showerror("Error", "No columns found in the selected sheet.")
                return
            
            # Update all key column comboboxes
            for key_var, frame in self.key_columns:
                # Find the combobox in this frame
                for child in frame.winfo_children():
                    if isinstance(child, ttk.Combobox):
                        child['values'] = columns
                        # Set a default value if available
                        if columns:
                            child.current(0)
            
            # If we don't have any key columns yet, add one automatically
            if not self.key_columns:
                self._add_custom_key_column()
                # Find the combobox in the newly added frame
                key_var, frame = self.key_columns[0]
                for child in frame.winfo_children():
                    if isinstance(child, ttk.Combobox):
                        child['values'] = columns
                        # Set a default value if available
                        if columns:
                            child.current(0)
            
            messagebox.showinfo("Success", f"Loaded {len(columns)} columns from sheet '{selected_sheet}'.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load columns: {str(e)}")

    def _create_custom_filter_widgets(self, parent):
        """Create filter criteria widgets for custom mode"""
        # Create header with label and help button
        header_frame = ttk.Frame(parent)
        header_frame.pack(fill=tk.X, pady=(0, 0))
        
        # Add the section label and help button
        filter_label = ttk.Label(header_frame, text="Custom Filter Criteria", font=("", 10, "bold"))
        filter_label.pack(side=tk.LEFT, padx=5, pady=5)
        
        help_btn = ttk.Button(
            header_frame, 
            text="?", 
            width=2,
            command=lambda: self._show_help_window("custom_filter_criteria")
        )
        help_btn.pack(side=tk.LEFT, padx=2, pady=5)
        
        # Create the filter content frame
        filter_frame = ttk.LabelFrame(parent, text="(Leave empty to match all)", padding="10")
        filter_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Container for custom filters - will hold filter frames for each key column
        self.custom_filter_container = ttk.Frame(filter_frame)
        self.custom_filter_container.pack(fill=tk.X, pady=5)
        
        # Add filter button
        ttk.Button(
            filter_frame,
            text="+ Add Custom Filter",
            command=self._add_custom_filter,
            width=15
        ).pack(anchor=tk.W, padx=5, pady=5)

    def _add_custom_filter(self):
        """Add filter based on custom key columns"""
        # Ensure we have key columns first
        if not self.key_columns:
            messagebox.showerror("Error", "Please define key columns first using 'Load Columns'")
            return
        
        # Show filter selection dialog
        filter_dialog = tk.Toplevel(self.root)
        filter_dialog.title("Add Custom Filter")
        filter_dialog.geometry("400x300")
        filter_dialog.transient(self.root)
        filter_dialog.grab_set()
        
        # Create dialog content
        ttk.Label(
            filter_dialog,
            text="Select column to filter by:",
            font=("", 10, "bold")
        ).pack(padx=10, pady=10)
        
        # Create a listbox with all key columns
        frame = ttk.Frame(filter_dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Create a scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create listbox
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=40, height=10)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Populate listbox with available columns from key columns
        available_columns = []
        for key_var, _ in self.key_columns:
            column_name = key_var.get()
            if column_name:
                available_columns.append(column_name)
                listbox.insert(tk.END, column_name)
        
        # Function to create the filter
        def create_filter():
            try:
                selected_idx = listbox.curselection()
                if not selected_idx:
                    messagebox.showerror("Error", "Please select a column.")
                    return
                
                column_name = available_columns[selected_idx[0]]
                
                # Check if filter already exists
                if column_name in self.custom_filters:
                    messagebox.showerror("Error", f"Filter for '{column_name}' already exists.")
                    filter_dialog.destroy()
                    return
                
                # Create a new filter frame for this column
                filter_frame = ttk.LabelFrame(self.custom_filter_container, text=f"{column_name} Filters")
                filter_frame.pack(fill=tk.X, padx=5, pady=5)
                
                # Initialize with empty filter
                self.custom_filters[column_name] = {
                    'frame': filter_frame,
                    'filters': [tk.StringVar()],
                    'widgets': []
                }
                
                # Refresh the custom filters UI
                self._refresh_custom_filter_widgets(column_name)
                
                filter_dialog.destroy()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create filter: {str(e)}")
                import traceback
                traceback.print_exc()
                filter_dialog.destroy()
        
        # Add buttons
        button_frame = ttk.Frame(filter_dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(button_frame, text="Create Filter", command=create_filter).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=filter_dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def _refresh_custom_filter_widgets(self, column_name):
        """Refresh the filter widgets for a specific column"""
        if column_name not in self.custom_filters:
            return
            
        filter_data = self.custom_filters[column_name]
        filter_frame = filter_data['frame']
        
        # Clear existing widgets
        for widget in filter_frame.winfo_children():
            widget.destroy()
        filter_data['widgets'] = []
        
        # Create filter widgets horizontally
        for i, filter_var in enumerate(filter_data['filters']):
            # Create entry horizontally
            entry = ttk.Entry(filter_frame, textvariable=filter_var, width=20)
            entry.grid(row=0, column=i*2, padx=2, pady=5)
            
            # Button frame for icons
            btn_frame = ttk.Frame(filter_frame)
            btn_frame.grid(row=0, column=i*2+1, padx=2, pady=5)
            
            # Get values button
            get_btn = ttk.Button(
                btn_frame, 
                text="🔍", 
                width=2,
                command=lambda col=column_name, idx=i: self._get_custom_unique_values(col, idx)
            )
            get_btn.pack(fill="x", pady=1)
            
            # Delete button - only if we have more than one filter
            if len(filter_data['filters']) > 1:
                del_btn = ttk.Button(
                    btn_frame, 
                    text="❌", 
                    width=2,
                    command=lambda col=column_name, idx=i: self._delete_custom_filter(col, idx)
                )
                del_btn.pack(fill="x", pady=1)
            
            # Store widget references
            filter_data['widgets'].append((entry, btn_frame))
        
        # Add button at the end
        add_btn = ttk.Button(
            filter_frame, 
            text="+", 
            width=2,
            command=lambda col=column_name: self._add_custom_filter_value(col)
        )
        add_btn.grid(row=0, column=len(filter_data['filters'])*2, padx=2, pady=5)

    def _add_custom_filter_value(self, column_name):
        """Add a new filter value for a custom filter"""
        if column_name in self.custom_filters:
            self.custom_filters[column_name]['filters'].append(tk.StringVar())
            self._refresh_custom_filter_widgets(column_name)

    def _delete_custom_filter(self, column_name, index):
        """Delete a filter value from a custom filter"""
        if column_name not in self.custom_filters:
            return
            
        filter_data = self.custom_filters[column_name]
        
        # Ensure we always keep at least one filter
        if len(filter_data['filters']) <= 1:
            messagebox.showwarning("Warning", "Cannot delete the last filter entry.")
            return
        
        # Remove the filter
        filter_data['filters'].pop(index)
        
        # Refresh the widgets
        self._refresh_custom_filter_widgets(column_name)

    def _get_custom_unique_values(self, column_name, index):
        """Get unique values for a custom filter"""
        # Find the first selected sheet
        selected_sheet = None
        for sheet, var in self.sheet_vars.items():
            if var.get():
                selected_sheet = sheet
                break
        
        if not selected_sheet:
            messagebox.showerror("Error", "Please select at least one sheet.")
            return
        
        old_file = self.old_file_path.get()
        
        if not old_file:
            messagebox.showerror("Error", "Please select the old Excel file first.")
            return
        
        try:
            # Use the configured header row
            header_row = self.header_row.get() - 1  # Convert to 0-based for pandas
            
            # Use openpyxl to read values
            wb = openpyxl.load_workbook(old_file, data_only=True)
            sheet = wb[selected_sheet]
            
            # Find the column index
            col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row+1, column=col).value
                if str(cell_value) == column_name:
                    col_idx = col
                    break
            
            if col_idx is None:
                messagebox.showerror("Error", f"Could not find column '{column_name}' in sheet.")
                return
                
            # Collect unique non-empty values from this column
            unique_values = set()
            for row in range(header_row+2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:  # Only add non-empty values
                    unique_values.add(str(cell_value))
            
            # Sort the values
            unique_values = sorted(list(unique_values))
            
            # Create selection dialog with search
            self._show_value_selection_dialog(
                unique_values, 
                column_name, 
                self.custom_filters[column_name]['filters'][index]
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get unique values: {str(e)}")
            import traceback
            traceback.print_exc()

    def _show_value_selection_dialog(self, values, title, target_var):
        """Show a dialog for selecting from a list of values"""
        # Create a new dialog
        value_dialog = tk.Toplevel(self.root)
        value_dialog.title(f"Select {title} Value")
        value_dialog.geometry("400x300")
        value_dialog.transient(self.root)
        value_dialog.grab_set()
        
        # Create a frame with a scrollbar
        frame = ttk.Frame(value_dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create a scrollable listbox
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, width=50, height=15)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=listbox.yview)
        
        # Add search functionality
        search_frame = ttk.Frame(value_dialog)
        search_frame.pack(fill=tk.X, padx=10, pady=(10, 0))
        
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        def filter_values(*args):
            search_term = search_var.get().lower()
            listbox.delete(0, tk.END)
            for value in values:
                if search_term in str(value).lower():
                    listbox.insert(tk.END, str(value))
                    
        search_var.trace_add("write", filter_values)
        
        # Populate the listbox initially
        for value in values:
            listbox.insert(tk.END, str(value))
        
        # Function to set the selected value
        def set_selected_value():
            try:
                selected_indices = listbox.curselection()
                if selected_indices:
                    selected_value = listbox.get(selected_indices[0])
                    target_var.set(selected_value)
                value_dialog.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to set value: {str(e)}")
                value_dialog.destroy()
        
        # Add buttons
        button_frame = ttk.Frame(value_dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(button_frame, text="Select", command=set_selected_value).pack(side=tk.RIGHT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=value_dialog.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Focus on search box
        search_entry.focus_set()

    def _custom_compare_and_update(self):
        """Compare and update using custom key columns"""
        try:
            self.custom_status_var.set("Starting comparison...")
            self.custom_progress_var.set(0)
            self.root.update_idletasks()
            
            # Get input values
            old_file = self.old_file_path.get()
            new_file = self.new_file_path.get()
            header_row = self.header_row.get()
            
            # Track updated cells and rows
            updated_cells = {}
            updated_rows = {}
            
            # Get key columns
            key_columns = [key_var.get() for key_var, _ in self.key_columns]
            if not key_columns or not all(key_columns):
                messagebox.showerror("Error", "Please define at least one key column")
                self.custom_status_var.set("Ready")
                self.custom_progress_var.set(0)
                return
            
            # Debug confirmation of key columns
            print(f"Using key columns for matching: {key_columns}")
                
            # Get selected sheets
            selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
            if not selected_sheets:
                messagebox.showerror("Error", "No sheets selected for processing.")
                self.custom_status_var.set("Ready")
                self.custom_progress_var.set(0)
                return
            
            # Get filter values
            custom_filter_values = {}
            for column, filter_data in self.custom_filters.items():
                values = [f.get().strip() for f in filter_data['filters'] if f.get().strip()]
                if values:
                    custom_filter_values[column] = values
                    print(f"Filter for {column}: {values}")
            
            # Load workbooks
            self.custom_status_var.set("Loading workbooks...")
            self.custom_progress_var.set(10)
            self.root.update_idletasks()
            
            # Force closure of any open workbooks
            self._ensure_workbooks_closed()
            
            old_wb_raw = openpyxl.load_workbook(old_file, data_only=False)  # For preserving formulas
            old_wb_eval = openpyxl.load_workbook(old_file, data_only=True)  # For evaluating formulas
            new_wb = openpyxl.load_workbook(new_file, data_only=True)  # Always use evaluated values
            
            # Get formula relationships if enabled
            formula_map = self.formula_relationships if self.formula_aware.get() else {}
            
            total_updates = 0
            sheets_processed = 0
            
            for sheet_name in selected_sheets:
                # Initialize tracking for this sheet
                updated_cells[sheet_name] = []
                updated_rows[sheet_name] = set()
                
                self.custom_status_var.set(f"Processing sheet: {sheet_name}...")
                self.custom_progress_var.set(20 + (sheets_processed / len(selected_sheets) * 60))
                self.root.update_idletasks()
                
                # Get sheet objects
                try:
                    old_sheet_raw = old_wb_raw[sheet_name]  # Contains formulas
                    old_sheet_eval = old_wb_eval[sheet_name]  # Contains formula results
                    new_sheet = new_wb[sheet_name]
                except KeyError:
                    messagebox.showerror("Error", f"Sheet '{sheet_name}' not found in one of the workbooks.")
                    continue
                
                # Create column mappings from header row
                headers = {}
                col_to_name = {}
                
                for col in range(1, old_sheet_raw.max_column + 1):
                    cell_value = old_sheet_raw.cell(row=header_row, column=col).value
                    if cell_value:
                        headers[str(cell_value)] = col
                        col_to_name[col] = str(cell_value)
                
                # Print found headers for debugging
                print(f"Found {len(headers)} headers in sheet {sheet_name}")
                
                # Validate all key columns exist in the headers
                missing_columns = [col for col in key_columns if col not in headers]
                if missing_columns:
                    messagebox.showerror("Error", 
                                        f"Key column(s) not found in sheet '{sheet_name}': {', '.join(missing_columns)}")
                    continue
                
                # Function to normalize cell values for consistent comparison
                def normalize_value(value):
                    if value is None:
                        return ""
                    elif isinstance(value, (int, float)):
                        # Convert numbers to strings with no decimal for integers
                        if isinstance(value, int) or value.is_integer():
                            return str(int(value))
                        return str(value)
                    return str(value).strip()
                
                # Create keys and maps based on the pattern in the Excel file
                old_keys = {}  # Maps composite key to row number
                new_keys = {}
                
                # For old file - use evaluated values (formula results)
                print(f"Processing {old_sheet_eval.max_row - header_row} rows in old sheet")
                for row in range(header_row + 1, old_sheet_eval.max_row + 1):
                    # Build the composite key from all key columns
                    key_parts = []
                    all_empty = True
                    
                    for key_col in key_columns:
                        col_idx = headers[key_col]
                        value = normalize_value(old_sheet_eval.cell(row=row, column=col_idx).value)
                        key_parts.append(value)
                        if value:  # Check if any part has content
                            all_empty = False
                    
                    # Skip if all key parts are empty
                    if all_empty:
                        continue
                        
                    # Create composite key
                    key = "|".join(key_parts)
                    old_keys[key] = row
                    
                    # Debug print first few rows
                    if len(old_keys) <= 5:
                        print(f"Old sheet row {row} key: {key}")
                
                # For new file - use evaluated values
                print(f"Processing {new_sheet.max_row - header_row} rows in new sheet")
                for row in range(header_row + 1, new_sheet.max_row + 1):
                    # Build the composite key from all key columns
                    key_parts = []
                    all_empty = True
                    invalid_key = False
                    
                    for key_col in key_columns:
                        # Check if column exists in new file
                        if key_col not in headers:
                            invalid_key = True
                            break
                            
                        col_idx = headers[key_col]
                        value = normalize_value(new_sheet.cell(row=row, column=col_idx).value)
                        key_parts.append(value)
                        if value:  # Check if any part has content
                            all_empty = False
                    
                    # Skip if key is invalid or all parts are empty
                    if invalid_key or all_empty:
                        continue
                        
                    # Create composite key
                    key = "|".join(key_parts)
                    new_keys[key] = row
                    
                    # Debug print first few rows
                    if len(new_keys) <= 5:
                        print(f"New sheet row {row} key: {key}")
                
                # Find common keys
                all_common_keys = set(old_keys.keys()).intersection(set(new_keys.keys()))
                print(f"Found {len(old_keys)} keys in old file, {len(new_keys)} keys in new file")
                print(f"Common keys before filtering: {len(all_common_keys)}")
                
                # Apply filters if specified
                filtered_keys = set()
                
                if custom_filter_values:  # If we have filters
                    for key in all_common_keys:
                        # Split the key into its components
                        key_parts = key.split('|')
                        
                        # Check each filter
                        passes_all_filters = True
                        
                        for i, key_col in enumerate(key_columns):
                            if key_col in custom_filter_values and i < len(key_parts):
                                # Get the value from this key part
                                value = key_parts[i]
                                
                                # Check if it matches any of the filter values (case-insensitive)
                                matches_filter = False
                                for filter_val in custom_filter_values[key_col]:
                                    if value.lower() == filter_val.lower():
                                        matches_filter = True
                                        break
                                
                                if not matches_filter:
                                    passes_all_filters = False
                                    break
                        
                        if passes_all_filters:
                            filtered_keys.add(key)
                else:
                    # No filters, use all common keys
                    filtered_keys = all_common_keys
                
                print(f"Keys after filtering: {len(filtered_keys)}")
                
                # If no keys match after filtering, inform the user but continue with other sheets
                if not filtered_keys:
                    print(f"No matching rows found in sheet {sheet_name} after applying filters")
                    continue
                
                # Detect formula cells to preserve them
                formula_cells = set()
                
                # Check for formulas in cells up to 100 rows past header row
                max_formula_check_row = min(header_row + 100, old_sheet_raw.max_row + 1)
                
                # Find formula cells for preservation
                for row in range(header_row + 1, max_formula_check_row):
                    for col in range(1, old_sheet_raw.max_column + 1):
                        try:
                            cell = old_sheet_raw.cell(row=row, column=col)
                            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value and cell.value.startswith('=')):
                                formula_cells.add((row, col))
                        except Exception as e:
                            print(f"Error checking formula at ({row}, {col}): {e}")
                            continue
                
                print(f"Detected {len(formula_cells)} formula cells to preserve")
                
                # Create a set of formula columns to avoid
                formula_columns = set()
                for formula_col, src_col in formula_map.items():
                    if formula_col in headers:
                        formula_columns.add(headers[formula_col])
                        print(f"Will preserve formula column: {formula_col}")
                
                # Process updates
                updates_made = 0
                skipped_rows = 0
                skipped_formula = 0
                
                # For each matching key, update cells
                for key in filtered_keys:
                    old_row = old_keys[key]
                    new_row = new_keys[key]
                    
                    # Debug for first few updates
                    if updates_made < 3:
                        print(f"Comparing match: old row {old_row}, new row {new_row}, key {key}")
                    
                    cells_updated = False
                    
                    # Loop through all columns to update
                    for col in range(1, min(old_sheet_raw.max_column, new_sheet.max_column) + 1):
                        # Skip header row
                        if old_row <= header_row:
                            continue
                        
                        # Skip if column has no name or is not in headers
                        if col not in col_to_name:
                            continue
                        
                        column_name = col_to_name[col]
                        
                        # Skip our key columns - don't update the keys themselves
                        if column_name in key_columns:
                            continue
                            
                        # Skip formula columns we identified in the mapping
                        if col in formula_columns:
                            skipped_formula += 1
                            continue
                        
                        # Skip cells with formulas
                        if (old_row, col) in formula_cells:
                            skipped_formula += 1
                            continue
                            
                        # Get cells for comparison
                        try:
                            new_cell = new_sheet.cell(row=new_row, column=col)
                            old_cell = old_sheet_raw.cell(row=old_row, column=col)
                            
                            # Extra formula check (for safety)
                            if isinstance(old_cell.value, str) and old_cell.value and old_cell.value.startswith('='):
                                skipped_formula += 1
                                continue
                            
                            # Compare values, handling None, empty strings, and different types
                            old_value = old_cell.value
                            new_value = new_cell.value
                            
                            # Convert to strings for comparison
                            old_str = str(old_value) if old_value is not None else ""
                            new_str = str(new_value) if new_value is not None else ""
                            
                            if old_str != new_str:
                                # Debug for first few cell updates
                                if updates_made < 3 and cells_updated < 3:
                                    print(f"Updating cell ({old_row}, {col}) {column_name}: '{old_value}' -> '{new_value}'")
                                
                                # Update the cell value
                                old_cell.value = new_value
                                cells_updated = True
                                
                                # Track for highlighting and popup
                                updated_cells[sheet_name].append((old_row, col, old_value, new_value))
                                
                        except Exception as e:
                            print(f"Error updating cell ({old_row}, {col}): {e}")
                            continue
                    
                    if cells_updated:
                        updates_made += 1
                        updated_rows[sheet_name].add(old_row)
                    else:
                        skipped_rows += 1
                
                total_updates += updates_made
                print(f"Sheet {sheet_name}: Updated {updates_made} rows, skipped {skipped_rows} rows, skipped {skipped_formula} formula cells")
                sheets_processed += 1
            
            # Generate output filename based on selected save mode
            if self.save_mode.get() == "new":
                # Create new file with filter info
                base, ext = os.path.splitext(old_file)
                filter_info = "_custom"
                if custom_filter_values:
                    filter_info += "_filtered"
                output_file = f"{base}_updated{filter_info}{ext}"
            else:
                # Replace original file
                if not messagebox.askyesno("Confirm Replace", 
                    "Are you sure you want to overwrite the original file?\nThis cannot be undone.", 
                    icon="warning"):
                    self.custom_status_var.set("Operation cancelled")
                    return
                output_file = old_file
            
            # Save the updated workbook
            self.custom_status_var.set("Saving updated workbook...")
            self.custom_progress_var.set(90)
            self.root.update_idletasks()
            
            try:
                old_wb_raw.save(output_file)
                print(f"Successfully saved to {output_file}")
            except Exception as save_error:
                messagebox.showerror("Save Error", f"Failed to save workbook: {str(save_error)}")
                self.custom_status_var.set("Save failed")
                return
            
            # Create highlighted file if option is enabled
            if self.create_highlighted_file.get() and any(cells for cells in updated_cells.values()):
                self.custom_status_var.set("Creating highlighted changes file...")
                self.custom_progress_var.set(95)
                self.root.update_idletasks()
                highlighted_file = self._create_highlighted_file(old_file, output_file, updated_cells)
                if highlighted_file:
                    messagebox.showinfo("Highlighted File Created", 
                                       f"A file with highlighted changes has been created at:\n{highlighted_file}")
            
            # Show popup with updated rows if option is enabled
            if self.show_update_popup.get() and any(rows for rows in updated_rows.values()):
                self._show_update_index_popup(updated_rows, header_row)
            
            # Close all workbooks
            old_wb_raw.close()
            old_wb_eval.close()
            new_wb.close()
            
            # Reset if needed
            if self.clear_after_update.get():
                self.old_file_path.set("")
                self.new_file_path.set("")
                self.sheet_vars.clear()
            
            self.custom_status_var.set("Complete!")
            self.custom_progress_var.set(100)
            
            if total_updates == 0:
                messagebox.showinfo("No Updates", 
                                  f"No rows were updated. This could be because:\n"
                                  f"1. No matching rows were found based on your key columns\n"
                                  f"2. The filter criteria excluded all matches\n"
                                  f"3. No data differences were detected\n\n"
                                  f"Key columns used: {', '.join(key_columns)}\n" +
                                  (f"Filters applied: {', '.join(f'{col}: {vals}' for col, vals in custom_filter_values.items())}" 
                                   if custom_filter_values else "No filters applied"))
            else:
                # Show success message
                key_message = f"Key columns used: {', '.join(key_columns)}"
                filter_message = ""
                if custom_filter_values:
                    filter_message = "\n\nFilters applied:"
                    for column, values in custom_filter_values.items():
                        filter_message += f"\n- {column}: {', '.join(values)}"
                
                sheet_message = f"\nProcessed sheets: {', '.join(selected_sheets)}"
                
                messagebox.showinfo("Success", 
                                   f"Updated {total_updates} rows successfully!\n"
                                   f"Saved to: {output_file}\n"
                                   f"{key_message}{sheet_message}{filter_message}")
            
        except Exception as e:
            self.custom_status_var.set("Error occurred")
            self.custom_progress_var.set(0)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            import traceback
            traceback.print_exc()

    def _configure_button_styles(self):
        """Configure custom button styles for better UX"""
        style = ttk.Style()
        
        # Configure prominent button style
        style.configure(
            "Accent.TButton",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130",
            background="#0078D4",
            borderwidth=1,
            focuscolor='none'
        )
        
        style.map("Accent.TButton",
            background=[('active', '#106EBE'), ('pressed', '#005A9E')],
            foreground=[('active', '#323130'), ('pressed', '#323130')])
        
        # Configure success button style  
        style.configure(
            "Success.TButton",
            font=("Segoe UI", 9, "bold"),
            foreground="#FFFFFF",
            background="#107C10",
            borderwidth=1,
            focuscolor='none'
        )
        
        style.map("Success.TButton",
            background=[('active', '#0E6A0E'), ('pressed', '#0C5D0C')],
            foreground=[('active', '#FFFFFF'), ('pressed', '#FFFFFF')])
        
        # Configure warning button style
        style.configure(
            "Warning.TButton",
            font=("Segoe UI", 9, "bold"),
            foreground="#FFFFFF",
            background="#FF8C00",
            borderwidth=1,
            focuscolor='none'
        )
        
        style.map("Warning.TButton",
            background=[('active', '#E67C00'), ('pressed', '#CC6F00')],
            foreground=[('active', '#FFFFFF'), ('pressed', '#FFFFFF')])
        
        # Configure secondary button style
        style.configure(
            "Secondary.TButton",
            font=("Segoe UI", 9),
            foreground="#323130",
            background="#F3F2F1",
            borderwidth=1,
            focuscolor='none'
        )
        
        style.map("Secondary.TButton",
            background=[('active', '#EDEBE9'), ('pressed', '#E1DFDD')],
            foreground=[('active', '#323130'), ('pressed', '#323130')])
        
        # Configure icon button style
        style.configure(
            "Icon.TButton",
            font=("Segoe UI", 8),
            foreground="#605E5C",
            background="#FAFAFA",
            borderwidth=1,
            relief="solid",
            focuscolor='none'
        )
        
        style.map("Icon.TButton",
            background=[('active', '#F3F2F1'), ('pressed', '#EDEBE9')],
            foreground=[('active', '#323130'), ('pressed', '#323130')])
        
        style.configure(
        "Card.TLabelframe",
        background="#FFFFFF",
        borderwidth=1,
        relief="solid",
        bordercolor="#E1DFDD"
        )
    
        style.configure(
            "Card.TLabelframe.Label",
            background="#FFFFFF",
            foreground="#323130",
            font=("Segoe UI", 9, "bold")
        )
        
        style.configure(
            "Nested.TLabelframe",
            background="#F8F8F8",
            borderwidth=1,
            relief="solid",
            bordercolor="#E1DFDD"
        )
        
        style.configure(
            "Nested.TLabelframe.Label",
            background="#F8F8F8",
            foreground="#605E5C",
            font=("Segoe UI", 8, "bold")
        )
        
        # Configure modern checkbutton and radiobutton styles
        style.configure(
            "Modern.TCheckbutton",
            background="#FFFFFF",
            foreground="#323130",
            font=("Segoe UI", 9),
            focuscolor='none'
        )
        
        style.configure(
            "Modern.TRadiobutton",
            background="#FFFFFF",
            foreground="#323130",
            font=("Segoe UI", 9),
            focuscolor='none'
        )

    def _create_status_section(self, parent, custom_mode=False):
        # Status container
        status_container = ttk.Frame(parent)
        status_container.pack(fill=tk.X, pady=(20, 0))
        
        # Status header
        status_header = ttk.Frame(status_container)
        status_header.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(
            status_header,
            text="📊 Status & Progress",
            font=("Segoe UI", 11, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT)
        
        # Status content
        status_content = ttk.LabelFrame(
            status_container,
            text="",
            padding=15,
            style="Card.TLabelframe"
        )
        status_content.pack(fill=tk.X)
        
        # Status row
        status_row = ttk.Frame(status_content)
        status_row.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(
            status_row,
            text="Status:",
            font=("Segoe UI", 9, "bold"),
            foreground="#323130"
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        if custom_mode:
            status_label = ttk.Label(
                status_row,
                textvariable=self.custom_status_var,
                font=("Segoe UI", 9),
                foreground="#0078D4"
            )
        else:
            status_label = ttk.Label(
                status_row,
                textvariable=self.status_var,
                font=("Segoe UI", 9),
                foreground="#0078D4"
            )
        status_label.pack(side=tk.LEFT)
        
        # Progress bar
        if custom_mode:
            self.custom_progress_bar = ttk.Progressbar(
                status_content,
                orient=tk.HORIZONTAL,
                mode='determinate',
                variable=self.custom_progress_var,
                style="Modern.Horizontal.TProgressbar"
            )
            self.custom_progress_bar.pack(fill=tk.X, pady=(5, 0))
        else:
            self.progress_bar = ttk.Progressbar(
                status_content,
                orient=tk.HORIZONTAL,
                mode='determinate',
                variable=self.progress_var,
                style="Modern.Horizontal.TProgressbar"
            )
            self.progress_bar.pack(fill=tk.X, pady=(5, 0))

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelComparisonApp(root)
    root.mainloop()

