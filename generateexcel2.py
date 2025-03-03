import openpyxl
import random
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.drawing.image import Image
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter
import os
from faker import Faker

# Initialize faker to generate realistic data
fake = Faker()

def generate_test_id(index):
    """Generate test IDs with proper formatting"""
    return f"T{index:04d}"

def create_test_excel_files():
    # Create directory if it doesn't exist
    output_dir = "test_files"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Define our file paths
    old_file_path = os.path.join(output_dir, "old_file.xlsx")
    new_file_path = os.path.join(output_dir, "new_file.xlsx")
    
    # Create the old file
    old_wb = openpyxl.Workbook()
    
    # Create sheets with different purposes
    old_sheet1 = old_wb.active
    old_sheet1.title = "Test Data"
    old_sheet2 = old_wb.create_sheet("Config")
    old_sheet3 = old_wb.create_sheet("Summary Dashboard")
    old_sheet4 = old_wb.create_sheet("Test Details")
    old_sheet5 = old_wb.create_sheet("Reference Data")
    
    # Add title and logo placeholder in first rows
    old_sheet1.merge_cells('A1:G3')
    title_cell = old_sheet1['A1']
    title_cell.value = "TEST MANAGEMENT DASHBOARD"
    title_cell.font = Font(size=18, bold=True, color="0000FF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="E0E0FF", end_color="E0E0FF", fill_type="solid")
    
    # Set up column widths for better readability
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        old_sheet1.column_dimensions[column_letter].width = 15
    old_sheet1.column_dimensions['G'].width = 40  # Notes column wider
    
    # ===== Set up the first sheet - Test Data =====
    # Define headers
    headers = {
        "A4": "Test ID", 
        "B4": "Status",
        "C4": "Team2",
        "D4": "App Name",
        "E4": "Category of Testing",
        "F4": "Result",
        "G4": "Notes",
        "H4": "Last Run",
        "I4": "Priority",
        "J4": "Execution Time (min)",
        "K4": "Cost ($)",
        "L4": "Formula"
    }
    
    # Apply headers and formatting
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for cell_addr, value in headers.items():
        cell = old_sheet1[cell_addr]
        cell.value = value
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = thin_border
    
    # Generate a lot of sample data (200 rows)
    status_options = ["Completed", "In Progress", "Pending", "Blocked", "Deferred"]
    team_options = ["Team Alpha", "Team Beta", "Team Gamma", "Team Delta", "Team Epsilon", "Team Omega"]
    app_options = ["Inventory App", "CRM Portal", "Reporting Tool", "Mobile App", "Admin Panel", 
                   "Payment Gateway", "Analytics Dashboard", "Customer Portal", "API Gateway", "Data Warehouse"]
    category_options = ["Performance", "Security", "UI", "Integration", "Functional", "Regression", 
                       "Stress", "Load", "Compatibility", "Localization"]
    result_options = ["Pass", "Fail", "N/A", "Partial"]
    priority_options = ["Critical", "High", "Medium", "Low"]
    
    # Generate data
    data = []
    now = datetime.datetime.now()
    
    # Generate 200 rows of data
    for i in range(1, 201):
        test_id = generate_test_id(i)
        status = random.choice(status_options)
        team = random.choice(team_options)
        app = random.choice(app_options)
        category = random.choice(category_options)
        result = random.choice(result_options)
        notes = fake.sentence()
        last_run = (now - datetime.timedelta(days=random.randint(0, 60))).strftime("%Y-%m-%d")
        priority = random.choice(priority_options)
        execution_time = random.randint(5, 180)
        cost = round(execution_time * random.uniform(1.5, 3.2), 2)
        
        # Formula will be added separately
        data.append([test_id, status, team, app, category, result, notes, 
                    last_run, priority, execution_time, cost])
    
    # Insert data starting at row 5
    for row_idx, row_data in enumerate(data, start=5):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = old_sheet1.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Add formatting based on content
            if col_idx == 2:  # Status column
                if cell_value == "Completed":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell_value == "In Progress":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell_value == "Blocked":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            if col_idx == 6:  # Result column
                if cell_value == "Pass":
                    cell.font = Font(color="006100")
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell_value == "Fail":
                    cell.font = Font(color="9C0006")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            if col_idx == 9:  # Priority column
                if cell_value == "Critical":
                    cell.font = Font(color="9C0006", bold=True)
                elif cell_value == "High":
                    cell.font = Font(color="9C5700")
            
            # Add borders to all cells
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Add formulas for the Formula column (column L)
    for row in range(5, 5 + len(data)):
        # Formula calculates ROI: =IF(K{row}>0, (F{row}="Pass")*100/K{row}, 0)
        formula = f'=IF(K{row}>0, IF(F{row}="Pass", 100/K{row}, -50/K{row}), 0)'
        cell = old_sheet1.cell(row=row, column=12)
        cell.value = formula
        cell.number_format = '0.00'
        cell.border = thin_border
    
    # Add conditional formatting
    # Color scale for execution time
    color_scale = ColorScaleRule(
        start_type='min', start_color='90EE90',  # Light green
        mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow
        end_type='max', end_color='F8696B'  # Red
    )
    old_sheet1.conditional_formatting.add(f'J5:J{4+len(data)}', color_scale)
    
    # Highlight expensive tests
    expensive_rule = CellIsRule(
        operator='greaterThan', formula=['80'], 
        stopIfTrue=True, fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    )
    old_sheet1.conditional_formatting.add(f'K5:K{4+len(data)}', expensive_rule)
    
    # Add a summary row at the bottom with formulas
    summary_row = 5 + len(data) + 1
    old_sheet1.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    old_sheet1.merge_cells(f'A{summary_row}:B{summary_row}')
    
    # Count tests by status
    old_sheet1.cell(row=summary_row, column=3, value=f'=COUNTIF(B5:B{4+len(data)}, "Completed")')
    old_sheet1.cell(row=summary_row, column=4, value=f'=COUNTIF(B5:B{4+len(data)}, "In Progress")')
    old_sheet1.cell(row=summary_row, column=5, value=f'=COUNTIF(B5:B{4+len(data)}, "Pending")')
    old_sheet1.cell(row=summary_row, column=6, value=f'=COUNTIF(F5:F{4+len(data)}, "Pass")&"/"&COUNTIF(F5:F{4+len(data)}, "Fail")')
    
    # Total cost
    old_sheet1.cell(row=summary_row, column=11, value=f'=SUM(K5:K{4+len(data)})').number_format = '$#,##0.00'
    
    # Average formula result
    old_sheet1.cell(row=summary_row, column=12, value=f'=AVERAGE(L5:L{4+len(data)})').number_format = '0.00'
    
    # ===== Set up the Dashboard sheet =====
    dashboard = old_sheet3
    dashboard.merge_cells('A1:H1')
    dashboard['A1'] = "TEST EXECUTION DASHBOARD"
    dashboard['A1'].font = Font(size=16, bold=True)
    dashboard['A1'].alignment = Alignment(horizontal='center')
    dashboard['A1'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Add chart titles
    dashboard['A3'] = "Test Results by Status"
    dashboard['A3'].font = Font(bold=True)
    dashboard['E3'] = "Results by Category"
    dashboard['E3'].font = Font(bold=True)
    dashboard['A15'] = "Results by Team"
    dashboard['A15'].font = Font(bold=True)
    dashboard['E15'] = "Cost vs. Time Distribution"
    dashboard['E15'].font = Font(bold=True)
    
    # Add charts - we'll create these after saving to avoid complexity in this code
    # We'll reference the main data in calculations
    
    # ===== Set up the Config sheet =====
    old_sheet2["A1"] = "TEST CONFIGURATION"
    old_sheet2["A1"].font = Font(size=14, bold=True)
    old_sheet2.merge_cells('A1:D1')
    old_sheet2["A1"].alignment = Alignment(horizontal='center')
    old_sheet2["A1"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    old_sheet2["A4"] = "Setting"
    old_sheet2["B4"] = "Value"
    old_sheet2["C4"] = "Description"
    old_sheet2["D4"] = "Last Modified"
    
    old_sheet2["A4"].font = Font(bold=True)
    old_sheet2["B4"].font = Font(bold=True)
    old_sheet2["C4"].font = Font(bold=True)
    old_sheet2["D4"].font = Font(bold=True)
    
    config_data = [
        ["Environment", "Production", "Main production environment", "2024-02-15"],
        ["Version", "1.2.3", "Current system version", "2024-02-10"],
        ["Debug Mode", "No", "Enable detailed logging", "2024-01-20"],
        ["Test Timeout", "5000", "Milliseconds before test fails", "2024-01-15"],
        ["Retry Count", "3", "Number of retries for flaky tests", "2024-02-05"],
        ["Test Path", "C:/TestData", "Path to test files", "2024-02-01"],
        ["Notification Email", "test-alerts@example.com", "Email for alert notifications", "2023-12-10"],
        ["CI Integration", "Yes", "Integrated with CI pipeline", "2024-01-30"],
        ["Team Lead", "John Smith", "Contact person for test framework", "2024-02-12"]
    ]
    
    for row_idx, row_data in enumerate(config_data, start=5):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = old_sheet2.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
    
    # Set column widths in config sheet
    old_sheet2.column_dimensions['A'].width = 18
    old_sheet2.column_dimensions['B'].width = 25
    old_sheet2.column_dimensions['C'].width = 35
    old_sheet2.column_dimensions['D'].width = 15
    
    # ===== Set up Reference Data sheet =====
    ref_sheet = old_sheet5
    ref_sheet.merge_cells('A1:D1')
    ref_sheet['A1'] = "REFERENCE DATA"
    ref_sheet['A1'].font = Font(size=14, bold=True)
    ref_sheet['A1'].alignment = Alignment(horizontal='center')
    ref_sheet['A1'].fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    
    # Add team information table
    ref_sheet['A3'] = "Team Information"
    ref_sheet['A3'].font = Font(bold=True, size=12)
    
    team_headers = ["Team Name", "Lead", "Members", "Focus Area"]
    for i, header in enumerate(team_headers):
        cell = ref_sheet.cell(row=4, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.border = thin_border
    
    team_data = [
        ["Team Alpha", "Sarah Johnson", 8, "Backend Services"],
        ["Team Beta", "Michael Chen", 6, "User Interface"],
        ["Team Gamma", "Priya Patel", 7, "Mobile Development"],
        ["Team Delta", "James Wilson", 5, "Database Systems"],
        ["Team Epsilon", "Emma Rodriguez", 9, "Security Testing"],
        ["Team Omega", "David Kim", 4, "API Integration"]
    ]
    
    for row_idx, row_data in enumerate(team_data, start=5):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ref_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
    
    # Add application table
    ref_sheet['A12'] = "Application Information"
    ref_sheet['A12'].font = Font(bold=True, size=12)
    
    app_headers = ["App Name", "Version", "Owner", "Dependencies", "Risk Level"]
    for i, header in enumerate(app_headers):
        cell = ref_sheet.cell(row=13, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.border = thin_border
    
    app_data = [
        ["Inventory App", "3.2.1", "Operations", "Database, API Gateway", "Medium"],
        ["CRM Portal", "2.0.4", "Sales", "Auth Service, Database", "High"],
        ["Reporting Tool", "1.5.0", "Analytics", "Data Warehouse, BI Engine", "Low"],
        ["Mobile App", "4.1.3", "Customer Engagement", "API Gateway, Push Service", "Medium"],
        ["Admin Panel", "2.2.0", "IT", "Auth Service, Config Service", "High"],
        ["Payment Gateway", "3.0.2", "Finance", "Banking API, Encryption Service", "Critical"],
        ["Analytics Dashboard", "1.1.3", "Analytics", "Data Lake, Visualization Engine", "Medium"],
        ["Customer Portal", "2.5.1", "Customer Success", "Auth Service, CRM Integration", "High"],
        ["API Gateway", "4.0.0", "Platform", "Service Registry, Load Balancer", "Critical"],
        ["Data Warehouse", "3.1.0", "Data", "ETL Pipeline, Storage Service", "High"]
    ]
    
    for row_idx, row_data in enumerate(app_data, start=14):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ref_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            if col_idx == 5:  # Risk level column
                if cell_value == "Critical":
                    cell.font = Font(color="9C0006")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif cell_value == "High":
                    cell.font = Font(color="9C5700")
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Set column widths
    for sheet in [ref_sheet, old_sheet4]:
        for col, width in zip('ABCDE', [20, 15, 15, 30, 15]):
            sheet.column_dimensions[col].width = width
    
    # ===== Save the old file =====
    old_wb.save(old_file_path)
    
    # ===== Create the new file (with updated values) =====
    new_wb = openpyxl.load_workbook(old_file_path)
    new_sheet1 = new_wb["Test Data"]
    
    # Update some values to show differences (more changes for larger file)
    updates = {}
    
    # Update ~20% of the rows with different statuses or results
    for i in range(5, 5 + len(data)):
        if random.random() < 0.2:  # 20% chance of change
            change_type = random.choice(['status', 'result', 'notes', 'time', 'multiple'])
            
            if change_type == 'status':
                updates[f'B{i}'] = random.choice(status_options)
            elif change_type == 'result':
                old_result = new_sheet1[f'F{i}'].value
                new_result = "Pass" if old_result != "Pass" else "Fail"
                updates[f'F{i}'] = new_result
            elif change_type == 'notes':
                updates[f'G{i}'] = fake.sentence()
            elif change_type == 'time':
                old_time = new_sheet1[f'J{i}'].value
                new_time = old_time + random.randint(-10, 20)
                if new_time < 5:
                    new_time = 5
                updates[f'J{i}'] = new_time
                # Update cost too based on the new time
                cost = round(new_time * random.uniform(1.5, 3.2), 2)
                updates[f'K{i}'] = cost
            elif change_type == 'multiple':
                # Change multiple fields for the same row
                updates[f'B{i}'] = random.choice(status_options)
                updates[f'F{i}'] = random.choice(result_options)
                updates[f'G{i}'] = fake.sentence()
    
    # Apply all the updates
    for cell_ref, value in updates.items():
        new_sheet1[cell_ref] = value
    
    # Add some completely new rows to the new file
    max_row = new_sheet1.max_row
    start_new_rows = max_row + 1
    
    # Add 10 new rows
    for i in range(201, 211):
        test_id = generate_test_id(i)
        status = random.choice(status_options)
        team = random.choice(team_options)
        app = random.choice(app_options)
        category = random.choice(category_options)
        result = random.choice(result_options)
        notes = fake.sentence()
        last_run = (now - datetime.timedelta(days=random.randint(0, 10))).strftime("%Y-%m-%d")
        priority = random.choice(priority_options)
        execution_time = random.randint(5, 180)
        cost = round(execution_time * random.uniform(1.5, 3.2), 2)
        
        row_data = [test_id, status, team, app, category, result, notes, 
                    last_run, priority, execution_time, cost]
        
        row_idx = start_new_rows + (i - 201)
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = new_sheet1.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply same formatting as earlier rows
            if col_idx == 2:  # Status column
                if cell_value == "Completed":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell_value == "In Progress":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell_value == "Blocked":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            if col_idx == 6:  # Result column
                if cell_value == "Pass":
                    cell.font = Font(color="006100")
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell_value == "Fail":
                    cell.font = Font(color="9C0006")
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Add borders to all cells
            cell.border = thin_border
        
        # Add formula for column L
        formula = f'=IF(K{row_idx}>0, IF(F{row_idx}="Pass", 100/K{row_idx}, -50/K{row_idx}), 0)'
        cell = new_sheet1.cell(row=row_idx, column=12)
        cell.value = formula
        cell.number_format = '0.00'
        cell.border = thin_border
    
    # Update summary row to include new rows
    new_summary_row = start_new_rows + 10
    new_sheet1.cell(row=new_summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    new_sheet1.merge_cells(f'A{new_summary_row}:B{new_summary_row}')
    
    # Update summary formulas to include new rows
    new_sheet1.cell(row=new_summary_row, column=3, value=f'=COUNTIF(B5:B{new_summary_row-1}, "Completed")')
    new_sheet1.cell(row=new_summary_row, column=4, value=f'=COUNTIF(B5:B{new_summary_row-1}, "In Progress")')
    new_sheet1.cell(row=new_summary_row, column=5, value=f'=COUNTIF(B5:B{new_summary_row-1}, "Pending")')
    new_sheet1.cell(row=new_summary_row, column=6, value=f'=COUNTIF(F5:F{new_summary_row-1}, "Pass")&"/"&COUNTIF(F5:F{new_summary_row-1}, "Fail")')
    
    # Total cost
    new_sheet1.cell(row=new_summary_row, column=11, value=f'=SUM(K5:K{new_summary_row-1})').number_format = '$#,##0.00'
    
    # Average formula result
    new_sheet1.cell(row=new_summary_row, column=12, value=f'=AVERAGE(L5:L{new_summary_row-1})').number_format = '0.00'
    
    # Save the new file
    new_wb.save(new_file_path)
    
    print(f"Created enhanced test files in '{output_dir}' folder:")
    print(f"- Old file: {old_file_path}")
    print(f"- New file: {new_file_path}")

if __name__ == "__main__":
    create_test_excel_files()